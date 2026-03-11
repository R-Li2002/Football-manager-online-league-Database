from pathlib import Path

import pandas as pd
from openpyxl import Workbook


TEAM_HEADER_ROW = [
    None,
    None,
    None,
    "球队名",
    "主教",
    "级别",
    "额外工资",
    "税后",
    "备注",
]


DEFAULT_TEAM_ROWS = [
    {"info_key": "届数", "info_value": 85, "name": "Alpha FC", "manager": "Coach A", "level": "超级", "extra_wage": 0.2, "after_tax": 0, "notes": ""},
    {"info_key": "成长年龄上限", "info_value": 24, "name": "Beta FC", "manager": "Coach B", "level": "甲级", "extra_wage": 0.0, "after_tax": 0, "notes": "+0.1M"},
]


DEFAULT_PLAYER_ROWS = [
    {"uid": 1, "name": "张三", "age": 20, "initial_ca": 120, "ca": 125, "pa": 150, "position": "MC", "nationality": "CN", "club_name": "Alpha FC"},
    {"uid": 2, "name": "李四", "age": 21, "initial_ca": 118, "ca": 121, "pa": 148, "position": "GK", "nationality": "CN", "club_name": "Beta FC"},
]


def default_mapping_rows(player_rows):
    return [
        {
            "uid": player["uid"],
            "name": player["name"],
            "age": player["age"],
            "ca": player["ca"],
            "pa": player["pa"],
            "nationality": player["nationality"],
            "team_name": player.get("league_team_name", player["club_name"]),
            "position": player["position"],
        }
        for player in player_rows
    ]


def write_workbook(
    path: Path,
    *,
    include_position_column: bool = True,
    team_rows=None,
    player_rows=None,
    mapping_rows=None,
) -> None:
    team_rows = team_rows or DEFAULT_TEAM_ROWS
    player_rows = player_rows or DEFAULT_PLAYER_ROWS
    mapping_rows = mapping_rows or default_mapping_rows(player_rows)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "信息总览"

    overview.append(["说明：测试导入", None, None, None, None, None, None, None, None])
    overview.append(TEAM_HEADER_ROW)
    for team in team_rows:
        overview.append(
            [
                team["info_key"],
                team["info_value"],
                None,
                team["name"],
                team["manager"],
                team["level"],
                team["extra_wage"],
                team["after_tax"],
                team["notes"],
            ]
        )
    overview.append(["8M名额", 1.5])
    overview.append(["7M名额", 1.3])
    overview.append(["可成长名额", 1.1])
    overview.append(["非名PA6M", 0.9])
    overview.append(["非名身价1M", 1.0])
    overview.append(["非名其他", 0.7])
    overview.append(["注：名额GK系数为", 1])

    players_sheet = workbook.create_sheet("联赛名单")
    player_headers = [
        "编号",
        "姓名",
        "年龄",
        "初始CA",
        "当前CA",
        "PA",
        "位置",
        "国籍",
        "俱乐部",
    ]
    include_league_team_column = any("league_team_name" in player for player in player_rows)
    if include_league_team_column:
        player_headers.append("联赛球队")
    if not include_position_column:
        player_headers.remove("位置")
    players_sheet.append(player_headers)
    for player in player_rows:
        row = [
            player["uid"],
            player["name"],
            player["age"],
            player["initial_ca"],
            player["ca"],
            player["pa"],
            player["position"],
            player["nationality"],
            player["club_name"],
        ]
        if include_league_team_column:
            row.append(player.get("league_team_name", ""))
        if include_position_column:
            players_sheet.append(row)
        else:
            if include_league_team_column:
                players_sheet.append([row[0], row[1], row[2], row[3], row[4], row[5], row[7], row[8], row[9]])
            else:
                players_sheet.append([row[0], row[1], row[2], row[3], row[4], row[5], row[7], row[8]])

    mapping_sheet = workbook.create_sheet("球员对应球队")
    mapping_sheet.append(["UID", "球员名", "年龄", "CA", "PA", "国籍", "球队", "位置", "联赛名单有无"])
    for player in mapping_rows:
        mapping_sheet.append(
            [
                player["uid"],
                player["name"],
                player["age"],
                player["ca"],
                player["pa"],
                player["nationality"],
                player["team_name"],
                player["position"],
                "",
            ]
        )

    workbook.save(path)


def write_attributes_csv(path: Path) -> None:
    df = pd.DataFrame(
        [
            {"UID": 1, "姓名": "张三", "位置": "MC", "年龄": 20, "ca": 125, "pa": 150, "国籍": "CN", "俱乐部": "Alpha FC", "角球": 12},
            {"UID": 2, "姓名": "李四", "位置": "GK", "年龄": 21, "ca": 121, "pa": 148, "国籍": "CN", "俱乐部": "Beta FC", "角球": 5},
        ]
    )
    df.to_csv(path, index=False, encoding="utf-8-sig")
