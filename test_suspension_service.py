import unittest
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base
from models import Match, Player, PlayerSuspensionRecord, Team
from schemas_write import SiteNoteUpdateRequest, SuspensionRecordUpdateRequest
from services import export_service, site_note_service, suspension_service


class SuspensionServiceTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.db = Session()
        self.team = Team(name="Alpha", level="超级", manager="Alpha Boss")
        self.hidden_team = Team(name="Hidden", level="隐藏", manager="Hidden Boss")
        self.db.add_all([self.team, self.hidden_team])
        self.db.commit()
        self.players = [
            Player(uid=101, name="Alpha One", team_id=self.team.id, team_name=self.team.name),
            Player(uid=102, name="Alpha Two", team_id=self.team.id, team_name=self.team.name),
            Player(uid=103, name="Alpha Three", team_id=self.team.id, team_name=self.team.name),
            Player(uid=201, name="Hidden One", team_id=self.hidden_team.id, team_name=self.hidden_team.name),
        ]
        self.db.add_all(self.players)
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def _save(self, uid, **payload):
        request = SuspensionRecordUpdateRequest(player_uid=uid, **payload)
        return suspension_service.update_suspension_record(self.db, "editor", request, lambda *_args: None)

    def _add_match(self, round_no, status="scheduled", home_score=None, away_score=None):
        match = Match(
            level="超级",
            round_no=round_no,
            home_team_id=self.team.id,
            home_team_name=self.team.name,
            away_team_name=f"Opponent {round_no}",
            home_score=home_score,
            away_score=away_score,
            status=status,
        )
        self.db.add(match)
        self.db.commit()
        return match

    def _set_team_round(self, round_no):
        site_note_service.update_site_note(
            self.db,
            "editor",
            site_note_service.build_suspension_team_note_key(self.team.id),
            SiteNoteUpdateRequest(text=f"核对至第 {round_no} 轮", round_no=round_no),
            lambda *_args: None,
        )

    def test_suspensions_are_grouped_by_status(self):
        self._save(101, yellow_cards=1, notes="一黄备注")
        self._save(102, yellow_cards=2)
        self._save(103, yellow_cards=3, red_card_suspended=True, red_injury_suspended=True, notes="停赛备注")

        response = suspension_service.get_suspensions(self.db)
        alpha = next(team for team in response.teams if team.team_name == "Alpha")

        self.assertEqual([item.player_uid for item in alpha.one_yellow], [101])
        self.assertEqual([item.player_uid for item in alpha.two_yellows], [102])
        self.assertEqual([item.player_uid for item in alpha.suspended], [103])
        self.assertIn("Alpha Three: 停赛备注", alpha.notes)

    def test_suspensions_can_be_limited_to_one_level(self):
        first_team = Team(name="First Team", level="甲级", manager="First Boss")
        self.db.add(first_team)
        self.db.commit()

        response = suspension_service.get_suspensions(self.db, level="甲级")

        self.assertEqual(response.levels, ["甲级"])
        self.assertEqual([team.team_name for team in response.teams], ["First Team"])

    def test_suspension_progress_ahead_suppresses_unsynced_old_fixtures(self):
        for round_no in (1, 2, 3):
            self._add_match(round_no)
        self._set_team_round(2)

        response = suspension_service.get_suspensions(self.db)
        progress = next(team for team in response.teams if team.team_name == "Alpha").progress

        self.assertEqual(progress.state, "ahead")
        self.assertEqual(progress.match_completed_round, 0)
        self.assertEqual(progress.suspension_checked_round, 2)
        self.assertEqual(progress.applies_from_round, 3)
        self.assertEqual(progress.progress_floor_round, 2)
        self.assertEqual(progress.next_match_round, 3)
        self.assertIn("赛果尚未同步", progress.detail)

    def test_suspension_progress_stale_when_recorded_results_are_ahead(self):
        for round_no in range(1, 6):
            self._add_match(round_no, status="played", home_score=round_no, away_score=0)
        self._add_match(6)
        self._set_team_round(2)

        response = suspension_service.get_suspensions(self.db)
        progress = next(team for team in response.teams if team.team_name == "Alpha").progress

        self.assertEqual(progress.state, "stale")
        self.assertEqual(progress.match_completed_round, 5)
        self.assertEqual(progress.suspension_checked_round, 2)
        self.assertEqual(progress.next_match_round, 6)
        self.assertIn("落后 3 轮", progress.detail)

    def test_result_gap_reports_continuous_and_latest_rounds_without_hiding_gap(self):
        for round_no in (1, 2, 3):
            self._add_match(round_no, status="played", home_score=round_no, away_score=0)
        gap_match = self._add_match(4)
        self._add_match(5, status="played", home_score=5, away_score=0)
        self._add_match(6)
        self._set_team_round(3)

        response = suspension_service.get_suspensions(self.db)
        progress = next(team for team in response.teams if team.team_name == "Alpha").progress

        self.assertEqual(progress.state, "gap")
        self.assertEqual(progress.match_completed_round, 5)
        self.assertEqual(progress.match_latest_recorded_round, 5)
        self.assertEqual(progress.match_continuous_completed_round, 3)
        self.assertEqual(progress.match_gap_rounds, [4])
        self.assertEqual(progress.next_match_id, gap_match.id)
        self.assertEqual(progress.next_match_round, 4)
        self.assertTrue(progress.next_match_is_gap)
        self.assertIn("连续完成至第 3 轮", progress.detail)
        self.assertIn("第 4 轮尚未确认", progress.detail)

    def test_explicit_postponement_remains_the_next_match(self):
        postponed = self._add_match(1, status="postponed")
        self._add_match(2)
        self._add_match(3)
        self._set_team_round(2)

        response = suspension_service.get_suspensions(self.db)
        progress = next(team for team in response.teams if team.team_name == "Alpha").progress

        self.assertEqual(progress.next_match_id, postponed.id)
        self.assertEqual(progress.next_match_round, 1)
        self.assertTrue(progress.next_match_is_postponed)
        self.assertIn("延期的第 1 轮", progress.detail)

    def test_empty_payload_clears_existing_record(self):
        self._save(101, yellow_cards=2)
        self._save(101, yellow_cards=0, red_card_suspended=False, red_injury_suspended=False, notes="")

        response = suspension_service.get_suspensions(self.db)
        alpha = next(team for team in response.teams if team.team_name == "Alpha")
        self.assertEqual(alpha.one_yellow, [])
        self.assertEqual(alpha.two_yellows, [])
        self.assertEqual(alpha.suspended, [])

    def test_hidden_team_player_is_rejected(self):
        with self.assertRaises(HTTPException):
            self._save(201, yellow_cards=1)

    def test_departed_player_record_is_visible_and_can_be_cleared(self):
        self._save(101, yellow_cards=2, notes="离队前记录")
        player = self.db.query(Player).filter(Player.uid == 101).one()
        self.db.delete(player)
        self.db.commit()

        response = suspension_service.get_suspensions(self.db)
        orphaned = next(team for team in response.teams if team.level == "超级" and team.is_orphaned)
        self.assertEqual([item.player_uid for item in orphaned.two_yellows], [101])
        clear_request = SuspensionRecordUpdateRequest(player_uid=101)
        self.assertEqual(suspension_service.get_suspension_request_level(self.db, clear_request), "超级")

        suspension_service.update_suspension_record(self.db, "editor", clear_request, lambda *_args: None)
        self.assertIsNone(
            self.db.query(PlayerSuspensionRecord).filter(PlayerSuspensionRecord.player_uid == 101).first()
        )

    def test_transferred_player_stale_record_moves_to_orphaned_section(self):
        self._save(102, yellow_cards=1)
        new_team = Team(name="Beta", level="甲级", manager="Beta Boss")
        self.db.add(new_team)
        self.db.flush()
        player = self.db.query(Player).filter(Player.uid == 102).one()
        player.team_id = new_team.id
        player.team_name = new_team.name
        self.db.commit()

        response = suspension_service.get_suspensions(self.db)
        alpha = next(team for team in response.teams if team.team_name == "Alpha")
        orphaned = next(team for team in response.teams if team.level == "超级" and team.is_orphaned)

        self.assertEqual(alpha.one_yellow, [])
        self.assertEqual([item.player_uid for item in orphaned.one_yellow], [102])

    def test_team_update_note_is_listed_and_exported_with_suspensions(self):
        self._save(101, yellow_cards=2, notes="球员说明")
        note_key = site_note_service.build_suspension_team_note_key(self.team.id)
        site_note_service.update_site_note(
            self.db,
            "editor",
            note_key,
            SiteNoteUpdateRequest(text="更新至第 8 轮赛后", round_no=8),
            lambda *_args: None,
        )

        self.assertEqual(site_note_service.get_suspension_note_level(self.db, note_key), "超级")
        notes = {item.key: item for item in site_note_service.list_site_notes(self.db)}
        self.assertEqual(notes[note_key].text, "更新至第 8 轮赛后")
        self.assertEqual(notes[note_key].round_no, 8)

        output, filename = export_service.build_suspensions_excel(self.db, "超级")
        self.assertTrue(filename.endswith(".xlsx"))
        workbook = load_workbook(BytesIO(output.getvalue()), data_only=True)
        self.assertEqual(workbook.sheetnames, ["球队汇总", "伤停明细"])

        summary_rows = list(workbook["球队汇总"].iter_rows(values_only=True))
        summary_headers = list(summary_rows[0])
        alpha_summary = dict(zip(summary_headers, summary_rows[1]))
        self.assertEqual(alpha_summary["球队"], "Alpha")
        self.assertEqual(alpha_summary["2张黄牌人数"], 1)
        self.assertEqual(alpha_summary["球队更新备注"], "更新至第 8 轮赛后")

        detail_rows = list(workbook["伤停明细"].iter_rows(values_only=True))
        detail_headers = list(detail_rows[0])
        player_detail = dict(zip(detail_headers, detail_rows[1]))
        self.assertEqual(player_detail["球员"], "Alpha One")
        self.assertEqual(player_detail["状态分类"], "2张黄牌")
        self.assertEqual(player_detail["球员备注"], "球员说明")
        self.assertEqual(player_detail["球队更新备注"], "更新至第 8 轮赛后")

    def test_team_note_rejects_hidden_or_missing_team(self):
        hidden_key = site_note_service.build_suspension_team_note_key(self.hidden_team.id)
        with self.assertRaises(HTTPException):
            site_note_service.update_site_note(
                self.db,
                "editor",
                hidden_key,
                SiteNoteUpdateRequest(text="不允许"),
                lambda *_args: None,
            )

    def test_note_round_accepts_boundaries_and_rejects_out_of_range(self):
        note_key = site_note_service.build_suspension_note_key("超级")
        for round_no in (0, 34):
            site_note_service.update_site_note(
                self.db,
                "editor",
                note_key,
                SiteNoteUpdateRequest(text=f"第 {round_no} 轮", round_no=round_no),
                lambda *_args: None,
            )
            note = next(item for item in site_note_service.list_site_notes(self.db) if item.key == note_key)
            self.assertEqual(note.round_no, round_no)

        for round_no in (-1, 35):
            with self.assertRaises(HTTPException):
                site_note_service.update_site_note(
                    self.db,
                    "editor",
                    note_key,
                    SiteNoteUpdateRequest(text="越界", round_no=round_no),
                    lambda *_args: None,
                )


if __name__ == "__main__":
    unittest.main()
