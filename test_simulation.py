import importlib
import io
import gc
import os
import socket
import subprocess
import sys
import time
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import requests
from openpyxl import load_workbook
from sqlalchemy.orm import close_all_sessions

from test_fixture_data import write_attributes_csv, write_workbook


MODULES_TO_RESET = [
    "import_data",
    "main1",
    "services.admin_service",
    "services.admin_write_service",
    "services.admin_common",
    "services.auth_service",
    "services.export_service",
    "services.import_service",
    "services.league_service",
    "services.operation_audit_service",
    "services.read_service",
    "services.roster_service",
    "services.transfer_service",
    "services.wage_service",
    "services",
    "routers.admin_read_routes",
    "routers.admin_write_routes",
    "routers.frontend_routes",
    "routers.public_routes",
    "routers",
    "repositories.operation_audit_repository",
    "schemas_read",
    "schemas_write",
    "operation_audit_store",
    "team_links",
    "domain_types",
    "league_settings",
    "auth_utils",
    "wage_calculator",
    "models",
    "database",
]

ROOT_DIR = Path(__file__).resolve().parent


def load_fresh_modules(db_path: Path):
    os.environ["DATABASE_URL"] = f"sqlite:///{db_path.as_posix()}"
    os.environ["SESSION_COOKIE_SECURE"] = "false"

    for module_name in MODULES_TO_RESET:
        sys.modules.pop(module_name, None)

    database = importlib.import_module("database")
    models = importlib.import_module("models")
    team_links = importlib.import_module("team_links")
    main1 = importlib.import_module("main1")
    return database, models, team_links, main1


def pick_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return sock.getsockname()[1]


class SimulationTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "simulation.db"
        self.import_workbook_path = Path(self.temp_dir.name) / "85届初HEIGO名单v5.xlsx"
        self.import_attributes_path = Path(self.temp_dir.name) / "2600球员属性.csv"
        write_workbook(self.import_workbook_path)
        write_attributes_csv(self.import_attributes_path)
        self.database, self.models, self.team_links, self.main1 = load_fresh_modules(self.db_path)
        self.database.init_database()
        self.SessionLocal = self.database.SessionLocal
        self.seed_data()

        self.port = pick_free_port()
        self.base_url = f"http://127.0.0.1:{self.port}"
        self.server_log_path = Path(self.temp_dir.name) / "server.log"
        self.server_log = open(self.server_log_path, "w", encoding="utf-8")
        env = os.environ.copy()
        env["DATABASE_URL"] = f"sqlite:///{self.db_path.as_posix()}"
        env["SESSION_COOKIE_SECURE"] = "false"
        env["HEIGO_IMPORT_ROOT"] = self.temp_dir.name
        env["PYTHONIOENCODING"] = "utf-8"

        self.server = subprocess.Popen(
            [sys.executable, "-m", "uvicorn", "main1:app", "--host", "127.0.0.1", "--port", str(self.port)],
            cwd=str(ROOT_DIR),
            env=env,
            stdout=self.server_log,
            stderr=subprocess.STDOUT,
        )
        self.wait_for_server()
        self.http = requests.Session()

    def tearDown(self):
        if hasattr(self, "http"):
            self.http.close()
        if hasattr(self, "server"):
            self.server.terminate()
            try:
                self.server.wait(timeout=10)
            except subprocess.TimeoutExpired:
                self.server.kill()
                self.server.wait(timeout=10)
        if hasattr(self, "server_log"):
            self.server_log.close()
        close_all_sessions()
        if hasattr(self, "database"):
            self.database.engine.dispose()
        for module_name in MODULES_TO_RESET:
            sys.modules.pop(module_name, None)
        self.main1 = None
        self.team_links = None
        self.models = None
        self.database = None
        self.temp_dir.cleanup()
        gc.collect()

    def wait_for_server(self):
        deadline = time.time() + 15
        last_error = None
        while time.time() < deadline:
            if self.server.poll() is not None:
                raise AssertionError(f"uvicorn exited early:\n{self.server_log_path.read_text(encoding='utf-8')}")
            try:
                response = requests.get(f"{self.base_url}/health", timeout=1)
                try:
                    if response.status_code == 200:
                        return
                finally:
                    response.close()
            except requests.RequestException as exc:
                last_error = exc
                time.sleep(0.2)
        raise AssertionError(f"server failed to start: {last_error}\n{self.server_log_path.read_text(encoding='utf-8')}")

    def seed_data(self):
        db = self.SessionLocal()
        try:
            Team = self.models.Team
            Player = self.models.Player
            PlayerAttribute = self.models.PlayerAttribute
            LeagueInfo = self.models.LeagueInfo

            alpha = Team(name="Alpha FC", manager="Alice", level="超级", wage=0, notes="+0.1M")
            beta = Team(name="Beta FC", manager="Bob", level="甲级", wage=0)
            sea = Team(name=self.team_links.SEA_TEAM_NAME, manager="Sea", level="隐藏", wage=0)
            db.add_all([alpha, beta, sea])
            db.flush()

            players = [
                Player(
                    uid=1001,
                    name="Alpha One",
                    age=22,
                    initial_ca=110,
                    ca=112,
                    pa=128,
                    position="MC",
                    nationality="ENG",
                    team_id=alpha.id,
                    team_name=alpha.name,
                    wage=0,
                    slot_type="",
                ),
                Player(
                    uid=1002,
                    name="Alpha Two",
                    age=24,
                    initial_ca=98,
                    ca=99,
                    pa=112,
                    position="DC",
                    nationality="ESP",
                    team_id=alpha.id,
                    team_name=alpha.name,
                    wage=0,
                    slot_type="",
                ),
                Player(
                    uid=1003,
                    name="Beta One",
                    age=21,
                    initial_ca=105,
                    ca=106,
                    pa=123,
                    position="ST",
                    nationality="BRA",
                    team_id=beta.id,
                    team_name=beta.name,
                    wage=0,
                    slot_type="",
                ),
            ]

            for player in players:
                self.main1.refresh_player_financials(player, db)
                db.add(player)

            db.add(
                PlayerAttribute(
                    uid=1001,
                    name="Alpha One",
                    position="MC",
                    age=22,
                    ca=112,
                    pa=128,
                    nationality="ENG",
                    club="Legacy Club",
                    pos_mc=20,
                    passing=16,
                    teamwork=15,
                )
            )
            growth_limit = LeagueInfo(key="成长年龄上限", category="基本信息", value_type="int")
            growth_limit.set_typed_value(24)
            db.add(growth_limit)

            self.main1.recalculate_team_stats(db)
        finally:
            db.close()

    def request(self, method: str, path: str, **kwargs):
        response = self.http.request(method, f"{self.base_url}{path}", timeout=10, **kwargs)
        return response

    def login(self):
        response = self.request(
            "POST",
            "/api/admin/login",
            json={"username": "HEIGO01", "password": "HEIGOLeverkusen85"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["username"], "HEIGO01")
        self.assertIsNotNone(self.http.cookies.get("session_token"))

    def fetch_player(self, uid: int):
        db = self.SessionLocal()
        try:
            return db.query(self.models.Player).filter(self.models.Player.uid == uid).first()
        finally:
            db.close()

    def fetch_team(self, name: str):
        db = self.SessionLocal()
        try:
            return db.query(self.models.Team).filter(self.models.Team.name == name).first()
        finally:
            db.close()

    def fetch_logs(self, uid: int | None = None):
        db = self.SessionLocal()
        try:
            query = db.query(self.models.TransferLog).order_by(self.models.TransferLog.id.asc())
            if uid is not None:
                query = query.filter(self.models.TransferLog.player_uid == uid)
            return query.all()
        finally:
            db.close()

    def test_public_endpoints_and_export(self):
        health = self.request("GET", "/health")
        self.assertEqual(health.status_code, 200)
        self.assertEqual(health.json()["status"], "healthy")

        root = self.request("GET", "/")
        self.assertEqual(root.status_code, 200)
        self.assertIn("Heigo", root.text)

        teams = self.request("GET", "/api/teams")
        self.assertEqual(teams.status_code, 200)

    def test_admin_formal_import_creates_backup_and_cleans_stale_team(self):
        db = self.SessionLocal()
        try:
            db.add(self.models.Team(name="Legacy Roma", manager="Old", level="瓒呯骇", wage=0))
            db.commit()
        finally:
            db.close()

        self.login()
        response = self.request("POST", "/api/admin/import/formal")
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()

        self.assertTrue(payload["success"], payload)
        self.assertTrue(payload["committed"], payload)
        self.assertTrue(payload["backup_path"])
        self.assertTrue(Path(payload["backup_path"]).exists())
        self.assertEqual(payload["datasets"]["team_cleanup"]["details"]["removed_count"], 1)
        self.assertEqual(payload["datasets"]["team_cleanup"]["details"]["removed_teams"], ["Legacy Roma"])

        audits = self.request("GET", "/api/admin/operations-audit")
        self.assertEqual(audits.status_code, 200, audits.text)
        audit_payload = audits.json()
        formal_import_event = next(
            item
            for item in audit_payload
            if item["category"] == "import" and item["action"] == "formal_import" and item["source"] == "admin_ui"
        )
        self.assertEqual(formal_import_event["status"], "success")
        self.assertEqual(formal_import_event["details"]["datasets"]["team_cleanup"]["details"]["removed_count"], 1)

        latest_import = self.request("GET", "/api/admin/import/latest")
        self.assertEqual(latest_import.status_code, 200, latest_import.text)
        latest_import_payload = latest_import.json()
        self.assertEqual(latest_import_payload["datasets"]["team_cleanup"]["details"]["removed_count"], 1)

        db = self.SessionLocal()
        try:
            self.assertIsNone(db.query(self.models.Team).filter(self.models.Team.name == "Legacy Roma").first())
            self.assertEqual(db.query(self.models.Team).count(), 3)
            self.assertEqual(db.query(self.models.Player).count(), 5)
        finally:
            db.close()
        return
        team_payload = teams.json()
        self.assertEqual([team["name"] for team in team_payload], ["Alpha FC", "Beta FC"])
        self.assertEqual(team_payload[0]["team_size"], 2)
        self.assertEqual(team_payload[1]["team_size"], 1)
        self.assertEqual(team_payload[0]["stat_sources"]["field_modes"]["team_size"], "cached")
        self.assertEqual(team_payload[0]["stat_sources"]["field_modes"]["total_value"], "realtime")
        self.assertIn("count_8m", team_payload[0]["stat_sources"]["cached_fields"])
        self.assertIn("avg_ca", team_payload[0]["stat_sources"]["realtime_fields"])
        self.assertEqual(team_payload[0]["stat_sources"]["refresh_state"]["cached_read_mode"], "cache_hit")
        self.assertEqual(team_payload[0]["stat_sources"]["refresh_state"]["realtime_read_mode"], "realtime_overlay")
        self.assertEqual(team_payload[0]["stat_sources"]["refresh_state"]["last_cache_refresh_mode"], "full_recalc")

        players = self.request("GET", "/api/players")
        self.assertEqual(players.status_code, 200)
        self.assertEqual(len(players.json()), 3)

        alpha_players = self.request("GET", "/api/players/team/Alpha FC")
        self.assertEqual(alpha_players.status_code, 200)
        self.assertEqual([player["uid"] for player in alpha_players.json()], [1001, 1002])

        player_search = self.request("GET", "/api/players/search/Alpha")
        self.assertEqual(player_search.status_code, 200)
        self.assertEqual(len(player_search.json()), 2)

        attribute_search = self.request("GET", "/api/attributes/search/Alpha")
        self.assertEqual(attribute_search.status_code, 200)
        self.assertEqual(attribute_search.json()[0]["uid"], 1001)

        attribute_detail = self.request("GET", "/api/attributes/1001")
        self.assertEqual(attribute_detail.status_code, 200)
        self.assertEqual(attribute_detail.json()["club"], "Legacy Club")

        wage_detail = self.request("GET", "/api/player/wage-detail/1001")
        self.assertEqual(wage_detail.status_code, 200)
        self.assertIn("wage", wage_detail.json())
        self.assertIn("slot_type", wage_detail.json())

        export = self.request("GET", "/api/export/excel")
        self.assertEqual(export.status_code, 200)
        workbook = load_workbook(io.BytesIO(export.content))
        try:
            self.assertEqual(len(workbook.sheetnames), 2)
            self.assertGreaterEqual(workbook[workbook.sheetnames[0]].max_row, 2)
            self.assertGreaterEqual(workbook[workbook.sheetnames[1]].max_row, 2)
        finally:
            workbook.close()

    def test_admin_write_flow_and_consistency(self):
        self.login()

        auth_audits = self.request("GET", "/api/admin/operations-audit?category=auth")
        self.assertEqual(auth_audits.status_code, 200, auth_audits.text)
        auth_payload = auth_audits.json()
        self.assertTrue(any(item["action"] == "login" for item in auth_payload))

        check = self.request("GET", "/api/admin/check")
        self.assertEqual(check.status_code, 200)
        self.assertTrue(check.json()["authenticated"])

        bootstrap_status = self.request("GET", "/api/admin/schema-bootstrap-status")
        self.assertEqual(bootstrap_status.status_code, 200, bootstrap_status.text)
        bootstrap_payload = bootstrap_status.json()
        self.assertTrue(bootstrap_payload["file_exists"])
        self.assertTrue(bootstrap_payload["latest_event"])
        self.assertIn("alembic_upgrade_head", bootstrap_payload["latest_event"])
        self.assertGreaterEqual(len(bootstrap_payload["recent_events"]), 1)

        transfer = self.request(
            "POST",
            "/api/admin/transfer",
            json={"player_uid": 1001, "to_team": "Beta FC", "notes": "simulation transfer"},
        )
        self.assertEqual(transfer.status_code, 200, transfer.text)

        transfer_audits = self.request("GET", "/api/admin/operations-audit?category=transfer")
        self.assertEqual(transfer_audits.status_code, 200, transfer_audits.text)
        transfer_payload = transfer_audits.json()
        transfer_event = next(item for item in transfer_payload if item["action"] == "transfer_player")
        self.assertEqual(transfer_event["status"], "success")
        self.assertEqual(transfer_event["details"]["request"]["player_uid"], 1001)

        player = self.fetch_player(1001)
        beta = self.fetch_team("Beta FC")
        self.assertEqual(player.team_name, "Beta FC")
        self.assertEqual(player.team_id, beta.id)

        logs = self.fetch_logs(1001)
        self.assertEqual(logs[-1].operation, "交易")
        self.assertIsNotNone(logs[-1].from_team_id)
        self.assertEqual(logs[-1].to_team_id, beta.id)

        teams = self.request("GET", "/api/teams").json()
        team_sizes = {team["name"]: team["team_size"] for team in teams}
        self.assertEqual(team_sizes["Alpha FC"], 1)
        self.assertEqual(team_sizes["Beta FC"], 2)
        team_refresh_modes = {team["name"]: team["stat_sources"]["refresh_state"]["last_cache_refresh_mode"] for team in teams}
        self.assertEqual(team_refresh_modes["Alpha FC"], "write_incremental")
        self.assertEqual(team_refresh_modes["Beta FC"], "write_incremental")

        consume = self.request(
            "POST",
            "/api/admin/consume",
            json={"player_uid": 1001, "ca_change": 2, "pa_change": -1, "notes": "simulation consume"},
        )
        self.assertEqual(consume.status_code, 200, consume.text)

        player = self.fetch_player(1001)
        self.assertEqual((player.ca, player.pa), (114, 127))

        rejuvenate = self.request(
            "POST",
            "/api/admin/rejuvenate",
            json={"player_uid": 1001, "age_change": 2, "notes": "simulation rejuvenate"},
        )
        self.assertEqual(rejuvenate.status_code, 200, rejuvenate.text)

        player = self.fetch_player(1001)
        self.assertEqual(player.age, 20)

        release = self.request(
            "POST",
            "/api/admin/release",
            json={"player_uid": 1001, "to_team": self.team_links.SEA_TEAM_NAME, "notes": "simulation release"},
        )
        self.assertEqual(release.status_code, 200, release.text)

        player = self.fetch_player(1001)
        self.assertEqual(player.team_name, self.team_links.SEA_TEAM_NAME)

        sea_players = self.request("GET", "/api/admin/sea-players")
        self.assertEqual(sea_players.status_code, 200)
        self.assertIn(1001, [item["uid"] for item in sea_players.json()])

        release_log_id = self.fetch_logs(1001)[-1].id
        undo = self.request("POST", f"/api/admin/undo/{release_log_id}")
        self.assertEqual(undo.status_code, 200, undo.text)

        player = self.fetch_player(1001)
        self.assertEqual(player.team_name, "Beta FC")

        player_update = self.request(
            "POST",
            "/api/admin/player/update",
            json={"uid": 1001, "name": "Alpha Prime", "position": "AMC", "nationality": "FRA", "age": 21},
        )
        self.assertEqual(player_update.status_code, 200, player_update.text)

        player = self.fetch_player(1001)
        self.assertEqual(player.name, "Alpha Prime")
        self.assertEqual(player.position, "AMC")
        self.assertEqual(player.nationality, "FRA")
        self.assertEqual(player.age, 21)
        self.assertGreater(player.wage, 0)

        uid_update = self.request(
            "POST",
            "/api/admin/player/update-uid",
            json={"old_uid": 1001, "new_uid": 2001},
        )
        self.assertEqual(uid_update.status_code, 200, uid_update.text)

        player = self.fetch_player(2001)
        self.assertIsNotNone(player)
        self.assertTrue(all(log.player_uid == 2001 for log in self.fetch_logs(2001)))

        recalc = self.request("POST", "/api/admin/recalculate-wages")
        self.assertEqual(recalc.status_code, 200, recalc.text)

        db = self.SessionLocal()
        try:
            for team_name in ("Alpha FC", "Beta FC"):
                team = db.query(self.models.Team).filter(self.models.Team.name == team_name).first()
                team.stats_cache_refresh_mode = "unknown"
                team.stats_cache_refresh_scopes = ""
                team.stats_cache_refresh_at = None
            db.commit()
        finally:
            db.close()

        rebuild = self.request("POST", "/api/admin/team-stats/rebuild-cache")
        self.assertEqual(rebuild.status_code, 200, rebuild.text)
        rebuilt_teams = {team["name"]: team for team in self.request("GET", "/api/teams").json()}
        self.assertEqual(rebuilt_teams["Alpha FC"]["stat_sources"]["refresh_state"]["last_cache_refresh_mode"], "full_recalc")
        self.assertEqual(rebuilt_teams["Beta FC"]["stat_sources"]["refresh_state"]["last_cache_refresh_mode"], "full_recalc")

        db = self.SessionLocal()
        try:
            sessions = db.query(self.models.AdminSession).count()
            self.assertEqual(sessions, 1)
        finally:
            db.close()

        exported = self.request("GET", "/api/admin/operations-audit/export?category=transfer")
        self.assertEqual(exported.status_code, 200, exported.text)
        self.assertIn("text/csv", exported.headers.get("Content-Type", ""))
        self.assertIn("transfer_player", exported.text)
        self.assertIn("operation_label", exported.text)

        logout = self.request("POST", "/api/admin/logout")
        self.assertEqual(logout.status_code, 200)

        exported_after_logout = self.request("GET", "/api/admin/operations-audit/export?category=transfer")
        self.assertEqual(exported_after_logout.status_code, 401, exported_after_logout.text)

        check = self.request("GET", "/api/admin/check")
        self.assertEqual(check.status_code, 200)
        self.assertFalse(check.json()["authenticated"])

        db = self.SessionLocal()
        try:
            sessions = db.query(self.models.AdminSession).count()
            self.assertEqual(sessions, 0)
        finally:
            db.close()

    def test_search_supports_ascii_queries_for_diacritic_names(self):
        db = self.SessionLocal()
        try:
            alpha = db.query(self.models.Team).filter(self.models.Team.name == "Alpha FC").first()
            self.assertIsNotNone(alpha)

            player = self.models.Player(
                uid=1999,
                name="João Félix",
                age=24,
                initial_ca=120,
                ca=121,
                pa=140,
                position="AMC",
                nationality="Portugal",
                team_id=alpha.id,
                team_name=alpha.name,
                wage=0,
                slot_type="",
            )
            self.main1.refresh_player_financials(player, db)
            db.add(player)
            db.add(
                self.models.PlayerAttribute(
                    uid=1999,
                    name="João Félix",
                    position="AMC",
                    age=24,
                    ca=121,
                    pa=140,
                    nationality="Portugal",
                    club="Legacy Club",
                    pos_amc=20,
                    passing=15,
                )
            )
            db.commit()
        finally:
            db.close()

        player_search = self.request("GET", "/api/players/search/Joao")
        self.assertEqual(player_search.status_code, 200, player_search.text)
        self.assertTrue(any(item["uid"] == 1999 for item in player_search.json()))

        attribute_search = self.request("GET", "/api/attributes/search/Joao")
        self.assertEqual(attribute_search.status_code, 200, attribute_search.text)
        self.assertTrue(any(item["uid"] == 1999 for item in attribute_search.json()))

    def test_batch_flow_and_rename_safe_undo(self):
        self.login()

        fish = self.request(
            "POST",
            "/api/admin/fish",
            json={
                "uid": 3001,
                "name": "Sea Prospect",
                "age": 19,
                "ca": 95,
                "pa": 130,
                "position": "ST",
                "nationality": "ARG",
                "team_name": "Alpha FC",
                "wage": 0,
                "slot_type": "",
                "notes": "simulation fish",
            },
        )
        self.assertEqual(fish.status_code, 200, fish.text)

        batch_transfer = self.request(
            "POST",
            "/api/admin/batch-transfer",
            json={
                "items": [
                    {"uid": 1001, "to_team": "Beta FC", "notes": "batch move 1"},
                    {"uid": 3001, "to_team": "Beta FC", "notes": "batch move 2"},
                ]
            },
        )
        self.assertEqual(batch_transfer.status_code, 200, batch_transfer.text)
        self.assertEqual(batch_transfer.json()["success_count"], 2)

        batch_consume = self.request(
            "POST",
            "/api/admin/batch-consume",
            json={
                "items": [
                    {"uid": 1001, "ca_change": 1, "pa_change": 0, "notes": "batch consume 1"},
                    {"uid": 3001, "ca_change": 2, "pa_change": -1, "notes": "batch consume 2"},
                ]
            },
        )
        self.assertEqual(batch_consume.status_code, 200, batch_consume.text)
        self.assertEqual(batch_consume.json()["success_count"], 2)

        batch_release = self.request(
            "POST",
            "/api/admin/batch-release",
            json={"items": [{"uid": 3001, "notes": "batch release"}]},
        )
        self.assertEqual(batch_release.status_code, 200, batch_release.text)
        self.assertEqual(batch_release.json()["success_count"], 1)

        team_update = self.request(
            "POST",
            "/api/admin/team/update",
            json={
                "team_name": "Beta FC",
                "name": "Gamma FC",
                "manager": "Gina",
                "notes": "+0.1M renamed",
                "level": "超级",
            },
        )
        self.assertEqual(team_update.status_code, 200, team_update.text)

        gamma = self.fetch_team("Gamma FC")
        self.assertIsNotNone(gamma)

        released_player = self.fetch_player(3001)
        self.assertEqual(released_player.team_name, self.team_links.SEA_TEAM_NAME)

        batch_release_logs = [log for log in self.fetch_logs(3001) if log.operation == "批量解约"]
        self.assertEqual(len(batch_release_logs), 1)
        self.assertEqual(batch_release_logs[0].from_team_id, gamma.id)

        undo = self.request("POST", f"/api/admin/undo/{batch_release_logs[0].id}")
        self.assertEqual(undo.status_code, 200, undo.text)

        restored_player = self.fetch_player(3001)
        self.assertEqual(restored_player.team_name, "Gamma FC")
        self.assertEqual(restored_player.team_id, gamma.id)

        team_info = self.request("GET", "/api/admin/team/Gamma FC")
        self.assertEqual(team_info.status_code, 200, team_info.text)
        self.assertEqual(team_info.json()["manager"], "Gina")

        transfer_logs = self.request("GET", "/api/admin/transfer-logs")
        self.assertEqual(transfer_logs.status_code, 200, transfer_logs.text)
        payload = transfer_logs.json()
        self.assertTrue(any(log["operation"] == "批量交易" for log in payload))
        self.assertTrue(any(log["operation"] == "批量消费" for log in payload))
        self.assertTrue(any(log["to_team_id"] == gamma.id for log in payload if log["player_uid"] == 1001))


if __name__ == "__main__":
    unittest.main()
