import unittest
from pathlib import Path
from types import SimpleNamespace
from tempfile import TemporaryDirectory
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from services import export_service


class ExportServiceTemplateTests(unittest.TestCase):
    def test_home_export_preserves_import_workbook_and_updates_current_data(self):
        with TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "87届HEIGO名单.xlsx"
            workbook = Workbook()
            overview = workbook.active
            overview.title = "信息总览"
            overview.append(["届数", 86, None, "序号", "级别", "球队名", "主教", "球队人数", "门将人数", "工资", "额外工资", "税后", "最终工资", "8M", "7M", "伪名", "总身价", "平均身价", "平均CA", "平均PA", "成长总计", "备注"])
            overview.append(["成长年龄上限", 24, None, "序号", "级别", "球队名", "主教", "球队人数", "门将人数", "工资", "额外工资", "税后", "最终工资", "8M", "7M", "伪名", "总身价", "平均身价", "平均CA", "平均PA", "成长总计", "备注"])
            overview.append([None, None, "Old FC", 1, "超级", "=C3", "Old Coach"])
            overview["H3"] = '=COUNTIF(联赛名单!O:O,信息总览!F3)'
            overview["I3"] = '=COUNTIFS(联赛名单!O:O,信息总览!F3,联赛名单!M:M,"GK")'
            overview["J3"] = '=SUMIF(联赛名单!O:O,信息总览!F3,联赛名单!X:X)'
            overview["L3"] = '=IF(J3+K3-$B$5>0.3,"拍卖",J3+K3)'
            overview["M3"] = '=IF(L3<8,8,L3)'
            overview["N3"] = '=COUNTIFS(联赛名单!O:O,信息总览!F3,联赛名单!Z:Z,"8M")'
            overview["O3"] = '=COUNTIFS(联赛名单!O:O,信息总览!F3,联赛名单!Z:Z,"7M")'
            overview["P3"] = '=COUNTIFS(联赛名单!O:O,信息总览!F3,联赛名单!Z:Z,"伪名")'
            overview["Q3"] = '=SUMIF(联赛名单!O:O,信息总览!F3,联赛名单!V:V)'
            overview["R3"] = '=Q3/H3'
            overview["S3"] = '=SUMIF(联赛名单!O:O,信息总览!F3,联赛名单!H:H)/H3'
            overview["T3"] = '=SUMIF(联赛名单!O:O,信息总览!F3,联赛名单!J:J)/H3'
            overview["U3"] = '=SUMIF(联赛名单!O:O,信息总览!F3,联赛名单!H:H)-SUMIF(联赛名单!O:O,信息总览!F3,联赛名单!G:G)'
            overview["A8"] = "总工资"
            overview["B8"] = "=SUM(M3:M3)"
            overview["F3"].fill = PatternFill("solid", fgColor="DDEEFF")

            players = workbook.create_sheet("联赛名单")
            players.append(["名额", "编号", "存档有无", "姓名", "原年龄", "年龄", "初始CA", "当前CA", "成长CA", "PA", "消费PA", "消费后PA", "位置", "国籍", "俱乐部", "更新俱乐部", "俱乐部改变", "级别", "初始身价", "当前身价", "潜力身价", "身价", "系数", "工资", "初始", "名额"])
            players.append(["=Z2", 1, '=IF(B2="","",1)', "Old Player", None, 20, 100, 110, "=H2-G2", 130, None, None, "MC", "CN", "Old FC", "Old FC"])
            players["D2"].fill = PatternFill("solid", fgColor="CCFFCC")
            extra = workbook.create_sheet("步骤")
            extra["A1"] = "保留此辅助页"
            workbook.save(template_path)

            team = SimpleNamespace(
                id=10, level="超级", name="New FC", manager="New Coach", team_size=1, gk_count=0,
                wage=7.5, extra_wage=0.2, wage_cap=10.25, after_tax=0, final_wage=7.7, count_8m=0, count_7m=1,
                count_fake=0, total_value=8, avg_value=8, avg_ca=120, avg_pa=140, total_growth=20,
                notes="当前备注",
            )
            player = SimpleNamespace(
                uid=1, name="Current Player", age=21, initial_ca=105, ca=120, pa=145,
                position="MC", nationality="China", team_name="New FC", wage=7.5, slot_type="7M",
            )
            db = MagicMock()
            db.query.return_value.all.return_value = [SimpleNamespace(key="总工资", value=9999)]

            with (
                patch.object(export_service, "resolve_import_root", return_value=Path(temp_dir)),
                patch.object(export_service, "_get_export_teams", return_value=[team]),
                patch.object(export_service, "_get_export_players", return_value=[player]),
                patch.object(export_service, "collect_team_stat_overlays", return_value={}),
            ):
                output, filename = export_service.build_export_excel(db)

            exported = load_workbook(output, data_only=False)
            self.assertEqual(exported.sheetnames, ["信息总览", "联赛名单", "步骤"])
            self.assertEqual(exported["步骤"]["A1"].value, "保留此辅助页")
            self.assertEqual(exported["信息总览"]["C3"].value, "New FC")
            self.assertEqual(exported["信息总览"]["F3"].value, "New FC")
            self.assertEqual(exported["信息总览"]["F3"].fill.fgColor.rgb, "00DDEEFF")
            self.assertEqual(exported["信息总览"]["W2"].value, "工资帽")
            self.assertEqual(exported["信息总览"]["W3"].value, 10.25)
            self.assertTrue(exported["信息总览"]["H3"].value.startswith("=COUNTIF"))
            self.assertEqual(exported["信息总览"]["K3"].value, "=ROUND(0.2,3)")
            self.assertIn("W3", exported["信息总览"]["L3"].value)
            self.assertIn("ROUND", exported["信息总览"]["L3"].value)
            self.assertNotIn("-0.1", exported["信息总览"]["L3"].value)
            self.assertIn('L3="拍卖"', exported["信息总览"]["M3"].value)
            self.assertTrue(exported["信息总览"]["N3"].value.startswith("=COUNTIFS"))
            self.assertTrue(exported["信息总览"]["Q3"].value.startswith("=SUMIF"))
            self.assertEqual(exported["信息总览"]["B8"].value, "=SUM(M3:M3)")
            self.assertEqual(exported["联赛名单"]["D2"].value, "Current Player")
            self.assertEqual(exported["联赛名单"]["O2"].value, "New FC")
            self.assertEqual(exported["联赛名单"]["P2"].value, "New FC")
            self.assertEqual(exported["联赛名单"]["C2"].value, '=IF(B2="","",1)')
            self.assertEqual(exported["联赛名单"]["S2"].value, '=IF(G2<115,1,INT((G2-95)/10))')
            self.assertIn("'信息总览'!$B$4", exported["联赛名单"]["V2"].value)
            self.assertIn("UPPER(TRIM(M2))", exported["联赛名单"]["W2"].value)
            self.assertIn("Y2=1", exported["联赛名单"]["W2"].value)
            self.assertIn("(Y2+T2)/2=1", exported["联赛名单"]["W2"].value)
            self.assertEqual(exported["联赛名单"]["X2"].value, "=ROUND(V2*W2,3)")
            self.assertIn("伪名", exported["联赛名单"]["Z2"].value)
            self.assertEqual(exported["联赛名单"]["D2"].fill.fgColor.rgb, "00CCFFCC")
            self.assertEqual(exported.calculation.calcMode, "auto")
            self.assertTrue(exported.calculation.fullCalcOnLoad)
            self.assertTrue(exported.calculation.forceFullCalc)
            self.assertTrue(exported.calculation.calcOnSave)
            self.assertFalse(exported.calculation.calcCompleted)
            self.assertTrue(exported.calculation.fullPrecision)
            self.assertEqual(exported.calculation.calcId, 0)
            self.assertTrue(filename.endswith(".xlsx"))


if __name__ == "__main__":
    unittest.main()
