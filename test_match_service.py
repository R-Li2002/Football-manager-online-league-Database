from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base
from models import Match, MatchPlayerEvent, Player, Team
from schemas_write import MatchBatchUpdateItem, MatchBatchUpdateRequest, MatchPlayerEventUpdateItem, MatchUpdateRequest
from services import competition_work_service, match_service, player_ranking_service


class MatchServiceTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.db = Session()
        for name, level in (
            ("Alpha", "超级"),
            ("Beta", "超级"),
            ("Gamma", "超级"),
            ("Delta", "超级"),
        ):
            self.db.add(Team(name=name, level=level, manager=f"{name} Boss"))
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def test_parse_block_schedule_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "超级赛程"
        ws.append([None, None, 1, None, None, None, 2])
        ws.append([None, "Alpha", "vs", "Beta", None, "Beta", "vs", "Alpha"])
        ws.append([None, "Gamma", "vs", "Delta", None, "Delta", "vs", "Gamma"])
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "schedule.xlsx"
            wb.save(path)
            fixtures = match_service.parse_schedule_workbook(path)

        self.assertEqual(len(fixtures), 4)
        self.assertEqual((fixtures[0].level, fixtures[0].round_no), ("超级", 1))
        self.assertEqual((fixtures[0].home_team_name, fixtures[0].away_team_name), ("Alpha", "Beta"))
        self.assertEqual((fixtures[1].round_no, fixtures[1].home_team_name, fixtures[1].away_team_name), (2, "Beta", "Alpha"))

    def test_parse_schedule_does_not_treat_numeric_team_names_as_round_headers(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "乙级赛程"
        ws.append([None, None, 1, None, None, None, 2])
        ws.append([None, "Bayer 04", "vs", "FC Heidenheim 1846", None, "FC Heidenheim 1846", "vs", "Bayer 04"])
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "numeric-team-names.xlsx"
            wb.save(path)
            fixtures = match_service.parse_schedule_workbook(path)

        self.assertEqual(len(fixtures), 2)
        self.assertEqual((fixtures[0].round_no, fixtures[0].home_team_name, fixtures[0].away_team_name), (1, "Bayer 04", "FC Heidenheim 1846"))
        self.assertEqual((fixtures[1].round_no, fixtures[1].home_team_name, fixtures[1].away_team_name), (2, "FC Heidenheim 1846", "Bayer 04"))

    def test_standings_are_calculated_from_played_matches(self):
        alpha = self.db.query(Team).filter(Team.name == "Alpha").one()
        beta = self.db.query(Team).filter(Team.name == "Beta").one()
        self.db.add_all(
            [
                Match(
                    level="超级",
                    round_no=1,
                    home_team_id=alpha.id,
                    home_team_name="Alpha Short",
                    away_team_id=beta.id,
                    away_team_name="Beta Short",
                    home_score=2,
                    away_score=0,
                    status="played",
                ),
                Match(level="超级", round_no=1, home_team_name="Gamma", away_team_name="Delta", home_score=1, away_score=1, status="played"),
                Match(level="超级", round_no=2, home_team_name="Alpha", away_team_name="Gamma", status="scheduled"),
            ]
        )
        self.db.commit()

        standings = match_service.get_standings(self.db)
        rows = [row for row in standings.rows if row.level == "超级"]

        self.assertEqual([row.team_name for row in rows], ["Alpha", "Delta", "Gamma", "Beta"])
        self.assertEqual(rows[0].points, 3)
        self.assertEqual(rows[0].goal_difference, 2)
        self.assertEqual(rows[0].win_rate, 100.0)
        self.assertEqual((rows[0].home_wins, rows[0].home_draws, rows[0].home_losses), (1, 0, 0))
        self.assertEqual((rows[0].away_wins, rows[0].away_draws, rows[0].away_losses), (0, 0, 0))
        self.assertEqual(rows[1].points, 1)
        self.assertEqual((rows[1].away_wins, rows[1].away_draws, rows[1].away_losses), (0, 1, 0))

    def test_standings_can_be_limited_to_one_level(self):
        first_home = Team(name="First Home", level="甲级", manager="First Home Boss")
        first_away = Team(name="First Away", level="甲级", manager="First Away Boss")
        self.db.add_all([first_home, first_away])
        self.db.commit()
        self.db.add(
            Match(
                level="甲级",
                round_no=1,
                home_team_id=first_home.id,
                home_team_name=first_home.name,
                away_team_id=first_away.id,
                away_team_name=first_away.name,
                home_score=1,
                away_score=0,
                status="played",
            )
        )
        self.db.commit()

        standings = match_service.get_standings(self.db, level="甲级")

        self.assertEqual(standings.levels, ["甲级"])
        self.assertEqual([row.team_name for row in standings.rows], ["First Home", "First Away"])
        self.assertTrue(all(row.level == "甲级" for row in standings.rows))

    def test_standings_resolve_legacy_schedule_alias_names(self):
        self.db.add(Team(name="RB Leipzig", level="超级", manager="RB Leipzig Boss"))
        self.db.commit()
        self.db.add(
            Match(
                level="超级",
                round_no=1,
                home_team_name="RBL",
                away_team_name="Alpha",
                home_score=2,
                away_score=1,
                status="played",
            )
        )
        self.db.commit()

        standings = match_service.get_standings(self.db)
        leipzig = next(row for row in standings.rows if row.team_name == "RB Leipzig")
        alpha = next(row for row in standings.rows if row.team_name == "Alpha")

        self.assertEqual(leipzig.points, 3)
        self.assertEqual(leipzig.played, 1)
        self.assertEqual(leipzig.goals_for, 2)
        self.assertEqual(alpha.losses, 1)

    def test_import_latest_schedule_resolves_alias_team_ids(self):
        self.db.add(Team(name="RB Leipzig", level="超级", manager="RB Leipzig Boss"))
        self.db.commit()
        wb = Workbook()
        ws = wb.active
        ws.title = "超级赛程"
        ws.append([None, None, 1])
        ws.append([None, "RBL", "vs", "Alpha"])
        with tempfile.TemporaryDirectory() as tmpdir:
            schedule_root = Path(tmpdir)
            path = schedule_root / "rounds.xlsx"
            wb.save(path)
            original_finder = match_service.find_latest_schedule_file
            match_service.find_latest_schedule_file = lambda: path
            try:
                logs = []
                response = match_service.import_latest_schedule(
                    self.db,
                    "admin",
                    lambda operation, details, operator: logs.append((operation, details, operator)),
                )
            finally:
                match_service.find_latest_schedule_file = original_finder

        leipzig = self.db.query(Team).filter(Team.name == "RB Leipzig").one()
        match = (
            self.db.query(Match)
            .filter(Match.home_team_name == "RBL", Match.away_team_name == "Alpha")
            .first()
        )

        self.assertTrue(response.success)
        self.assertEqual(response.warnings, [])
        self.assertIsNotNone(match)
        self.assertEqual(match.home_team_id, leipzig.id)
        self.assertEqual(logs[0][0], "赛程导入")

    def test_schedule_alias_resolves_bayer_04(self):
        team = Team(name="Bayer 04 Leverkusen", level="乙级", manager="Bayer Boss")
        self.db.add(team)
        self.db.commit()

        resolved = match_service._resolve_schedule_team(self.db, None, "Bayer 04")

        self.assertIsNotNone(resolved)
        self.assertEqual(resolved.id, team.id)

    def test_update_match_result_marks_played_and_recalculates_standings(self):
        match = Match(level="超级", round_no=1, home_team_name="Alpha", away_team_name="Beta", status="scheduled")
        self.db.add(match)
        self.db.commit()

        logs = []
        response = match_service.update_match_result(
            self.db,
            "admin",
            match.id,
            MatchUpdateRequest(home_score=0, away_score=3, status="played"),
            lambda operation, details, operator: logs.append((operation, details, operator)),
        )

        self.assertTrue(response["success"])
        standings = match_service.get_standings(self.db)
        beta = next(row for row in standings.rows if row.team_name == "Beta")
        alpha = next(row for row in standings.rows if row.team_name == "Alpha")
        self.assertEqual(beta.points, 3)
        self.assertEqual(alpha.losses, 1)
        self.assertEqual(logs[0][0], "赛程比分编辑")

    def test_batch_update_match_results_infers_played_and_resets_scheduled(self):
        played = Match(level="超级", round_no=1, home_team_name="Alpha", away_team_name="Beta", status="scheduled")
        reset = Match(level="超级", round_no=2, home_team_name="Gamma", away_team_name="Delta", home_score=2, away_score=1, status="played")
        self.db.add_all([played, reset])
        self.db.commit()

        logs = []
        response = match_service.batch_update_match_results(
            self.db,
            "admin",
            MatchBatchUpdateRequest(
                matches=[
                    MatchBatchUpdateItem(match_id=played.id, home_score=1, away_score=1, status=""),
                    MatchBatchUpdateItem(match_id=reset.id, home_score=None, away_score=None, status=""),
                ]
            ),
            lambda operation, details, operator: logs.append((operation, details, operator)),
        )

        self.assertTrue(response["success"])
        self.db.refresh(played)
        self.db.refresh(reset)
        self.assertEqual(played.status, "played")
        self.assertEqual((played.home_score, played.away_score), (1, 1))
        self.assertEqual(reset.status, "scheduled")
        self.assertIsNone(reset.home_score)
        self.assertIsNone(reset.away_score)
        self.assertEqual(logs[0][0], "赛程比分批量编辑")

    def test_player_rankings_are_calculated_from_schedule_events(self):
        alpha = self.db.query(Team).filter(Team.name == "Alpha").one()
        beta = self.db.query(Team).filter(Team.name == "Beta").one()
        scorer = Player(uid=101, name="Alpha Scorer", team_id=alpha.id, team_name="Alpha")
        creator = Player(uid=102, name="Alpha Creator", team_id=alpha.id, team_name="Alpha")
        mvp = Player(uid=201, name="Beta Keeper", team_id=beta.id, team_name="Beta")
        match = Match(
            level="超级",
            round_no=1,
            home_team_id=alpha.id,
            home_team_name="Alpha",
            away_team_id=beta.id,
            away_team_name="Beta",
            status="scheduled",
        )
        self.db.add_all([scorer, creator, mvp, match])
        self.db.commit()

        response = match_service.update_match_result(
            self.db,
            "admin",
            match.id,
            MatchUpdateRequest(
                home_score=2,
                away_score=1,
                status="played",
                events=[
                    MatchPlayerEventUpdateItem(team_name="Alpha", player_uid=101, player_name="Alpha Scorer", event_type="goal", quantity=2),
                    MatchPlayerEventUpdateItem(team_name="Alpha", player_uid=102, player_name="Alpha Creator", event_type="assist", quantity=1),
                    MatchPlayerEventUpdateItem(team_name="Beta", player_uid=201, player_name="Beta Keeper", event_type="mvp", quantity=1),
                ],
            ),
            lambda operation, details, operator: None,
        )

        self.assertTrue(response["success"])
        rankings = player_ranking_service.get_player_rankings(self.db)
        rows_by_name = {row.player_name: row for row in rankings.rows if row.level == "超级"}
        coverage = next(row for row in rankings.coverage if row.level == "超级")

        self.assertEqual(rows_by_name["Alpha Scorer"].goals, 2)
        self.assertEqual(rows_by_name["Alpha Creator"].assists, 1)
        self.assertEqual(rows_by_name["Beta Keeper"].mvps, 1)
        self.assertEqual(coverage.played_matches, 1)
        self.assertEqual(coverage.matches_with_events, 1)
        self.assertEqual(coverage.matches_missing_events, 0)
        self.assertEqual(coverage.goal_quantity, 2)
        self.assertEqual(coverage.assist_quantity, 1)
        self.assertEqual(coverage.mvp_quantity, 1)

    def test_own_goal_completes_score_without_crediting_player_ranking(self):
        alpha = self.db.query(Team).filter(Team.name == "Alpha").one()
        beta = self.db.query(Team).filter(Team.name == "Beta").one()
        alpha_scorer = Player(uid=301, name="Alpha Scorer", team_id=alpha.id, team_name="Alpha")
        beta_scorer = Player(uid=302, name="Beta Scorer", team_id=beta.id, team_name="Beta")
        match = Match(
            level="超级",
            round_no=1,
            home_team_id=alpha.id,
            home_team_name="Alpha",
            away_team_id=beta.id,
            away_team_name="Beta",
            status="scheduled",
        )
        self.db.add_all([alpha_scorer, beta_scorer, match])
        self.db.commit()

        response = match_service.update_match_result(
            self.db,
            "admin",
            match.id,
            MatchUpdateRequest(
                home_score=2,
                away_score=1,
                status="played",
                events=[
                    MatchPlayerEventUpdateItem(team_name="Alpha", player_uid=301, player_name="Alpha Scorer", event_type="goal"),
                    MatchPlayerEventUpdateItem(team_name="Alpha", player_uid=None, player_name="乌龙球", event_type="own_goal"),
                    MatchPlayerEventUpdateItem(team_name="Beta", player_uid=302, player_name="Beta Scorer", event_type="goal"),
                    MatchPlayerEventUpdateItem(team_name="Beta", player_uid=302, player_name="Beta Scorer", event_type="mvp"),
                ],
            ),
            lambda operation, details, operator: None,
        )

        self.assertTrue(response["success"])
        events = self.db.query(MatchPlayerEvent).filter(MatchPlayerEvent.match_id == match.id).all()
        own_goal = next(event for event in events if event.event_type == "own_goal")
        self.assertIsNone(own_goal.player_uid)
        self.assertEqual(own_goal.player_name, "乌龙球")
        is_ready, events_ready, issue_codes, issue_messages = competition_work_service.evaluate_match_readiness(match, events)
        self.assertTrue(is_ready)
        self.assertTrue(events_ready)
        self.assertNotIn("missing_events", issue_codes)
        self.assertFalse(any("还缺少" in message for message in issue_messages))

        rankings = player_ranking_service.get_player_rankings(self.db)
        rows_by_name = {row.player_name: row for row in rankings.rows if row.level == "超级"}
        self.assertEqual(rows_by_name["Alpha Scorer"].goals, 1)
        self.assertEqual(rows_by_name["Beta Scorer"].goals, 1)
        self.assertNotIn("乌龙球", rows_by_name)

    def test_player_ranking_coverage_reports_played_matches_missing_events(self):
        self.db.add(
            Match(
                level="超级",
                round_no=1,
                home_team_name="Alpha",
                away_team_name="Beta",
                home_score=1,
                away_score=0,
                status="played",
            )
        )
        self.db.commit()

        rankings = player_ranking_service.get_player_rankings(self.db)
        coverage = next(row for row in rankings.coverage if row.level == "超级")

        self.assertEqual(coverage.played_matches, 1)
        self.assertEqual(coverage.matches_with_events, 0)
        self.assertEqual(coverage.matches_missing_events, 1)

    def test_player_rankings_can_be_limited_to_one_level(self):
        rankings = player_ranking_service.get_player_rankings(self.db, level="甲级")

        self.assertEqual(rankings.levels, ["甲级"])
        self.assertEqual(rankings.rows, [])
        self.assertEqual([item.level for item in rankings.coverage], ["甲级"])


if __name__ == "__main__":
    unittest.main()
