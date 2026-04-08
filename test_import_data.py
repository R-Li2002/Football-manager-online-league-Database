import unittest
from unittest.mock import patch
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base, init_database
from import_data import (
    DEFAULT_LEAGUE_INFO_VALUES,
    HIDDEN_SEA_TEAM_LEVEL,
    HIDDEN_SEA_TEAM_NAME,
    decode_player_habit_value,
    run_import,
)
from models import LeagueInfo, Player, PlayerAttribute, PlayerAttributeVersion, Team
import imports_runtime.persistence as runtime_persistence
from services import read_service


TEAM_HEADER_ROW = [
    None,
    None,
    None,
    "球队名",
    "主教",
    "级别",
    "额外工资",
    "税后",
    "备注",
]


DEFAULT_TEAM_ROWS = [
    {"info_key": "届数", "info_value": 85, "name": "Alpha FC", "manager": "Coach A", "level": "超级", "extra_wage": 0.2, "after_tax": 0, "notes": ""},
    {"info_key": "成长年龄上限", "info_value": 24, "name": "Beta FC", "manager": "Coach B", "level": "甲级", "extra_wage": 0.0, "after_tax": 0, "notes": "+0.1M"},
]

DEFAULT_PLAYER_ROWS = [
    {"uid": 1, "name": "张三", "age": 20, "initial_ca": 120, "ca": 125, "pa": 150, "position": "MC", "nationality": "CN", "club_name": "Alpha FC"},
    {"uid": 2, "name": "李四", "age": 21, "initial_ca": 118, "ca": 121, "pa": 148, "position": "GK", "nationality": "CN", "club_name": "Beta FC"},
]


def default_mapping_rows(player_rows):
    return [
        {
            "uid": player["uid"],
            "name": player["name"],
            "age": player["age"],
            "ca": player["ca"],
            "pa": player["pa"],
            "nationality": player["nationality"],
            "team_name": player.get("league_team_name", player["club_name"]),
            "position": player["position"],
        }
        for player in player_rows
    ]


def write_workbook(
    path: Path,
    *,
    include_position_column: bool = True,
    team_rows=None,
    player_rows=None,
    mapping_rows=None,
) -> None:
    team_rows = team_rows or DEFAULT_TEAM_ROWS
    player_rows = player_rows or DEFAULT_PLAYER_ROWS
    mapping_rows = mapping_rows or default_mapping_rows(player_rows)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "信息总览"

    overview.append(["说明：测试导入", None, None, None, None, None, None, None, None])
    overview.append(TEAM_HEADER_ROW)
    for team in team_rows:
        overview.append(
            [
                team["info_key"],
                team["info_value"],
                None,
                team["name"],
                team["manager"],
                team["level"],
                team["extra_wage"],
                team["after_tax"],
                team["notes"],
            ]
        )
    overview.append(["8M名额", 1.5])
    overview.append(["7M名额", 1.3])
    overview.append(["可成长名额", 1.1])
    overview.append(["非名PA6M", 0.9])
    overview.append(["非名身价1M", 1.0])
    overview.append(["非名其他", 0.7])
    overview.append(["注：名额GK系数为 1"])

    players_sheet = workbook.create_sheet("联赛名单")
    player_headers = [
        "编\u200b号",
        "姓\u200b名",
        "年\u200b龄",
        "初始CA",
        "当前CA",
        "PA",
        "位\u200b置",
        "国籍",
        "俱乐部",
    ]
    include_league_team_column = any("league_team_name" in player for player in player_rows)
    if include_league_team_column:
        player_headers.append("联赛球队")
    if not include_position_column:
        player_headers.remove("位\u200b置")
    players_sheet.append(player_headers)
    for player in player_rows:
        row = [
            player["uid"],
            player["name"],
            player["age"],
            player["initial_ca"],
            player["ca"],
            player["pa"],
            player["position"],
            player["nationality"],
            player["club_name"],
        ]
        if include_league_team_column:
            row.append(player.get("league_team_name", ""))
        if include_position_column:
            players_sheet.append(row)
        else:
            if include_league_team_column:
                players_sheet.append([row[0], row[1], row[2], row[3], row[4], row[5], row[7], row[8], row[9]])
            else:
                players_sheet.append([row[0], row[1], row[2], row[3], row[4], row[5], row[7], row[8]])

    mapping_sheet = workbook.create_sheet("球员对应球队")
    mapping_sheet.append(["UID", "球员名", "年龄", "CA", "PA", "国籍", "球队", "位置", "联赛名单有无"])
    for player in mapping_rows:
        mapping_sheet.append(
            [
                player["uid"],
                player["name"],
                player["age"],
                player["ca"],
                player["pa"],
                player["nationality"],
                player["team_name"],
                player["position"],
                "",
            ]
        )

    workbook.save(path)


def write_attributes_csv(path: Path) -> None:
    df = pd.DataFrame(
        [
            {"UID": 1, "姓名": "张三", "位置": "MC", "年龄": 20, "ca": 125, "pa": 150, "国籍": "CN", "俱乐部": "Alpha FC", "角球": 12},
            {"UID": 2, "姓名": "李四", "位置": "GK", "年龄": 21, "ca": 121, "pa": 148, "国籍": "CN", "俱乐部": "Beta FC", "角球": 5},
        ]
    )
    df.to_csv(path, index=False, encoding="utf-8-sig")


def write_attributes_xlsx_2620(path: Path) -> None:
    headers = [
        "名字", "UID", "年龄", "当前ca", "当前PA", "球队", "位置", "传球", "传中", "盯人", "技术", "盘带", "抢断", "射门", "停球",
        "头球", "远射", "点球", "角球", "界外球", "任意球", "想象力", "防守站位", "工作投入", "集中", "决断", "领导力", "侵略性",
        "视野", "团队合作", "无球跑动", "意志力", "勇敢", "预判", "镇定", "爆发力", "弹跳", "灵活", "耐力", "平衡2", "强壮",
        "速度", "体质", "左脚", "右脚", "出击", "大脚开球", "反应", "拦截传中", "击球倾向", "手控球", "手抛球", "一对一",
        "神经指数", "指挥防守", "制空能力", "稳定性", "肮脏动作", "大赛发挥", "多样性", "受伤倾向", "适应性", "野心", "争论倾向",
        "忠诚", "抗压能力", "职业素养", "体育精神", "情绪控制", "国籍", "前腰", "左前腰", "右前腰", "中后卫", "左后卫", "右后卫",
        "后腰", "门将", "中前卫", "左前卫", "右前卫", "前锋", "进攻型左边卫", "进攻型右边卫", "习惯", "身高", "负潜",
    ]

    rows = [
        {
            "名字": "张三",
            "UID": 1,
            "年龄": 20,
            "当前ca": 125,
            "当前PA": 150,
            "球队": "Alpha FC",
            "位置": "MC",
            "国籍": "CN",
            "传球": 12,
            "传中": 10,
            "盯人": 6,
            "技术": 14,
            "盘带": 16,
            "抢断": 9,
            "射门": 8,
            "停球": 15,
            "头球": 7,
            "远射": 11,
            "点球": 9,
            "角球": 8,
            "界外球": 5,
            "任意球": 6,
            "想象力": 13,
            "防守站位": 12,
            "工作投入": 9,
            "集中": 11,
            "决断": 13,
            "领导力": 7,
            "侵略性": 8,
            "视野": 14,
            "团队合作": 15,
            "无球跑动": 14,
            "意志力": 14,
            "勇敢": 10,
            "预判": 12,
            "镇定": 11,
            "爆发力": 10,
            "弹跳": 13,
            "灵活": 15,
            "耐力": 11,
            "平衡2": 13,
            "强壮": 9,
            "速度": 14,
            "体质": 12,
            "左脚": 12,
            "右脚": 15,
            "出击": 2,
            "大脚开球": 3,
            "反应": 4,
            "拦截传中": 5,
            "击球倾向": 3,
            "手控球": 2,
            "手抛球": 1,
            "一对一": 4,
            "神经指数": 7,
            "指挥防守": 6,
            "制空能力": 3,
            "稳定性": 10,
            "肮脏动作": 4,
            "大赛发挥": 8,
            "多样性": 9,
            "受伤倾向": 5,
            "适应性": 11,
            "野心": 12,
            "争论倾向": 3,
            "忠诚": 10,
            "抗压能力": 13,
            "职业素养": 14,
            "体育精神": 15,
            "情绪控制": 9,
            "中前卫": 18,
            "习惯": 525440,
            "身高": 182,
            "负潜": -10,
        },
        {
            "名字": "李四",
            "UID": 2,
            "年龄": 21,
            "当前ca": 121,
            "当前PA": 148,
            "球队": "Beta FC",
            "位置": "GK",
            "国籍": "CN",
            "传球": 8,
            "传中": 2,
            "盯人": 5,
            "技术": 7,
            "盘带": 4,
            "抢断": 3,
            "射门": 2,
            "停球": 9,
            "头球": 6,
            "远射": 2,
            "点球": 1,
            "角球": 1,
            "界外球": 2,
            "任意球": 1,
            "想象力": 6,
            "防守站位": 13,
            "工作投入": 9,
            "集中": 15,
            "决断": 12,
            "领导力": 11,
            "侵略性": 6,
            "视野": 7,
            "团队合作": 10,
            "无球跑动": 3,
            "意志力": 16,
            "勇敢": 14,
            "预判": 13,
            "镇定": 12,
            "爆发力": 9,
            "弹跳": 14,
            "灵活": 8,
            "耐力": 12,
            "平衡2": 11,
            "强壮": 13,
            "速度": 10,
            "体质": 12,
            "左脚": 8,
            "右脚": 15,
            "出击": 11,
            "大脚开球": 13,
            "反应": 18,
            "拦截传中": 14,
            "击球倾向": 6,
            "手控球": 17,
            "手抛球": 11,
            "一对一": 16,
            "神经指数": 5,
            "指挥防守": 12,
            "制空能力": 15,
            "稳定性": 16,
            "肮脏动作": 2,
            "大赛发挥": 14,
            "多样性": 4,
            "受伤倾向": 6,
            "适应性": 8,
            "野心": 10,
            "争论倾向": 2,
            "忠诚": 9,
            "抗压能力": 15,
            "职业素养": 16,
            "体育精神": 14,
            "情绪控制": 12,
            "门将": 20,
            "习惯": float(2 ** 62),
            "身高": 190,
            "负潜": -10,
        },
    ]

    matrix = [list(range(1, len(headers) + 1)), headers]
    matrix.extend([[row.get(header, "") for header in headers] for row in rows])
    pd.DataFrame(matrix).to_excel(path, sheet_name="数据", header=False, index=False)


class ImportDataTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = TemporaryDirectory()
        self.root_dir = Path(self.temp_dir.name)
        self.workbook_path = self.root_dir / "fixture.xlsx"
        self.attributes_csv_path = self.root_dir / "fixture.csv"
        write_workbook(self.workbook_path)
        write_attributes_csv(self.attributes_csv_path)

        self.db_path = self.root_dir / "import_test.db"
        self.engine = create_engine(f"sqlite:///{self.db_path}")
        Base.metadata.create_all(bind=self.engine)
        init_database(target_engine=self.engine)
        self.SessionLocal = sessionmaker(bind=self.engine, autocommit=False, autoflush=False)

    def tearDown(self):
        self.engine.dispose()
        self.temp_dir.cleanup()

    def count_rows(self) -> dict[str, int]:
        session = self.SessionLocal()
        try:
            return {
                "league_info": session.query(LeagueInfo).count(),
                "teams": session.query(Team).count(),
                "players": session.query(Player).count(),
                "player_attributes": session.query(PlayerAttribute).count(),
                "player_attribute_versions": session.query(PlayerAttributeVersion).count(),
            }
        finally:
            session.close()

    def test_dry_run_validates_without_committing(self):
        report = run_import(
            workbook_path=self.workbook_path.name,
            attributes_csv_path=self.attributes_csv_path.name,
            root_dir=self.root_dir,
            target_engine=self.engine,
            dry_run=True,
        )

        self.assertFalse(report.has_errors)
        self.assertFalse(report.committed)
        self.assertEqual(report.datasets["league_info"].created, len(DEFAULT_LEAGUE_INFO_VALUES))
        self.assertEqual(report.datasets["teams"].created, 3)
        self.assertEqual(report.datasets["players"].created, 2)
        self.assertEqual(report.datasets["player_attributes"].created, 2)
        self.assertEqual(report.datasets["team_cache_rebuild"].updated, 2)
        self.assertEqual(
            self.count_rows(),
            {"league_info": 0, "teams": 0, "players": 0, "player_attributes": 0, "player_attribute_versions": 0},
        )

    def test_default_cli_root_uses_project_root_for_auto_attribute_discovery(self):
        fake_runtime_dir = self.root_dir / "imports_runtime"
        fake_runtime_dir.mkdir(exist_ok=True)
        fake_persistence_file = fake_runtime_dir / "persistence.py"
        fake_persistence_file.touch()
        auto_discovery_attributes_path = self.root_dir / "测试球员属性.csv"
        self.attributes_csv_path.replace(auto_discovery_attributes_path)
        self.attributes_csv_path = auto_discovery_attributes_path

        with patch.dict("os.environ", {}, clear=False):
            with patch.object(runtime_persistence, "__file__", str(fake_persistence_file)):
                report = run_import(
                    workbook_path=self.workbook_path.name,
                    target_engine=self.engine,
                )

        self.assertFalse(report.has_errors)
        self.assertTrue(report.committed)
        self.assertEqual(Path(report.attributes_csv_path).name, self.attributes_csv_path.name)
        self.assertEqual(self.count_rows()["player_attributes"], 2)

    def test_real_import_is_idempotent(self):
        first_report = run_import(
            workbook_path=self.workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )

        self.assertFalse(first_report.has_errors)
        self.assertTrue(first_report.committed)
        self.assertEqual(
            self.count_rows(),
            {
                "league_info": len(DEFAULT_LEAGUE_INFO_VALUES),
                "teams": 3,
                "players": 2,
                "player_attributes": 2,
                "player_attribute_versions": 2,
            },
        )

        session = self.SessionLocal()
        try:
            visible_teams = session.query(Team).filter(Team.level != HIDDEN_SEA_TEAM_LEVEL).all()
            self.assertEqual(len(visible_teams), 2)
            self.assertTrue(all(team.stats_cache_refresh_mode == "full_recalc" for team in visible_teams))
            self.assertTrue(all(team.stats_cache_refresh_scopes == "roster,wage" for team in visible_teams))

            hidden_team = session.query(Team).filter(Team.name == HIDDEN_SEA_TEAM_NAME).one()
            self.assertEqual(hidden_team.level, HIDDEN_SEA_TEAM_LEVEL)
        finally:
            session.close()

        second_report = run_import(
            workbook_path=self.workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )

        self.assertFalse(second_report.has_errors)
        self.assertTrue(second_report.committed)
        self.assertEqual(second_report.datasets["league_info"].created, 0)
        self.assertEqual(second_report.datasets["league_info"].unchanged, len(DEFAULT_LEAGUE_INFO_VALUES))
        self.assertEqual(second_report.datasets["teams"].created, 0)
        self.assertEqual(second_report.datasets["teams"].updated, 0)
        self.assertEqual(second_report.datasets["teams"].unchanged, 3)
        self.assertEqual(second_report.datasets["players"].created, 0)
        self.assertEqual(second_report.datasets["players"].updated, 0)
        self.assertEqual(second_report.datasets["players"].unchanged, 2)
        self.assertEqual(second_report.datasets["player_attributes"].created, 0)
        self.assertEqual(second_report.datasets["player_attributes"].updated, 0)
        self.assertEqual(second_report.datasets["player_attributes"].unchanged, 2)
        self.assertEqual(
            self.count_rows(),
            {
                "league_info": len(DEFAULT_LEAGUE_INFO_VALUES),
                "teams": 3,
                "players": 2,
                "player_attributes": 2,
                "player_attribute_versions": 2,
            },
        )

    def test_validation_errors_rollback_all_changes(self):
        bad_workbook_path = self.root_dir / "bad_fixture.xlsx"
        write_workbook(bad_workbook_path, include_position_column=False)

        report = run_import(
            workbook_path=bad_workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )

        self.assertTrue(report.has_errors)
        self.assertFalse(report.committed)
        self.assertTrue(report.datasets["players"].errors)
        self.assertIn("缺少列", report.datasets["players"].errors[0])
        self.assertEqual(
            self.count_rows(),
            {"league_info": 0, "teams": 0, "players": 0, "player_attributes": 0, "player_attribute_versions": 0},
        )

    def test_mapping_sheet_backfills_position_and_normalizes_team_aliases(self):
        fallback_workbook_path = self.root_dir / "mapping_fallback.xlsx"
        write_workbook(
            fallback_workbook_path,
            team_rows=[
                {"info_key": "届数", "info_value": 85, "name": "Bournemouth", "manager": "Coach A", "level": "超级", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
                {"info_key": "成长年龄上限", "info_value": 24, "name": "As Roma", "manager": "Coach B", "level": "甲级", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
            ],
            player_rows=[
                {"uid": 10, "name": "王五", "age": 20, "initial_ca": 120, "ca": 123, "pa": 150, "position": "", "nationality": "CN", "club_name": "Bournemouth", "team_name": "AFC Bournemouth"},
                {"uid": 11, "name": "赵六", "age": 21, "initial_ca": 118, "ca": 121, "pa": 148, "position": "", "nationality": "CN", "club_name": "Associazione Sportiva Roma"},
            ],
            mapping_rows=[
                {"uid": 10, "name": "王五", "age": 20, "ca": 123, "pa": 150, "nationality": "CN", "team_name": "AFC Bournemouth", "position": "MC"},
                {"uid": 11, "name": "赵六", "age": 21, "ca": 121, "pa": 148, "nationality": "CN", "team_name": "Associazione Sportiva Roma", "position": "GK"},
            ],
        )

        report = run_import(
            workbook_path=fallback_workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
            strict_mode=False,
        )

        self.assertFalse(report.has_errors)
        self.assertTrue(report.committed)
        self.assertEqual(report.datasets["players"].details["player_team_overrides"], 2)
        self.assertEqual(report.datasets["players"].details["team_alias_hits"], 2)

        session = self.SessionLocal()
        try:
            players = session.query(Player).order_by(Player.uid).all()
            self.assertEqual([player.team_name for player in players], ["Bournemouth", "As Roma"])
            self.assertTrue(all(player.team_id for player in players))
            self.assertEqual([player.position for player in players], ["MC", "GK"])
        finally:
            session.close()

    def test_falls_back_to_club_when_explicit_league_team_is_not_in_league(self):
        fallback_workbook_path = self.root_dir / "club_fallback.xlsx"
        write_workbook(
            fallback_workbook_path,
            team_rows=[
                {"info_key": "届数", "info_value": 85, "name": "Eintracht Frankfurt", "manager": "Coach A", "level": "超级", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
                {"info_key": "成长年龄上限", "info_value": 24, "name": "Beta FC", "manager": "Coach B", "level": "甲级", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
            ],
            player_rows=[
                {"uid": 21, "name": "阿德莫拉", "age": 27, "initial_ca": 160, "ca": 160, "pa": 164, "position": "ST", "nationality": "NG", "club_name": "Eintracht Frankfurt", "league_team_name": "Fenerbahçe"},
            ],
            mapping_rows=[],
        )

        report = run_import(
            workbook_path=fallback_workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
            strict_mode=False,
        )

        self.assertFalse(report.has_errors)
        self.assertTrue(report.committed)
        self.assertEqual(report.datasets["players"].details["club_team_fallback_hits"], 1)

        session = self.SessionLocal()
        try:
            player = session.query(Player).filter(Player.uid == 21).one()
            self.assertEqual(player.team_name, "Eintracht Frankfurt")
            self.assertIsNotNone(player.team_id)
        finally:
            session.close()

    def test_strict_mode_reports_missing_position_and_team_mismatch_clearly(self):
        strict_workbook_path = self.root_dir / "strict_failure.xlsx"
        write_workbook(
            strict_workbook_path,
            team_rows=[
                {"info_key": "届数", "info_value": 85, "name": "Bournemouth", "manager": "Coach A", "level": "超级", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
                {"info_key": "成长年龄上限", "info_value": 24, "name": "As Roma", "manager": "Coach B", "level": "甲级", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
            ],
            player_rows=[
                {"uid": 31, "name": "王五", "age": 20, "initial_ca": 120, "ca": 123, "pa": 150, "position": "", "nationality": "CN", "club_name": "Bournemouth", "team_name": "AFC Bournemouth"},
                {"uid": 32, "name": "赵六", "age": 21, "initial_ca": 118, "ca": 121, "pa": 148, "position": "GK", "nationality": "CN", "club_name": "Associazione Sportiva Roma"},
            ],
            mapping_rows=[
                {"uid": 31, "name": "王五", "age": 20, "ca": 123, "pa": 150, "nationality": "CN", "team_name": "Bournemouth", "position": "MC"},
                {"uid": 32, "name": "赵六", "age": 21, "ca": 121, "pa": 148, "nationality": "CN", "team_name": "As Roma", "position": "GK"},
            ],
        )

        report = run_import(
            workbook_path=strict_workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )

        self.assertTrue(report.has_errors)
        self.assertFalse(report.committed)
        self.assertTrue(report.strict_mode)
        self.assertEqual(report.datasets["players"].details["error_counts"]["missing_position"], 1)
        self.assertEqual(report.datasets["players"].details["error_counts"]["team_name_mismatch"], 1)
        self.assertEqual(
            report.datasets["players"].details["strict_mode_guidance"],
            {
                "position_rule": "联赛名单.位置 必须非空",
                "team_rule": "联赛名单中的球队名必须与 信息总览.球队名 完全一致",
            },
        )
        mismatch_sample = report.datasets["players"].details["error_samples"]["team_name_mismatch"][0]
        self.assertEqual(mismatch_sample["provided_team_name"], "Associazione Sportiva Roma")
        self.assertEqual(mismatch_sample["club_team_name"], "Associazione Sportiva Roma")
        self.assertEqual(
            self.count_rows(),
            {"league_info": 0, "teams": 0, "players": 0, "player_attributes": 0, "player_attribute_versions": 0},
        )

    def test_real_import_removes_stale_orphan_visible_teams(self):
        session = self.SessionLocal()
        try:
            session.add(Team(name="Legacy Roma", manager="Old", level="鐢茬骇", wage=0))
            session.commit()
        finally:
            session.close()

        report = run_import(
            workbook_path=self.workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )

        self.assertFalse(report.has_errors)
        self.assertTrue(report.committed)
        self.assertEqual(report.datasets["team_cleanup"].details["removed_count"], 1)
        self.assertEqual(report.datasets["team_cleanup"].details["removed_teams"], ["Legacy Roma"])

        session = self.SessionLocal()
        try:
            self.assertIsNone(session.query(Team).filter(Team.name == "Legacy Roma").first())
            self.assertEqual(session.query(Team).count(), 3)
        finally:
            session.close()

    def test_real_import_full_sync_removes_players_missing_from_new_roster(self):
        initial_workbook_path = self.root_dir / "full_sync_initial.xlsx"
        write_workbook(
            initial_workbook_path,
            team_rows=[
                {"info_key": "灞婃暟", "info_value": 85, "name": "Alpha FC", "manager": "Coach A", "level": "瓒呯骇", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
                {"info_key": "鎴愰暱骞撮緞涓婇檺", "info_value": 24, "name": "Beta FC", "manager": "Coach B", "level": "鐢茬骇", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
                {"info_key": "鏈増棣栧眾", "info_value": 84, "name": "Gamma FC", "manager": "Coach C", "level": "涔欑骇", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
            ],
            player_rows=[
                {"uid": 1, "name": "寮犱笁", "age": 20, "initial_ca": 120, "ca": 125, "pa": 150, "position": "MC", "nationality": "CN", "club_name": "Alpha FC"},
                {"uid": 2, "name": "鏉庡洓", "age": 21, "initial_ca": 118, "ca": 121, "pa": 148, "position": "GK", "nationality": "CN", "club_name": "Alpha FC"},
                {"uid": 3, "name": "鐜嬩簲", "age": 22, "initial_ca": 116, "ca": 119, "pa": 144, "position": "ST", "nationality": "CN", "club_name": "Gamma FC"},
            ],
        )
        updated_workbook_path = self.root_dir / "full_sync_updated.xlsx"
        write_workbook(
            updated_workbook_path,
            team_rows=[
                {"info_key": "灞婃暟", "info_value": 85, "name": "Alpha FC", "manager": "Coach A", "level": "瓒呯骇", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
                {"info_key": "鎴愰暱骞撮緞涓婇檺", "info_value": 24, "name": "Beta FC", "manager": "Coach B", "level": "鐢茬骇", "extra_wage": 0.0, "after_tax": 0, "notes": ""},
            ],
            player_rows=[
                {"uid": 1, "name": "寮犱笁", "age": 20, "initial_ca": 120, "ca": 125, "pa": 150, "position": "MC", "nationality": "CN", "club_name": "Alpha FC"},
            ],
        )

        first_report = run_import(
            workbook_path=initial_workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )
        self.assertFalse(first_report.has_errors)
        self.assertTrue(first_report.committed)

        second_report = run_import(
            workbook_path=updated_workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )

        self.assertFalse(second_report.has_errors)
        self.assertTrue(second_report.committed)
        self.assertEqual(second_report.datasets["player_sync_cleanup"].details["removed_count"], 2)
        self.assertEqual(
            [sample["uid"] for sample in second_report.datasets["player_sync_cleanup"].details["removed_players_sample"]],
            [2, 3],
        )
        self.assertEqual(
            second_report.datasets["player_sync_cleanup"].details["removed_team_breakdown"],
            {"Alpha FC": 1, "Gamma FC": 1},
        )
        self.assertEqual(second_report.datasets["team_cleanup"].details["removed_count"], 1)
        self.assertEqual(second_report.datasets["team_cleanup"].details["removed_teams"], ["Gamma FC"])

        session = self.SessionLocal()
        try:
            remaining_player = session.query(Player).filter(Player.uid == 1).one()
            alpha_team = session.query(Team).filter(Team.name == "Alpha FC").one()
            self.assertEqual(session.query(Player).count(), 1)
            self.assertIsNone(session.query(Player).filter(Player.uid == 2).first())
            self.assertIsNone(session.query(Player).filter(Player.uid == 3).first())
            self.assertEqual(alpha_team.team_size, 1)
            self.assertEqual(alpha_team.gk_count, 0)
            self.assertAlmostEqual(alpha_team.wage, remaining_player.wage)
            self.assertIsNone(session.query(Team).filter(Team.name == "Gamma FC").first())
            self.assertEqual(session.query(Team).count(), 3)
        finally:
            session.close()

    def test_decode_player_habit_value_keeps_suspicious_numeric_codes_unparsed(self):
        decoded = decode_player_habit_value(2251834173423610)

        self.assertIsNotNone(decoded)
        self.assertEqual(decoded["decoded_text"], "")
        self.assertEqual(decoded["raw_code"], "2251834173423610")
        self.assertEqual(decoded["high_bits"], "")
        self.assertFalse(decoded["is_reliable"])

    def test_import_supports_2620_attribute_workbook_and_derives_radar_metrics(self):
        attribute_workbook_path = self.root_dir / "2620_fixture.xlsx"
        write_attributes_xlsx_2620(attribute_workbook_path)

        report = run_import(
            workbook_path=self.workbook_path,
            attributes_csv_path=attribute_workbook_path,
            target_engine=self.engine,
        )

        self.assertFalse(report.has_errors)
        self.assertTrue(report.committed)
        attribute_summary = report.datasets["player_attributes"]
        self.assertEqual(attribute_summary.details["data_version"], "2620")
        self.assertEqual(attribute_summary.details["source_format"], "xlsx")
        self.assertIn("防守", attribute_summary.details["derived_columns"])
        self.assertEqual(attribute_summary.details["negative_pa_override_count"], 2)
        self.assertEqual(attribute_summary.details["player_habit_numeric_rows"], 2)
        self.assertEqual(attribute_summary.details["player_habit_decoded_rows"], 1)
        self.assertEqual(attribute_summary.details["player_habit_high_bits_rows"], 1)
        self.assertEqual(attribute_summary.details["player_habit_unresolved_rows"], 1)
        self.assertEqual(attribute_summary.details["header_renames"]["停球"], "接球")
        self.assertEqual(attribute_summary.details["header_renames"]["指挥防守"], "沟通")

        session = self.SessionLocal()
        try:
            outfield = session.query(PlayerAttribute).filter(PlayerAttribute.uid == 1).one()
            goalkeeper = session.query(PlayerAttribute).filter(PlayerAttribute.uid == 2).one()
            outfield_versioned = (
                session.query(PlayerAttributeVersion)
                .filter(
                    PlayerAttributeVersion.uid == 1,
                    PlayerAttributeVersion.data_version == "2620",
                )
                .one()
            )
            goalkeeper_versioned = (
                session.query(PlayerAttributeVersion)
                .filter(
                    PlayerAttributeVersion.uid == 2,
                    PlayerAttributeVersion.data_version == "2620",
                )
                .one()
            )

            self.assertEqual(outfield.pa, -10)
            self.assertEqual(outfield_versioned.pa, -10)
            self.assertEqual(outfield.first_touch, 15)
            self.assertEqual(outfield_versioned.first_touch, 15)
            self.assertEqual(outfield.communication, 6)
            self.assertEqual(outfield.player_habits, "经常尝试传身后球\n角度刁钻的射门\n回撤拿球")
            self.assertEqual(outfield.player_habits_raw_code, "525440")
            self.assertEqual(outfield.player_habits_high_bits, "")
            self.assertEqual(outfield_versioned.player_habits, outfield.player_habits)
            self.assertEqual(outfield_versioned.player_habits_raw_code, "525440")
            self.assertEqual(outfield_versioned.player_habits_high_bits, "")
            self.assertAlmostEqual(outfield.radar_defense, 9.0)
            self.assertAlmostEqual(outfield.radar_physical, 12.0)
            self.assertAlmostEqual(outfield.radar_speed, 12.0)
            self.assertAlmostEqual(outfield.radar_creativity, 13.0)
            self.assertAlmostEqual(outfield.radar_attack, 11.0)
            self.assertAlmostEqual(outfield.radar_technical, 15.0)
            self.assertAlmostEqual(outfield.radar_aerial, 10.0)
            self.assertAlmostEqual(outfield.radar_mental, 12.5)

            self.assertEqual(goalkeeper.pa, -10)
            self.assertEqual(goalkeeper_versioned.pa, -10)
            self.assertEqual(goalkeeper.communication, 12)
            self.assertEqual(goalkeeper.throwing, 11)
            self.assertEqual(goalkeeper.player_habits, "")
            self.assertEqual(goalkeeper.player_habits_raw_code, str(2 ** 62))
            self.assertEqual(goalkeeper.player_habits_high_bits, str(2 ** 62))
            self.assertEqual(goalkeeper_versioned.player_habits, "")
            self.assertEqual(goalkeeper_versioned.player_habits_raw_code, str(2 ** 62))
            self.assertEqual(goalkeeper_versioned.player_habits_high_bits, str(2 ** 62))
            self.assertAlmostEqual(goalkeeper.radar_gk_shot_stopping, 17.0)
            self.assertAlmostEqual(goalkeeper.radar_gk_command, 13.0)
            self.assertAlmostEqual(goalkeeper.radar_gk_aerial, 16.0)
            self.assertAlmostEqual(goalkeeper.radar_gk_kicking, 12.0)
            self.assertAlmostEqual(goalkeeper.radar_gk_physical, 11.0)
            self.assertAlmostEqual(goalkeeper.radar_gk_speed, 9.5)
            self.assertAlmostEqual(goalkeeper.radar_gk_mental, 13.3333333333, places=6)
            self.assertAlmostEqual(goalkeeper.radar_gk_eccentricity, 5.0)
        finally:
            session.close()

    def test_import_preserves_2600_and_2620_attribute_versions_side_by_side(self):
        first_report = run_import(
            workbook_path=self.workbook_path,
            attributes_csv_path=self.attributes_csv_path,
            target_engine=self.engine,
        )
        self.assertFalse(first_report.has_errors)
        self.assertTrue(first_report.committed)
        self.assertEqual(first_report.datasets["player_attributes"].details["data_version"], "2600")

        attribute_workbook_path = self.root_dir / "2620_fixture.xlsx"
        write_attributes_xlsx_2620(attribute_workbook_path)
        second_report = run_import(
            workbook_path=self.workbook_path,
            attributes_csv_path=attribute_workbook_path,
            target_engine=self.engine,
        )

        self.assertFalse(second_report.has_errors)
        self.assertTrue(second_report.committed)
        self.assertEqual(second_report.datasets["player_attributes"].details["data_version"], "2620")

        session = self.SessionLocal()
        try:
            self.assertEqual(session.query(PlayerAttribute).count(), 2)
            self.assertEqual(session.query(PlayerAttributeVersion).count(), 4)

            legacy_outfield = session.query(PlayerAttribute).filter(PlayerAttribute.uid == 1).one()
            outfield_2600 = (
                session.query(PlayerAttributeVersion)
                .filter(
                    PlayerAttributeVersion.uid == 1,
                    PlayerAttributeVersion.data_version == "2600",
                )
                .one()
            )
            outfield_2620 = (
                session.query(PlayerAttributeVersion)
                .filter(
                    PlayerAttributeVersion.uid == 1,
                    PlayerAttributeVersion.data_version == "2620",
                )
                .one()
            )

            self.assertEqual(outfield_2600.corner, 12)
            self.assertEqual(outfield_2620.first_touch, 15)
            self.assertEqual(legacy_outfield.first_touch, 15)

            version_summary = read_service.get_attribute_versions(session)
            self.assertEqual(version_summary.available_versions, ["2620", "2600"])
            self.assertEqual(version_summary.default_version, "2620")

            detail_2600 = read_service.get_player_attribute_detail(session, 1, data_version="2600")
            detail_2620 = read_service.get_player_attribute_detail(session, 1, data_version="2620")
            self.assertIsNotNone(detail_2600)
            self.assertIsNotNone(detail_2620)
            self.assertEqual(detail_2600.data_version, "2600")
            self.assertEqual(detail_2620.data_version, "2620")
            self.assertEqual(detail_2600.corner, 12)
            self.assertEqual(detail_2620.first_touch, 15)
            self.assertEqual(detail_2620.player_habits, "经常尝试传身后球\n角度刁钻的射门\n回撤拿球")
            self.assertEqual(detail_2620.player_habits_raw_code, "525440")
            self.assertEqual(detail_2620.player_habits_high_bits, "")
        finally:
            session.close()


if __name__ == "__main__":
    unittest.main()
