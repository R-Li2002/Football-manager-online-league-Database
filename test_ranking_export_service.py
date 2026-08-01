import io
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base
from models import RankingSeed, Team
from services.ranking_export_service import build_ranking_excel


class RankingExportServiceTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(self.engine)
        self.db = sessionmaker(bind=self.engine)()
        self.db.add_all([
            Team(id=1, name="FC Schalke 04", level="超级"),
            Team(id=2, name="R. Madrid", level="甲级"),
        ])
        self.db.add_all([
            RankingSeed(team_id=1, team_name="FC Schalke 04", base_points=1000.736, matches=95, wins=58, draws=0, losses=37),
            RankingSeed(team_id=2, team_name="R. Madrid", base_points=1975.011, matches=13, wins=12, draws=0, losses=1),
        ])
        self.db.commit()
        self.temp_dir = TemporaryDirectory()

    def tearDown(self):
        self.temp_dir.cleanup()
        self.db.close()
        self.engine.dispose()

    def build_template(self):
        workbook = Workbook()
        current = workbook.active
        current.title = "原始 (3)"
        current.cell(2, 10, 291)
        old = workbook.create_sheet("原始")
        old.cell(2, 10, 12)
        for sheet in (current, old):
            for row_no in range(7, 67):
                sheet.cell(row_no, 1, row_no - 6)
                sheet.cell(row_no, 2, f"球队{row_no}")
                sheet.cell(row_no, 3, 1000)
                sheet.cell(row_no, 4, f"=E{row_no}+F{row_no}")
                sheet.cell(row_no, 5, 0)
                sheet.cell(row_no, 6, 0)
                sheet.cell(row_no, 7, f'=IFERROR(E{row_no}/D{row_no},"")')
                sheet.cell(row_no, 8, f"=C{row_no}+D{row_no}*20")
        path = Path(self.temp_dir.name) / "ranking.xlsx"
        workbook.save(path)
        return path

    def test_exports_standard_names_and_keeps_formulas(self):
        output, filename = build_ranking_excel(self.db, self.build_template())
        self.assertTrue(filename.endswith(".xlsx"))
        payload = output.getvalue()
        workbook = load_workbook(io.BytesIO(payload), data_only=False)
        sheet = workbook["原始 (3)"]

        self.assertEqual(sheet["B7"].value, "FC Schalke 04")
        self.assertEqual(sheet["B8"].value, "R. Madrid")
        self.assertEqual(sheet["C7"].value, 1000.736)
        self.assertEqual(sheet["D7"].value, "=E7+F7+0")
        self.assertEqual(sheet["G7"].value, '=IFERROR(E7/D7,"")')
        self.assertEqual(sheet["H7"].value, "=C7+D7*20")
        self.assertIsNone(sheet["B9"].value)
        self.assertEqual(workbook["原始"]["B7"].value, "球队7")
        self.assertEqual(workbook.calculation.calcMode, "auto")
        self.assertTrue(workbook.calculation.fullCalcOnLoad)
        self.assertTrue(workbook.calculation.forceFullCalc)
        self.assertEqual(workbook.calculation.calcId, 0)

        with ZipFile(io.BytesIO(payload)) as archive:
            self.assertNotIn("xl/calcChain.xml", archive.namelist())
            sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
            self.assertIn("<f>E7+F7+0</f>", sheet_xml)
            self.assertIn("<f>C7+D7*20</f>", sheet_xml)


if __name__ == "__main__":
    unittest.main()
