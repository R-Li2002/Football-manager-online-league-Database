import unittest
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base
from models import Match, Team
from services import export_service


class StandingsExportServiceTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.db = Session()
        alpha = Team(name="Alpha", level="超级", manager="Alpha Boss")
        beta = Team(name="Beta", level="超级", manager="Beta Boss")
        hidden = Team(name="Hidden", level="隐藏")
        self.db.add_all([alpha, beta, hidden])
        self.db.flush()
        self.db.add_all(
            [
                Match(
                    season_label="S1",
                    level="超级",
                    round_no=1,
                    home_team_id=alpha.id,
                    home_team_name=alpha.name,
                    away_team_id=beta.id,
                    away_team_name=beta.name,
                    home_score=2,
                    away_score=1,
                    status="played",
                    match_date=datetime(2026, 7, 29, 20, 0),
                ),
                Match(
                    season_label="S1",
                    level="超级",
                    round_no=2,
                    home_team_id=beta.id,
                    home_team_name=beta.name,
                    away_team_id=alpha.id,
                    away_team_name=alpha.name,
                    status="scheduled",
                ),
            ]
        )
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def test_workbook_has_formula_standings_and_editable_schedule(self):
        output, filename = export_service.build_standings_excel(self.db, "超级")

        self.assertRegex(filename, r"^HEIGO_S_standings_.*\.xlsx$")
        workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
        self.assertEqual(workbook.sheetnames, ["主积分榜", "赛程"])

        standings = workbook["主积分榜"]
        schedule = workbook["赛程"]
        self.assertEqual([standings.cell(3, column).value for column in range(1, 12)], [
            "排名", "球队", "主教练", "场次", "胜", "平", "负", "进球", "失球", "净胜球", "积分",
        ])
        self.assertEqual(standings["B4"].value, "Alpha")
        self.assertTrue(standings["A4"].value.startswith("=1+COUNTIF"))
        self.assertIn("'赛程'!$B$4:$B$1003", standings["D4"].value)
        self.assertIn("SUMPRODUCT", standings["E4"].value)
        self.assertIn('="客队判负"', standings["E4"].value)
        self.assertNotIn('="主队判负"', standings["E4"].value)
        self.assertIn('="主队判负"', standings["F4"].value)
        self.assertEqual(standings["K4"].value, "=E4*3+F4")

        self.assertEqual([schedule.cell(3, column).value for column in range(1, 9)], [
            "轮次", "主队", "客队", "主队比分", "客队比分", "状态", "比赛日期", "备注",
        ])
        self.assertEqual(schedule["F4"].value, "已赛")
        self.assertEqual(schedule["F5"].value, "未赛")
        self.assertFalse(schedule["D4"].protection.locked)
        self.assertFalse(schedule["F5"].protection.locked)
        self.assertEqual(len(schedule.data_validations.dataValidation), 1)
        self.assertIn("主队判负", schedule.data_validations.dataValidation[0].formula1)
        self.assertEqual(standings["A1"].fill.fgColor.rgb, "00F2EEFF")
        self.assertEqual(standings["A3"].fill.fgColor.rgb, "00F8F9FA")
        self.assertEqual(schedule["D4"].fill.fgColor.rgb, "00EAF7EF")
        self.assertFalse(standings.sheet_view.showGridLines)

    def test_invalid_level_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "仅支持超级、甲级、乙级"):
            export_service.build_standings_excel(self.db, "冠军杯")


if __name__ == "__main__":
    unittest.main()
