import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database import Base
from models import Team
from services.ranking_import_service import preview_ranking_workbook


class RankingImportServiceTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(self.engine)
        self.db = sessionmaker(bind=self.engine)()
        self.db.add_all([
            Team(id=1, name="Tottenham Hotspur", level="超级"),
            Team(id=2, name="FK Bodø/Glimt", level="甲级"),
            Team(id=3, name="West Ham United", level="乙级"),
        ])
        self.db.commit()
        self.temp_dir = TemporaryDirectory()

    def tearDown(self):
        self.temp_dir.cleanup()
        self.db.close()
        self.engine.dispose()

    def build_workbook(self):
        workbook = Workbook()
        old = workbook.active
        old.title = "原始"
        old.cell(2, 10, 12)
        old.cell(7, 2, "热刺")
        old.cell(7, 3, 999)

        current = workbook.create_sheet("原始 (3)")
        current.cell(2, 10, 243)
        rows = [
            (7, "Tottenham", 1000, 0, 0, 0),
            (8, "热刺", 1175.5, 6, 4, 2),
            (9, "Glimt", 1080, 2, 1, 1),
            (10, "Celtic", 1200, 4, 3, 1),
            (11, "埃弗顿19", 1000, 0, 0, 0),
        ]
        for row_no, name, points, matches, wins, losses in rows:
            current.cell(row_no, 2, name)
            current.cell(row_no, 3, points)
            current.cell(row_no, 4, matches)
            current.cell(row_no, 5, wins)
            current.cell(row_no, 6, losses)
            current.cell(row_no, 8, points + matches * 20)
        path = Path(self.temp_dir.name) / "ranking.xlsx"
        workbook.save(path)
        return path

    def test_selects_latest_sheet_maps_aliases_and_initializes_missing_team(self):
        report = preview_ranking_workbook(self.db, self.build_workbook())

        self.assertEqual(report["sheet"], "原始 (3)")
        self.assertEqual(report["cutoff"], 243)
        self.assertEqual(report["team_count"], 3)
        self.assertEqual(report["mapped_count"], 2)
        self.assertEqual(report["initialized_count"], 1)
        self.assertEqual(report["missing_current_teams"], ["West Ham United"])

        spurs = next(row for row in report["rows"] if row["team_name"] == "Tottenham Hotspur")
        self.assertEqual((spurs["source_name"], spurs["base_points"], spurs["matches"]), ("热刺", 1175.5, 6))
        self.assertEqual(report["duplicates"][0]["standard_name"], "Tottenham Hotspur")
        self.assertEqual({item["source_name"] for item in report["skipped"]}, {"Celtic", "埃弗顿19"})


if __name__ == "__main__":
    unittest.main()
