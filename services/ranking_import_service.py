from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import RankingMatch, RankingSeed, Team
from services.ranking_service import APPEARANCE_BONUS, INITIAL_POINTS, LEAGUE_LEVELS

TEAM_NAME_ALIASES = {
    "Schalke 04": "FC Schalke 04",
    "布莱顿": "Brighton & Hove Albion",
    "毕尔巴勒": "A. Bilbao",
    "皇马": "R. Madrid",
    "热刺": "Tottenham Hotspur",
    "河床": "River Plate",
    "切尔西": "Chelsea",
    "阿贾克斯": "AFC Ajax",
    "罗马": "Associazione Sportiva Roma",
    "葡体": "Sporting Clube de Portugal",
    "森林": "Nottingham Forest",
    "埃弗顿": "Everton",
    "曼城": "Manchester City",
    "巴黎": "Paris Saint-Germain",
    "莱比锡": "RB Leipzig",
    "里昂": "Olympique Lyonnais",
    "博德": "FK Bodø/Glimt",
    "考文垂": "Coventry City",
    "利物浦": "Liverpool",
    "本菲卡": "Sport Lisboa e Benfica",
    "桑德兰": "Sunderland",
    "法兰克福": "Eintracht Frankfurt",
    "维拉": "Aston Villa",
    "尤文": "Juventus",
    "巴萨": "Barcelona",
    "那不勒斯": "Napoli",
    "斯特拉斯堡": "RC Strasbourg Alsace",
    "伯恩茅斯": "AFC Bournemouth",
    "Bayer 04": "Bayer 04 Leverkusen",
    "Boca": "Club Atlético Boca Juniors",
    "FC Bayern": "FC Bayern München",
    "Frankfurt": "Eintracht Frankfurt",
    "Leicester": "Leicester City",
    "Man UFC": "Manchester United",
    "Newcastle": "Newcastle United",
    "OM": "Olympique de Marseille",
    "Glimt": "FK Bodø/Glimt",
    "Tottenham": "Tottenham Hotspur",
    "Como": "Como 1907",
    "利兹联": "Leeds United",
    "勒沃库森": "Bayer 04 Leverkusen",
    "马竞": "A. Madrid",
    "雷恩": "Stade Rennais F.C.",
    "巨龙": "Oriental Dragon",
    "费耶诺德": "Feyenoord Rotterdam",
    "阿森纳": "Arsenal",
}


def _numeric(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _select_sheet(workbook):
    candidates = []
    for worksheet in workbook.worksheets:
        cutoff = _numeric(worksheet.cell(2, 10).value, -1)
        candidates.append((cutoff, worksheet.max_row, worksheet.title, worksheet))
    return max(candidates, key=lambda item: (item[0], item[1], item[2]))[3]


def preview_ranking_workbook(db: Session, workbook_path: str | Path) -> dict[str, Any]:
    source = Path(workbook_path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = _select_sheet(workbook)
    teams = db.query(Team).filter(Team.level.in_(LEAGUE_LEVELS)).order_by(Team.name).all()
    teams_by_name = {team.name: team for team in teams}
    candidates: dict[str, list[dict[str, Any]]] = {}
    skipped = []

    for row_no in range(7, worksheet.max_row + 1):
        raw_name = str(worksheet.cell(row_no, 2).value or "").strip()
        if not raw_name:
            continue
        standard_name = TEAM_NAME_ALIASES.get(raw_name, raw_name if raw_name in teams_by_name else None)
        item = {
            "source_row": row_no,
            "source_name": raw_name,
            "standard_name": standard_name,
            "source_rank": int(_numeric(worksheet.cell(row_no, 1).value, 0)),
            "base_points": _numeric(worksheet.cell(row_no, 3).value, INITIAL_POINTS),
            "matches": int(_numeric(worksheet.cell(row_no, 4).value, 0)),
            "wins": int(_numeric(worksheet.cell(row_no, 5).value, 0)),
            "losses": int(_numeric(worksheet.cell(row_no, 6).value, 0)),
            "source_total_points": _numeric(worksheet.cell(row_no, 8).value, INITIAL_POINTS),
        }
        item["draws"] = max(0, item["matches"] - item["wins"] - item["losses"])
        item["calculated_total_points"] = item["base_points"] + item["matches"] * APPEARANCE_BONUS
        if not standard_name:
            skipped.append({**item, "reason": "not_current_team"})
            continue
        candidates.setdefault(standard_name, []).append(item)

    selected = {}
    duplicates = []
    for standard_name, items in candidates.items():
        ordered = sorted(
            items,
            key=lambda item: (item["matches"], abs(item["base_points"] - INITIAL_POINTS), -item["source_row"]),
            reverse=True,
        )
        selected[standard_name] = ordered[0]
        if len(ordered) > 1:
            duplicates.append({"standard_name": standard_name, "selected_row": ordered[0]["source_row"], "ignored_rows": [item["source_row"] for item in ordered[1:]]})

    rows = []
    missing = []
    for team in teams:
        item = selected.get(team.name)
        if not item:
            item = {
                "source_row": None,
                "source_name": None,
                "standard_name": team.name,
                "source_rank": 0,
                "base_points": INITIAL_POINTS,
                "matches": 0,
                "wins": 0,
                "draws": 0,
                "losses": 0,
                "source_total_points": INITIAL_POINTS,
                "calculated_total_points": INITIAL_POINTS,
            }
            missing.append(team.name)
        rows.append({"team_id": team.id, "team_name": team.name, "level": team.level, **item})

    rows.sort(key=lambda item: (-item["calculated_total_points"], -item["base_points"], item["team_name"]))
    return {
        "workbook": source.name,
        "sheet": worksheet.title,
        "cutoff": worksheet.cell(2, 10).value,
        "team_count": len(rows),
        "mapped_count": len(rows) - len(missing),
        "initialized_count": len(missing),
        "missing_current_teams": missing,
        "duplicates": duplicates,
        "skipped": skipped,
        "rows": rows,
    }


def import_ranking_workbook(db: Session, workbook_path: str | Path, *, force: bool = False) -> dict[str, Any]:
    if db.query(RankingMatch).count() and not force:
        raise ValueError("已有排位比赛，拒绝覆盖导入基线；如确需重置请显式使用 --force")
    report = preview_ranking_workbook(db, workbook_path)
    now = datetime.now()
    existing = {row.team_id: row for row in db.query(RankingSeed).all()}
    for item in report["rows"]:
        seed = existing.get(item["team_id"])
        if not seed:
            seed = RankingSeed(team_id=item["team_id"])
            db.add(seed)
        seed.team_name = item["team_name"]
        seed.base_points = item["base_points"]
        seed.matches = item["matches"]
        seed.wins = item["wins"]
        seed.draws = item["draws"]
        seed.losses = item["losses"]
        seed.source_name = item["source_name"]
        seed.source_row = item["source_row"]
        seed.imported_at = now
        seed.updated_at = now
    db.commit()
    return report
