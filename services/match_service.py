from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import Match, MatchPlayerEvent, Player, Team
from repositories.match_repository import (
    delete_match_events,
    delete_matches_not_in_keys,
    find_match_by_fixture,
    get_match_by_id,
    list_match_events,
    list_matches,
)
from repositories.player_repository import get_players_by_team_name, get_team_players
from repositories.team_repository import get_team_by_id, get_team_by_name, list_visible_teams
from schemas_read import MatchPlayerEventResponse, MatchResponse, ScheduleResponse, StandingHistoryRoundResponse, StandingHistoryRowResponse, StandingHistoryTeamResponse, StandingRowResponse, StandingsHistoryResponse, StandingsPredictionSummaryResponse, StandingsResponse
from schemas_write import MatchBatchUpdateRequest, MatchPlayerEventUpdateItem, MatchUpdateRequest, ScheduleImportResponse
from services.admin_common import LogWriter, require_admin
from services import standings_prediction_service

LEVEL_ORDER = {"超级": 1, "甲级": 2, "乙级": 3}
VISIBLE_LEVEL = "隐藏"
FORFEIT_STATUSES = {"home_forfeit", "away_forfeit", "double_forfeit"}
MATCH_STATUSES = {"scheduled", "played", "postponed", "cancelled", *FORFEIT_STATUSES}
PLAYED_MATCH_STATUSES = {"played", *FORFEIT_STATUSES}
HISTORY_RESOLVED_MATCH_STATUSES = {*PLAYED_MATCH_STATUSES, "cancelled"}
MATCH_EVENT_TYPES = {"goal", "own_goal", "assist", "mvp"}
SCHEDULE_ROOT = Path("imports") / "schedules"
SCHEDULE_TEAM_ALIASES = {
    "A.Bilbao": "A. Bilbao",
    "Ajax": "AFC Ajax",
    "AS Roma": "Associazione Sportiva Roma",
    "At Madrid": "A. Madrid",
    "Bayer 04": "Bayer 04 Leverkusen",
    "Bayern": "FC Bayern München",
    "Benfica": "Sport Lisboa e Benfica",
    "Boca": "Club Atlético Boca Juniors",
    "Bournemouth": "AFC Bournemouth",
    "Brighton": "Brighton & Hove Albion",
    "Como": "Como 1907",
    "Coventry": "Coventry City",
    "Dortmund": "Borussia Dortmund",
    "Frankfurt": "Eintracht Frankfurt",
    "Heidenheim": "FC Heidenheim 1846",
    "Leeds": "Leeds United",
    "Leicester": "Leicester City",
    "Man Utd": "Manchester United",
    "Newcastle": "Newcastle United",
    "Nottm Forest": "Nottingham Forest",
    "OL": "Olympique Lyonnais",
    "OM": "Olympique de Marseille",
    "PSG": "Paris Saint-Germain",
    "R.Madrid": "R. Madrid",
    "RBL": "RB Leipzig",
    "Schalke": "FC Schalke 04",
    "Sheff Utd": "Sheffield United",
    "Sporing CP": "Sporting Clube de Portugal",
    "Strasbourg": "RC Strasbourg Alsace",
    "Sturm Graz": "Sportklub Sturm Graz",
    "Talleres": "Club Atlético Talleres de Córdoba",
    "Tottenham": "Tottenham Hotspur",
    "West Ham": "West Ham United",
    "Wolves": "Wolverhampton Wanderers",
    "Zhejiang": "Oriental Dragon",
}


@dataclass(frozen=True)
class ParsedFixture:
    level: str
    round_no: int
    home_team_name: str
    away_team_name: str


def _normalize_cell(value: Any) -> str:
    return str(value or "").strip()


def _normalize_team_lookup_name(name: str) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        "",
        re.sub(
            r"\b(fc|cf|club|football|futebol|sport|sporting|association|associazione|olympique|de|of|the)\b",
            "",
            str(name or "").lower().replace("&", "and"),
        ),
    ).strip()


def _resolve_schedule_team(db: Session, team_id: int | None, team_name: str) -> Team | None:
    if team_id:
        team = get_team_by_id(db, team_id)
        if team:
            return team
    raw_name = str(team_name or "").strip()
    alias_name = SCHEDULE_TEAM_ALIASES.get(raw_name, raw_name)
    for candidate_name in (raw_name, alias_name):
        team = get_team_by_name(db, candidate_name)
        if team:
            return team
    normalized_names = {_normalize_team_lookup_name(raw_name), _normalize_team_lookup_name(alias_name)}
    normalized_names.discard("")
    for team in list_visible_teams(db, VISIBLE_LEVEL):
        if _normalize_team_lookup_name(team.name) in normalized_names:
            return team
    return None


def _normalize_level(sheet_title: str) -> str:
    title = sheet_title.strip()
    for level in LEVEL_ORDER:
        if level in title:
            return level
    return title.replace("赛程", "").strip() or title


def _parse_round_no(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and int(value) == value:
        return int(value)
    match = re.fullmatch(r"(?:第\s*)?(\d+)(?:\s*轮)?", str(value).strip())
    return int(match.group(1)) if match else None


def _parse_fixture_sheet(ws, level: str) -> list[ParsedFixture]:
    fixtures: list[ParsedFixture] = []
    active_rounds: dict[int, int] = {}

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        round_positions = {
            index: round_no
            for index, value in enumerate(cells)
            if (round_no := _parse_round_no(value)) is not None
            and _normalize_cell(cells[index - 1] if index > 0 else "") == ""
        }
        if round_positions:
            active_rounds.update(round_positions)
            continue

        for round_col, round_no in active_rounds.items():
            home = _normalize_cell(cells[round_col - 1] if round_col - 1 >= 0 and round_col - 1 < len(cells) else "")
            marker = _normalize_cell(cells[round_col] if round_col < len(cells) else "").lower()
            away = _normalize_cell(cells[round_col + 1] if round_col + 1 < len(cells) else "")
            if home and away and marker == "vs":
                fixtures.append(
                    ParsedFixture(
                        level=level,
                        round_no=round_no,
                        home_team_name=home,
                        away_team_name=away,
                    )
                )
    return fixtures


def find_latest_schedule_file(root: Path = SCHEDULE_ROOT) -> Path:
    candidates = [
        path
        for pattern in ("*.xlsx", "*.xlsm")
        for path in root.glob(pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise HTTPException(status_code=404, detail=f"未在 {root} 找到赛程 Excel 文件")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def parse_schedule_workbook(path: Path) -> list[ParsedFixture]:
    workbook = load_workbook(path, data_only=True)
    fixtures: list[ParsedFixture] = []
    for ws in workbook.worksheets:
        level = _normalize_level(ws.title)
        fixtures.extend(_parse_fixture_sheet(ws, level))
    if not fixtures:
        raise HTTPException(status_code=400, detail="赛程文件中没有识别到任何 主队 vs 客队 对阵")
    return fixtures


def import_schedule_file(
    db: Session,
    admin: str | None,
    write_to_log: LogWriter,
    path: str | Path,
) -> ScheduleImportResponse:
    operator = require_admin(admin)
    path = Path(path)
    fixtures = parse_schedule_workbook(path)
    fixture_keys = {(item.level, item.round_no, item.home_team_name, item.away_team_name) for item in fixtures}
    warnings: list[str] = []
    created = updated = unchanged = 0

    for item in fixtures:
        home_team = _resolve_schedule_team(db, None, item.home_team_name)
        away_team = _resolve_schedule_team(db, None, item.away_team_name)
        if not home_team:
            warnings.append(f"未匹配主队：{item.home_team_name}")
        if not away_team:
            warnings.append(f"未匹配客队：{item.away_team_name}")

        match = find_match_by_fixture(
            db,
            level=item.level,
            round_no=item.round_no,
            home_team_name=item.home_team_name,
            away_team_name=item.away_team_name,
        )
        if not match:
            db.add(
                Match(
                    season_label=path.stem,
                    level=item.level,
                    round_no=item.round_no,
                    home_team_id=home_team.id if home_team else None,
                    home_team_name=item.home_team_name,
                    away_team_id=away_team.id if away_team else None,
                    away_team_name=item.away_team_name,
                    status="scheduled",
                    source_file=str(path),
                    created_at=datetime.now(),
                    updated_at=datetime.now(),
                )
            )
            created += 1
            continue

        changed = False
        for field_name, value in (
            ("season_label", path.stem),
            ("home_team_id", home_team.id if home_team else None),
            ("away_team_id", away_team.id if away_team else None),
            ("source_file", str(path)),
        ):
            if getattr(match, field_name) != value:
                setattr(match, field_name, value)
                changed = True
        if changed:
            match.updated_at = datetime.now()
            updated += 1
        else:
            unchanged += 1

    removed = delete_matches_not_in_keys(db, fixture_keys)
    db.commit()
    message = f"赛程导入完成：新增 {created}，更新 {updated}，未变 {unchanged}，移除 {removed}"
    write_to_log("赛程导入", f"{message}; source={path}", operator)
    return ScheduleImportResponse(
        success=True,
        message=message,
        source_file=str(path),
        created=created,
        updated=updated,
        unchanged=unchanged,
        removed=removed,
        warnings=sorted(set(warnings)),
    )


def import_latest_schedule(db: Session, admin: str | None, write_to_log: LogWriter) -> ScheduleImportResponse:
    require_admin(admin)
    return import_schedule_file(db, admin, write_to_log, find_latest_schedule_file())


def get_schedule(db: Session, *, level: str | None = None, round_no: int | None = None) -> ScheduleResponse:
    matches = list_matches(db, level=level, round_no=round_no)
    events_by_match: dict[int, list[MatchPlayerEventResponse]] = {}
    for event in list_match_events(db, {match.id for match in matches}):
        events_by_match.setdefault(event.match_id, []).append(MatchPlayerEventResponse.model_validate(event))
    responses = []
    for match in matches:
        response = MatchResponse.model_validate(match)
        response.events = events_by_match.get(match.id, [])
        responses.append(response)
    return ScheduleResponse(
        levels=sorted({match.level for match in matches}, key=lambda item: (LEVEL_ORDER.get(item, 99), item)),
        rounds=sorted({match.round_no for match in matches}),
        matches=responses,
    )


def get_standings(db: Session, *, level: str | None = None, include_predictions: bool = True) -> StandingsResponse:
    teams = list_visible_teams(db, VISIBLE_LEVEL)
    if level:
        teams = [team for team in teams if team.level == level]
    teams = sorted(teams, key=lambda team: (LEVEL_ORDER.get(team.level, 99), team.name))
    rows_by_team = {
        team.name: {
            "level": team.level,
            "team_id": team.id,
            "team_name": team.name,
            "manager": team.manager,
            "played": 0,
            "wins": 0,
            "draws": 0,
            "losses": 0,
            "goals_for": 0,
            "goals_against": 0,
            "goal_difference": 0,
            "points": 0,
            "goal_rate": 0.0,
            "win_rate": 0.0,
            "home_played": 0,
            "home_wins": 0,
            "home_draws": 0,
            "home_losses": 0,
            "home_goals_for": 0,
            "home_goals_against": 0,
            "home_goal_difference": 0,
            "home_points": 0,
            "home_win_rate": 0.0,
            "away_played": 0,
            "away_wins": 0,
            "away_draws": 0,
            "away_losses": 0,
            "away_goals_for": 0,
            "away_goals_against": 0,
            "away_goal_difference": 0,
            "away_points": 0,
            "away_win_rate": 0.0,
        }
        for team in teams
    }
    rows_by_team_id = {row["team_id"]: row for row in rows_by_team.values() if row["team_id"] is not None}
    rows_by_lookup_name = dict(rows_by_team)
    rows_by_normalized_name: dict[str, dict[str, Any]] = {}
    for team in teams:
        row = rows_by_team.get(team.name)
        if not row:
            continue
        rows_by_lookup_name.setdefault(team.name, row)
        normalized = _normalize_team_lookup_name(team.name)
        if normalized:
            rows_by_normalized_name.setdefault(normalized, row)
    for raw_name, canonical_name in SCHEDULE_TEAM_ALIASES.items():
        row = rows_by_lookup_name.get(canonical_name)
        if not row:
            continue
        rows_by_lookup_name.setdefault(raw_name, row)
        normalized = _normalize_team_lookup_name(raw_name)
        if normalized:
            rows_by_normalized_name.setdefault(normalized, row)

    def resolve_match_row(match: Match, side: str) -> dict[str, Any] | None:
        team_id = getattr(match, f"{side}_team_id")
        team_name = str(getattr(match, f"{side}_team_name") or "")
        row = rows_by_team_id.get(team_id) if team_id is not None else None
        row = row or rows_by_lookup_name.get(team_name)
        if not row:
            row = rows_by_normalized_name.get(_normalize_team_lookup_name(team_name))
        return row

    all_matches = list_matches(db, level=level)
    played_matches = [
        match for match in all_matches
        if match.status in PLAYED_MATCH_STATUSES
        and match.home_score is not None
        and match.away_score is not None
    ]
    for match in played_matches:
        home = resolve_match_row(match, "home")
        away = resolve_match_row(match, "away")
        if not home or not away:
            continue
        home_score = int(match.home_score or 0)
        away_score = int(match.away_score or 0)

        home["played"] += 1
        away["played"] += 1
        home["home_played"] += 1
        away["away_played"] += 1
        home["goals_for"] += home_score
        home["goals_against"] += away_score
        away["goals_for"] += away_score
        away["goals_against"] += home_score
        home["home_goals_for"] += home_score
        home["home_goals_against"] += away_score
        away["away_goals_for"] += away_score
        away["away_goals_against"] += home_score

        if match.status == "home_forfeit":
            home["draws"] += 1
            home["home_draws"] += 1
            away["draws"] += 1
            away["away_draws"] += 1
            home["points"] += 1
            away["points"] += 1
            home["home_points"] += 1
            away["away_points"] += 1
        elif match.status == "away_forfeit":
            home["wins"] += 1
            home["home_wins"] += 1
            home["points"] += 3
            home["home_points"] += 3
            away["losses"] += 1
            away["away_losses"] += 1
        elif match.status == "double_forfeit":
            home["losses"] += 1
            home["home_losses"] += 1
            away["losses"] += 1
            away["away_losses"] += 1
        elif home_score > away_score:
            home["wins"] += 1
            home["home_wins"] += 1
            home["points"] += 3
            home["home_points"] += 3
            away["losses"] += 1
            away["away_losses"] += 1
        elif home_score < away_score:
            away["wins"] += 1
            away["away_wins"] += 1
            away["points"] += 3
            away["away_points"] += 3
            home["losses"] += 1
            home["home_losses"] += 1
        else:
            home["draws"] += 1
            home["home_draws"] += 1
            away["draws"] += 1
            away["away_draws"] += 1
            home["points"] += 1
            away["points"] += 1
            home["home_points"] += 1
            away["away_points"] += 1

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows_by_team.values():
        row["goal_difference"] = row["goals_for"] - row["goals_against"]
        row["home_goal_difference"] = row["home_goals_for"] - row["home_goals_against"]
        row["away_goal_difference"] = row["away_goals_for"] - row["away_goals_against"]
        row["goal_rate"] = round((row["goals_for"] / row["played"]), 2) if row["played"] else 0.0
        row["win_rate"] = round((row["wins"] / row["played"] * 100), 1) if row["played"] else 0.0
        row["home_win_rate"] = round((row["home_wins"] / row["home_played"] * 100), 1) if row["home_played"] else 0.0
        row["away_win_rate"] = round((row["away_wins"] / row["away_played"] * 100), 1) if row["away_played"] else 0.0
        grouped.setdefault(row["level"], []).append(row)

    response_rows: list[StandingRowResponse] = []
    prediction_summaries: list[StandingsPredictionSummaryResponse] = []
    for level in sorted(grouped, key=lambda item: (LEVEL_ORDER.get(item, 99), item)):
        ranked_rows = sorted(
            grouped[level],
            key=lambda row: (
                -row["points"],
                -row["goal_difference"],
                -row["goals_for"],
                -row["wins"],
                row["team_name"],
            ),
        )
        for index, row in enumerate(ranked_rows, start=1):
            row["rank"] = index

        if not include_predictions:
            response_rows.extend(StandingRowResponse(**row) for row in ranked_rows)
            continue

        remaining_fixtures: list[tuple[str, str]] = []
        resolved_match_count = 0
        for match in all_matches:
            if str(match.level or "") != level:
                continue
            home = resolve_match_row(match, "home")
            away = resolve_match_row(match, "away")
            if not home or not away or home is away:
                continue
            is_played = (
                match.status in PLAYED_MATCH_STATUSES
                and match.home_score is not None
                and match.away_score is not None
            )
            is_remaining = (
                match.status in {"scheduled", "postponed"}
                and match.home_score is None
                and match.away_score is None
            )
            if not is_played and not is_remaining:
                continue
            resolved_match_count += 1
            if is_remaining:
                remaining_fixtures.append((str(home["team_name"]), str(away["team_name"])))

        prediction = standings_prediction_service.predict_level(
            level,
            ranked_rows,
            remaining_fixtures,
            total_match_count=resolved_match_count,
        )
        prediction_summaries.append(StandingsPredictionSummaryResponse(**prediction["summary"]))
        predictions_by_team = prediction["teams"]
        for row in ranked_rows:
            row.update(predictions_by_team.get(str(row["team_name"]), {}))
            response_rows.append(StandingRowResponse(**row))

    return StandingsResponse(
        levels=sorted(grouped, key=lambda item: (LEVEL_ORDER.get(item, 99), item)),
        rows=response_rows,
        prediction_summaries=prediction_summaries,
    )


def get_standings_history(db: Session, *, level: str) -> StandingsHistoryResponse:
    teams = sorted(
        [team for team in list_visible_teams(db, VISIBLE_LEVEL) if team.level == level],
        key=lambda team: team.name,
    )
    matches = sorted(list_matches(db, level=level), key=lambda match: (match.round_no, match.id))
    team_by_id = {int(team.id): team for team in teams}
    team_by_name = {team.name: team for team in teams}
    team_by_normalized_name = {
        _normalize_team_lookup_name(team.name): team
        for team in teams
        if _normalize_team_lookup_name(team.name)
    }
    for raw_name, canonical_name in SCHEDULE_TEAM_ALIASES.items():
        team = team_by_name.get(canonical_name)
        if not team:
            continue
        team_by_name.setdefault(raw_name, team)
        normalized = _normalize_team_lookup_name(raw_name)
        if normalized:
            team_by_normalized_name.setdefault(normalized, team)

    def resolve_team(match: Match, side: str) -> Team | None:
        team_id = getattr(match, f"{side}_team_id")
        team_name = str(getattr(match, f"{side}_team_name") or "")
        team = team_by_id.get(int(team_id)) if team_id is not None else None
        return team or team_by_name.get(team_name) or team_by_normalized_name.get(_normalize_team_lookup_name(team_name))

    stats = {
        int(team.id): {
            "team_id": int(team.id),
            "team_name": team.name,
            "played": 0,
            "wins": 0,
            "draws": 0,
            "losses": 0,
            "goals_for": 0,
            "goals_against": 0,
            "goal_difference": 0,
            "points": 0,
        }
        for team in teams
    }

    def ranked_snapshot(previous_ranks: dict[int, int]) -> tuple[list[StandingHistoryRowResponse], dict[int, int]]:
        ranked = sorted(
            stats.values(),
            key=lambda row: (
                -row["points"],
                -row["goal_difference"],
                -row["goals_for"],
                -row["wins"],
                row["team_name"],
            ),
        )
        ranks = {int(row["team_id"]): index for index, row in enumerate(ranked, start=1)}
        response_rows = [
            StandingHistoryRowResponse(
                **row,
                rank=ranks[int(row["team_id"])],
                previous_rank=previous_ranks.get(int(row["team_id"]), ranks[int(row["team_id"])]),
                rank_change=previous_ranks.get(int(row["team_id"]), ranks[int(row["team_id"])]) - ranks[int(row["team_id"])],
            )
            for row in ranked
        ]
        return response_rows, ranks

    opening_rows, previous_ranks = ranked_snapshot({})
    history_rounds = [
        StandingHistoryRoundResponse(
            round_no=0,
            round_label="开赛前",
            is_complete=True,
            rows=opening_rows,
        )
    ]
    matches_by_round: dict[int, list[Match]] = {}
    latest_recorded_round = 0
    for match in matches:
        matches_by_round.setdefault(int(match.round_no), []).append(match)
        if (
            match.status in PLAYED_MATCH_STATUSES
            and match.home_score is not None
            and match.away_score is not None
        ):
            latest_recorded_round = max(latest_recorded_round, int(match.round_no))

    latest_complete_round = 0
    continuous_rounds_complete = True
    for round_no in sorted(round_no for round_no in matches_by_round if round_no <= latest_recorded_round):
        round_matches = matches_by_round[round_no]
        played_match_count = 0
        round_complete = bool(round_matches)
        for match in round_matches:
            home_team = resolve_team(match, "home")
            away_team = resolve_team(match, "away")
            if not home_team or not away_team or home_team.id == away_team.id:
                round_complete = False
                continue
            is_played = (
                match.status in PLAYED_MATCH_STATUSES
                and match.home_score is not None
                and match.away_score is not None
            )
            is_resolved = match.status in HISTORY_RESOLVED_MATCH_STATUSES and (
                match.status == "cancelled" or (match.home_score is not None and match.away_score is not None)
            )
            if not is_resolved:
                round_complete = False
            if not is_played:
                continue
            played_match_count += 1
            home = stats[int(home_team.id)]
            away = stats[int(away_team.id)]
            home_score = int(match.home_score or 0)
            away_score = int(match.away_score or 0)
            home["played"] += 1
            away["played"] += 1
            home["goals_for"] += home_score
            home["goals_against"] += away_score
            away["goals_for"] += away_score
            away["goals_against"] += home_score
            if match.status == "home_forfeit":
                home["draws"] += 1
                away["draws"] += 1
                home["points"] += 1
                away["points"] += 1
            elif match.status == "away_forfeit":
                home["wins"] += 1
                home["points"] += 3
                away["losses"] += 1
            elif match.status == "double_forfeit":
                home["losses"] += 1
                away["losses"] += 1
            elif home_score > away_score:
                home["wins"] += 1
                home["points"] += 3
                away["losses"] += 1
            elif home_score < away_score:
                away["wins"] += 1
                away["points"] += 3
                home["losses"] += 1
            else:
                home["draws"] += 1
                away["draws"] += 1
                home["points"] += 1
                away["points"] += 1
            home["goal_difference"] = home["goals_for"] - home["goals_against"]
            away["goal_difference"] = away["goals_for"] - away["goals_against"]

        rows, previous_ranks = ranked_snapshot(previous_ranks)
        history_rounds.append(
            StandingHistoryRoundResponse(
                round_no=round_no,
                round_label=f"第{round_no}轮",
                is_complete=round_complete,
                played_match_count=played_match_count,
                total_match_count=len(round_matches),
                rows=rows,
            )
        )
        if continuous_rounds_complete and round_complete and round_no == latest_complete_round + 1:
            latest_complete_round = round_no
        else:
            continuous_rounds_complete = False

    return StandingsHistoryResponse(
        level=level,
        total_rounds=max(matches_by_round, default=0),
        latest_recorded_round=latest_recorded_round,
        latest_complete_round=latest_complete_round,
        teams=[
            StandingHistoryTeamResponse(
                team_id=int(team.id),
                team_name=team.name,
                manager=team.manager,
                logo_path=team.logo_path,
            )
            for team in teams
        ],
        rounds=history_rounds,
    )


def _parse_match_date(value: str | None):
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    raise HTTPException(status_code=400, detail="比赛日期格式应为 YYYY-MM-DD 或 YYYY-MM-DD HH:MM")


def _normalize_status(request: MatchUpdateRequest) -> str:
    status = str(request.status or "").strip().lower()
    has_score = request.home_score is not None and request.away_score is not None
    if not status:
        return "played" if has_score else "scheduled"
    if status not in MATCH_STATUSES:
        raise HTTPException(status_code=400, detail="比赛状态仅支持 scheduled、played、postponed、cancelled 或判负状态")
    return status


def _get_forfeit_score(status: str) -> tuple[int, int] | None:
    if status == "home_forfeit":
        return 0, 0
    if status == "away_forfeit":
        return 2, 0
    if status == "double_forfeit":
        return 0, 0
    return None


def _apply_match_result_update(match: Match, request: MatchUpdateRequest) -> None:
    status = _normalize_status(request)
    forfeit_score = _get_forfeit_score(status)
    if forfeit_score is not None:
        match.home_score, match.away_score = forfeit_score
        match.status = status
        if hasattr(request, "match_date"):
            match.match_date = _parse_match_date(request.match_date)
        match.notes = str(request.notes or "").strip() or None
        match.updated_at = datetime.now()
        return

    if status == "played":
        if request.home_score is None or request.away_score is None:
            raise HTTPException(status_code=400, detail="已赛比赛必须填写双方比分")
        if request.home_score < 0 or request.away_score < 0:
            raise HTTPException(status_code=400, detail="比分不能为负数")

    match.home_score = request.home_score if request.home_score is not None else None
    match.away_score = request.away_score if request.away_score is not None else None
    match.status = status
    if hasattr(request, "match_date"):
        match.match_date = _parse_match_date(request.match_date)
    match.notes = str(request.notes or "").strip() or None
    match.updated_at = datetime.now()


def _player_belongs_to_team(player: Player, *, team_id: int | None, team_name: str) -> bool:
    return (team_id is not None and player.team_id == team_id) or str(player.team_name or "") == team_name


def _resolve_event_player(
    team_players: list[Player],
    item: MatchPlayerEventUpdateItem,
    *,
    team_name: str,
) -> Player:
    if item.player_uid is not None:
        for player in team_players:
            if player.uid == item.player_uid:
                return player
        raise HTTPException(status_code=400, detail=f"{team_name} 中未找到球员 UID：{item.player_uid}")

    player_name = str(item.player_name or "").strip()
    matches = [player for player in team_players if str(player.name or "").strip() == player_name]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise HTTPException(status_code=400, detail=f"{team_name} 中未找到球员：{player_name}")
    raise HTTPException(status_code=400, detail=f"{team_name} 中存在重名球员，请使用 UID 选择：{player_name}")


def _replace_match_player_events(db: Session, match: Match, request: MatchUpdateRequest) -> None:
    delete_match_events(db, match.id)
    if match.status != "played":
        return

    side_by_team_name = {}
    for raw_team_id, raw_team_name, score in (
        (match.home_team_id, str(match.home_team_name), int(match.home_score or 0)),
        (match.away_team_id, str(match.away_team_name), int(match.away_score or 0)),
    ):
        team = _resolve_schedule_team(db, raw_team_id, raw_team_name)
        canonical_team_id = team.id if team else raw_team_id
        canonical_team_name = team.name if team else raw_team_name
        players = get_team_players(db, team) if team else get_players_by_team_name(db, raw_team_name)
        side = {
            "team_id": canonical_team_id,
            "team_name": canonical_team_name,
            "match_team_name": raw_team_name,
            "score": score,
            "players": players,
        }
        for lookup_name in {raw_team_name, canonical_team_name, SCHEDULE_TEAM_ALIASES.get(raw_team_name, raw_team_name)}:
            if lookup_name:
                side_by_team_name[str(lookup_name)] = side
    totals = {
        id(side): {"goal": 0, "assist": 0}
        for side in side_by_team_name.values()
    }

    for item in request.events or []:
        event_type = str(item.event_type or "").strip().lower()
        if event_type not in MATCH_EVENT_TYPES:
            raise HTTPException(status_code=400, detail="比赛事件仅支持 goal、own_goal、assist 或 mvp")
        quantity = int(item.quantity or 0)
        if quantity <= 0:
            raise HTTPException(status_code=400, detail="球员事件数量必须大于 0")
        if event_type == "mvp":
            quantity = 1
        team_name = str(item.team_name or "").strip()
        side = side_by_team_name.get(team_name)
        if not side:
            raise HTTPException(status_code=400, detail=f"球员事件球队不属于本场比赛：{team_name}")
        player = None
        if event_type != "own_goal":
            player = _resolve_event_player(side["players"], item, team_name=side["team_name"])
            if not _player_belongs_to_team(player, team_id=side["team_id"], team_name=side["team_name"]):
                raise HTTPException(status_code=400, detail=f"球员不属于 {side['team_name']}：{player.name}")

        side_key = id(side)
        if event_type in {"goal", "own_goal"}:
            totals[side_key]["goal"] += quantity
        elif event_type in totals[side_key]:
            totals[side_key][event_type] += quantity
        db.add(
            MatchPlayerEvent(
                match_id=match.id,
                team_id=side["team_id"],
                team_name=side["team_name"],
                player_uid=player.uid if player else None,
                player_name=player.name if player else "乌龙球",
                event_type=event_type,
                quantity=quantity,
                created_at=datetime.now(),
                updated_at=datetime.now(),
            )
        )

    checked_side_ids = set()
    for side in side_by_team_name.values():
        side_key = id(side)
        if side_key in checked_side_ids:
            continue
        checked_side_ids.add(side_key)
        if totals[side_key]["goal"] > side["score"]:
            raise HTTPException(status_code=400, detail=f"{side['match_team_name']} 的进球明细数量不能超过比分 {side['score']}")
        if totals[side_key]["assist"] > side["score"]:
            raise HTTPException(status_code=400, detail=f"{side['match_team_name']} 的助攻数量不能超过进球数")


def update_match_result(
    db: Session,
    admin: str | None,
    match_id: int,
    request: MatchUpdateRequest,
    write_to_log: LogWriter,
) -> dict[str, Any]:
    operator = require_admin(admin)
    match = get_match_by_id(db, match_id)
    if not match:
        raise HTTPException(status_code=404, detail="比赛不存在")

    _apply_match_result_update(match, request)
    _replace_match_player_events(db, match, request)
    db.commit()

    score_text = "-" if match.home_score is None or match.away_score is None else f"{match.home_score}-{match.away_score}"
    message = f"已更新第 {match.round_no} 轮 {match.home_team_name} vs {match.away_team_name}：{score_text}"
    write_to_log("赛程比分编辑", message, operator)
    return {"success": True, "message": message}


def batch_update_match_results(
    db: Session,
    admin: str | None,
    request: MatchBatchUpdateRequest,
    write_to_log: LogWriter,
) -> dict[str, Any]:
    operator = require_admin(admin)
    if not request.matches:
        raise HTTPException(status_code=400, detail="没有需要保存的比赛")

    seen_ids: set[int] = set()
    updated = 0
    for item in request.matches:
        if item.match_id in seen_ids:
            continue
        seen_ids.add(item.match_id)
        match = get_match_by_id(db, item.match_id)
        if not match:
            raise HTTPException(status_code=404, detail=f"比赛不存在：{item.match_id}")
        _apply_match_result_update(
            match,
            MatchUpdateRequest(
                home_score=item.home_score,
                away_score=item.away_score,
                status=item.status,
                notes=item.notes,
                events=item.events,
            ),
        )
        _replace_match_player_events(db, match, MatchUpdateRequest(
            home_score=item.home_score,
            away_score=item.away_score,
            status=item.status,
            notes=item.notes,
            events=item.events,
        ))
        updated += 1

    db.commit()
    message = f"已保存 {updated} 场比赛进展"
    write_to_log("赛程比分批量编辑", message, operator)
    return {"success": True, "message": message}
