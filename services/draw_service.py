from __future__ import annotations

import hashlib
import io
import json
import secrets
from datetime import datetime
from typing import Any

from fastapi import HTTPException
from sqlalchemy.orm import Session

from models import (
    CandidateList,
    CandidateListPlayer,
    CupGroupTeam,
    CupMatch,
    DrawPick,
    DrawPoolEntry,
    DrawSession,
    Match,
    Player,
    SeasonArchive,
    Team,
)
from repositories.attribute_repository import resolve_attribute_version
from schemas_read import (
    DrawPickResponse,
    DrawPoolEntryResponse,
    DrawSessionDetailResponse,
    DrawSessionSummaryResponse,
    SeasonArchiveDetailResponse,
    SeasonArchiveSummaryResponse,
)
from schemas_write import (
    DrawPoolEntryRequest,
    DrawSessionCreateRequest,
    DrawSessionUpdateRequest,
    LotteryPoolBuildRequest,
    SeasonArchiveCreateRequest,
)
from services import cup_service, match_service, player_ranking_service
from services.operation_audit_service import AUDIT_SOURCE_ADMIN_UI, persist_admin_operation_audit
from wage_calculator import calculate_wage


DRAW_TYPES = {
    "champions_group": {"competition": "champions_cup", "kind": "groups", "entry_count": 30, "group_count": 5, "pot_count": 6},
    "league_group": {"competition": "league_cup", "kind": "groups", "entry_count": 24, "group_count": 4, "pot_count": 6},
    "champions_r16": {"competition": "champions_cup", "kind": "seeded_pairs", "entry_count": 16, "pair_count": 8},
    "league_r16": {"competition": "league_cup", "kind": "seeded_pairs", "entry_count": 16, "pair_count": 8},
    "wumingjian_qualifying": {"competition": "wumingjian_cup", "kind": "random_pairs", "entry_count": 44, "pair_count": 22},
    "wumingjian_r32": {"competition": "wumingjian_cup", "kind": "random_pairs", "entry_count": 32, "pair_count": 16},
    "lottery": {"competition": None, "kind": "weighted", "entry_count": None},
    "custom_team": {"competition": None, "kind": "custom", "entry_count": None},
    "custom_player": {"competition": None, "kind": "custom", "entry_count": None},
}

CUSTOM_DRAW_TYPES = {"custom_team", "custom_player"}


def _now() -> datetime:
    return datetime.now()


def _loads(value: str | None) -> dict[str, Any]:
    try:
        payload = json.loads(value or "{}")
    except (TypeError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _audit(db: Session, operator: str, action: str, summary: str, details: dict[str, Any] | None = None) -> None:
    bind = db.get_bind()
    if bind is None:
        return
    persist_admin_operation_audit(
        bind,
        category="draw",
        action=action,
        operator=operator,
        status="success",
        summary=summary,
        source=AUDIT_SOURCE_ADMIN_UI,
        operation_label="抽签",
        extra_details=details or {},
    )


def _session_or_404(db: Session, session_id: int, *, public: bool = False) -> DrawSession:
    query = db.query(DrawSession).filter(DrawSession.id == int(session_id))
    if public:
        query = query.filter(DrawSession.status == "published")
    record = query.first()
    if not record:
        raise HTTPException(status_code=404, detail="抽签记录不存在")
    return record


def _entry_response(entry: DrawPoolEntry) -> DrawPoolEntryResponse:
    return DrawPoolEntryResponse(
        id=int(entry.id),
        entity_key=entry.entity_key,
        entity_type=entry.entity_type,
        team_id=entry.team_id,
        player_uid=entry.player_uid,
        entity_name=entry.entity_name,
        team_name=entry.team_name,
        level=entry.level,
        source_rank=entry.source_rank,
        pot_no=entry.pot_no,
        seed_status=entry.seed_status,
        self_save_count=int(entry.self_save_count or 0),
        weight=float(entry.weight or 1),
        final_value=entry.final_value,
        slot_type=entry.slot_type,
        is_active=bool(entry.is_active),
        metadata=_loads(entry.metadata_json),
    )


def _entries(db: Session, session_id: int) -> list[DrawPoolEntry]:
    return db.query(DrawPoolEntry).filter(DrawPoolEntry.session_id == session_id).order_by(DrawPoolEntry.id).all()


def _picks(db: Session, session_id: int) -> list[DrawPick]:
    return db.query(DrawPick).filter(DrawPick.session_id == session_id).order_by(DrawPick.sequence_no).all()


def _summary(db: Session, record: DrawSession) -> DrawSessionSummaryResponse:
    entries = _entries(db, record.id)
    picks = _picks(db, record.id)
    return DrawSessionSummaryResponse(
        id=record.id,
        name=record.name,
        draw_type=record.draw_type,
        competition=record.competition,
        season_label=record.season_label,
        status=record.status,
        random_seed=record.random_seed,
        pool_hash=record.pool_hash,
        entry_count=len(entries),
        active_entry_count=sum(bool(row.is_active) for row in entries),
        pick_count=sum(row.status == "active" for row in picks),
        candidate_list_id=record.candidate_list_id,
        created_by=record.created_by,
        updated_by=record.updated_by,
        created_at=record.created_at,
        updated_at=record.updated_at,
        locked_at=record.locked_at,
        completed_at=record.completed_at,
        published_at=record.published_at,
    )


def _detail(db: Session, record: DrawSession) -> DrawSessionDetailResponse:
    entries = _entries(db, record.id)
    by_id = {row.id: row for row in entries}
    picks = _picks(db, record.id)
    return DrawSessionDetailResponse(
        **_summary(db, record).model_dump(),
        config=_loads(record.config_json),
        result=_loads(record.result_json),
        entries=[_entry_response(row) for row in entries],
        picks=[
            DrawPickResponse(
                id=row.id,
                sequence_no=row.sequence_no,
                entry=_entry_response(by_id[row.entry_id]),
                paired_entry=_entry_response(by_id[row.paired_entry_id]) if row.paired_entry_id and row.paired_entry_id in by_id else None,
                target_group=row.target_group,
                target_slot=row.target_slot,
                random_value=row.random_value,
                status=row.status,
                reason=row.reason,
                created_at=row.created_at,
            )
            for row in picks
        ],
    )


def list_sessions(db: Session, *, public: bool = False) -> list[DrawSessionSummaryResponse]:
    query = db.query(DrawSession)
    if public:
        query = query.filter(DrawSession.status == "published")
    records = query.order_by(DrawSession.created_at.desc(), DrawSession.id.desc()).all()
    return [_summary(db, row) for row in records]


def get_session(db: Session, session_id: int, *, public: bool = False) -> DrawSessionDetailResponse:
    return _detail(db, _session_or_404(db, session_id, public=public))


def _entry_payload(db: Session, request: DrawPoolEntryRequest, draw_type: str) -> dict[str, Any]:
    if draw_type == "lottery":
        if request.player_uid is None:
            raise HTTPException(status_code=400, detail="乐透候选项缺少球员 UID")
        player = db.query(Player).filter(Player.uid == int(request.player_uid)).first()
        if not player:
            raise HTTPException(status_code=400, detail=f"球员不存在：{request.player_uid}")
        wage = calculate_wage(player.initial_ca or 0, player.ca or 0, player.pa or 0, player.age or 0, player.position or "")
        saves = int(request.self_save_count or 0)
        return {
            "entity_key": f"player:{player.uid}",
            "entity_type": "player",
            "team_id": player.team_id,
            "player_uid": player.uid,
            "entity_name": player.name,
            "team_name": player.team_name,
            "level": request.level,
            "source_rank": request.source_rank,
            "pot_no": None,
            "seed_status": None,
            "self_save_count": saves,
            "weight": float(2**saves),
            "final_value": float(wage["final_value"]),
            "slot_type": wage["slot_type"],
            "is_active": 1 if request.is_active else 0,
            "metadata_json": _dumps(request.metadata),
        }
    if draw_type == "custom_player":
        player = None
        if request.player_uid is not None:
            player = db.query(Player).filter(Player.uid == int(request.player_uid)).first()
            if not player:
                raise HTTPException(status_code=400, detail=f"球员不存在：{request.player_uid}")
        entity_name = str(player.name if player else request.entity_name or "").strip()
        if not entity_name:
            raise HTTPException(status_code=400, detail="自由球员抽签候选项缺少球员名称")
        entity_key = f"player:{player.uid}" if player else f"custom-player:{hashlib.sha256(entity_name.casefold().encode('utf-8')).hexdigest()[:20]}"
        return {
            "entity_key": entity_key,
            "entity_type": "player",
            "team_id": player.team_id if player else request.team_id,
            "player_uid": player.uid if player else None,
            "entity_name": entity_name,
            "team_name": player.team_name if player else str(request.team_name or "").strip() or None,
            "level": request.level,
            "source_rank": request.source_rank,
            "pot_no": None,
            "seed_status": None,
            "self_save_count": 0,
            "weight": 1.0,
            "final_value": None,
            "slot_type": None,
            "is_active": 1 if request.is_active else 0,
            "metadata_json": _dumps(request.metadata),
        }
    if draw_type == "custom_team" and request.team_id is None:
        entity_name = str(request.entity_name or "").strip()
        if not entity_name:
            raise HTTPException(status_code=400, detail="自由球队抽签候选项缺少球队名称")
        return {
            "entity_key": f"custom-team:{hashlib.sha256(entity_name.casefold().encode('utf-8')).hexdigest()[:20]}",
            "entity_type": "team",
            "team_id": None,
            "player_uid": None,
            "entity_name": entity_name,
            "team_name": entity_name,
            "level": request.level,
            "source_rank": request.source_rank,
            "pot_no": None,
            "seed_status": None,
            "self_save_count": 0,
            "weight": 1.0,
            "final_value": None,
            "slot_type": None,
            "is_active": 1 if request.is_active else 0,
            "metadata_json": _dumps(request.metadata),
        }
    if request.team_id is None:
        raise HTTPException(status_code=400, detail="杯赛候选项缺少球队")
    team = db.query(Team).filter(Team.id == int(request.team_id)).first()
    if not team or team.level == "隐藏":
        raise HTTPException(status_code=400, detail=f"球队不存在：{request.team_id}")
    return {
        "entity_key": f"team:{team.id}",
        "entity_type": "team",
        "team_id": team.id,
        "player_uid": None,
        "entity_name": team.name,
        "team_name": team.name,
        "level": request.level or team.level,
        "source_rank": request.source_rank,
        "pot_no": request.pot_no,
        "seed_status": request.seed_status,
        "self_save_count": 0,
        "weight": 1.0,
        "final_value": None,
        "slot_type": None,
        "is_active": 1 if request.is_active else 0,
        "metadata_json": _dumps(request.metadata),
    }


def _replace_entries(db: Session, record: DrawSession, requests: list[DrawPoolEntryRequest]) -> None:
    if record.status != "draft":
        raise HTTPException(status_code=409, detail="候选池锁定后不能直接修改，请作废整场后重新建立")
    payloads = [_entry_payload(db, request, record.draw_type) for request in requests]
    keys = [row["entity_key"] for row in payloads]
    if len(keys) != len(set(keys)):
        raise HTTPException(status_code=400, detail="候选池存在重复球队或球员")
    db.query(DrawPoolEntry).filter(DrawPoolEntry.session_id == record.id).delete(synchronize_session=False)
    for payload in payloads:
        db.add(DrawPoolEntry(session_id=record.id, created_at=_now(), updated_at=_now(), **payload))


def create_session(db: Session, operator: str, request: DrawSessionCreateRequest) -> DrawSessionDetailResponse:
    name = str(request.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="抽签名称不能为空")
    definition = DRAW_TYPES[request.draw_type]
    record = DrawSession(
        name=name,
        draw_type=request.draw_type,
        competition=definition["competition"],
        season_label=str(request.season_label or "").strip() or None,
        status="draft",
        random_seed=str(request.random_seed or secrets.token_hex(16)),
        config_json=_dumps(request.config),
        result_json="{}",
        created_by=operator,
        updated_by=operator,
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(record)
    db.flush()
    _replace_entries(db, record, request.entries)
    db.commit()
    db.refresh(record)
    _audit(db, operator, "create", f"创建抽签：{record.name}", {"draw_session_id": record.id, "draw_type": record.draw_type})
    return _detail(db, record)


def update_session(db: Session, operator: str, session_id: int, request: DrawSessionUpdateRequest) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.status != "draft":
        raise HTTPException(status_code=409, detail="候选池锁定后不能修改")
    if request.name is not None:
        name = str(request.name).strip()
        if not name:
            raise HTTPException(status_code=400, detail="抽签名称不能为空")
        record.name = name
    if request.season_label is not None:
        record.season_label = str(request.season_label).strip() or None
    if request.random_seed is not None:
        random_seed = str(request.random_seed).strip()
        if not random_seed:
            raise HTTPException(status_code=400, detail="随机种子不能为空")
        record.random_seed = random_seed
    if request.config is not None:
        record.config_json = _dumps(request.config)
    if request.entries is not None:
        _replace_entries(db, record, request.entries)
    record.updated_by = operator
    record.updated_at = _now()
    db.commit()
    db.refresh(record)
    _audit(db, operator, "update", f"更新抽签草稿：{record.name}", {"draw_session_id": record.id})
    return _detail(db, record)


def delete_session(db: Session, operator: str, session_id: int) -> dict[str, Any]:
    record = _session_or_404(db, session_id)
    if record.status not in {"draft", "void"}:
        raise HTTPException(status_code=409, detail="仅草稿或已作废任务可以删除；请先作废当前抽签")
    audit_details = {
        "draw_session_id": int(record.id),
        "draw_type": record.draw_type,
        "status": record.status,
        "entry_count": db.query(DrawPoolEntry).filter(DrawPoolEntry.session_id == record.id).count(),
        "pick_count": db.query(DrawPick).filter(DrawPick.session_id == record.id).count(),
    }
    name = record.name
    db.query(DrawPick).filter(DrawPick.session_id == record.id).delete(synchronize_session=False)
    db.query(DrawPoolEntry).filter(DrawPoolEntry.session_id == record.id).delete(synchronize_session=False)
    db.delete(record)
    db.commit()
    _audit(db, operator, "delete", f"删除抽签任务：{name}", audit_details)
    return {"success": True, "message": "抽签任务已删除"}


def _pool_hash(entries: list[DrawPoolEntry]) -> str:
    payload = [
        {
            "key": row.entity_key,
            "team": row.team_name,
            "pot": row.pot_no,
            "seed": row.seed_status,
            "weight": float(row.weight or 1),
            "value": row.final_value,
            "slot": row.slot_type,
            "active": bool(row.is_active),
        }
        for row in sorted(entries, key=lambda item: item.entity_key)
    ]
    return hashlib.sha256(_dumps(payload).encode("utf-8")).hexdigest()


def _validate_pool(record: DrawSession, entries: list[DrawPoolEntry]) -> None:
    definition = DRAW_TYPES[record.draw_type]
    active = [row for row in entries if row.is_active]
    expected = definition.get("entry_count")
    if expected is not None and len(active) != expected:
        raise HTTPException(status_code=400, detail=f"该抽签需要 {expected} 个有效候选，当前为 {len(active)} 个")
    if definition["kind"] == "groups":
        pot_count = definition["pot_count"]
        group_count = definition["group_count"]
        for pot_no in range(1, pot_count + 1):
            count = sum(row.pot_no == pot_no for row in active)
            if count != group_count:
                raise HTTPException(status_code=400, detail=f"第 {pot_no} 档需要 {group_count} 支球队，当前为 {count} 支")
    if definition["kind"] == "seeded_pairs":
        seeded = sum(row.seed_status == "seeded" for row in active)
        unseeded = sum(row.seed_status == "unseeded" for row in active)
        if seeded != 8 or unseeded != 8:
            raise HTTPException(status_code=400, detail=f"16强抽签需要8支种子和8支非种子，当前为 {seeded}/{unseeded}")
    if record.draw_type == "lottery":
        teams = {row.team_id for row in active if row.team_id}
        if not active or not teams:
            raise HTTPException(status_code=400, detail="乐透候选池为空")
    if record.draw_type in CUSTOM_DRAW_TYPES:
        if not active:
            raise HTTPException(status_code=400, detail="自由抽签候选池不能为空")
        config = _loads(record.config_json)
        mode = str(config.get("mode") or "list")
        if mode not in {"list", "groups", "pairs"}:
            raise HTTPException(status_code=400, detail="自由抽签结果形式无效")
        if mode == "groups":
            group_count = int(config.get("group_count") or 2)
            if group_count < 2 or group_count > min(26, len(active)):
                raise HTTPException(status_code=400, detail=f"分组数量应为 2 至 {min(26, len(active))} 组")
        if mode == "pairs" and len(active) < 2:
            raise HTTPException(status_code=400, detail="随机配对至少需要 2 个候选")
        if mode == "list":
            result_count = int(config.get("result_count") or len(active))
            if result_count < 1 or result_count > len(active):
                raise HTTPException(status_code=400, detail=f"名单数量应为 1 至 {len(active)}")


def lock_session(db: Session, operator: str, session_id: int) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.status != "draft":
        raise HTTPException(status_code=409, detail="只有草稿可以锁定")
    entries = _entries(db, record.id)
    _validate_pool(record, entries)
    record.pool_hash = _pool_hash(entries)
    record.status = "locked"
    record.locked_by = operator
    record.locked_at = _now()
    record.updated_by = operator
    record.updated_at = _now()
    db.commit()
    db.refresh(record)
    _audit(db, operator, "lock", f"锁定抽签候选池：{record.name}", {"draw_session_id": record.id, "pool_hash": record.pool_hash})
    return _detail(db, record)


def _hash_rank(seed: str, namespace: str, key: str) -> str:
    return hashlib.sha256(f"{seed}:{namespace}:{key}".encode("utf-8")).hexdigest()


def _complete(record: DrawSession, operator: str, result: dict[str, Any]) -> None:
    record.result_json = _dumps(result)
    record.status = "completed"
    record.completed_by = operator
    record.completed_at = _now()
    record.updated_by = operator
    record.updated_at = _now()


def _draw_groups(db: Session, record: DrawSession, operator: str) -> None:
    definition = DRAW_TYPES[record.draw_type]
    entries = [row for row in _entries(db, record.id) if row.is_active]
    sequence = 0
    groups: dict[str, list[dict[str, Any]]] = {}
    for pot_no in range(1, definition["pot_count"] + 1):
        pot = sorted(
            [row for row in entries if row.pot_no == pot_no],
            key=lambda row: _hash_rank(record.random_seed, f"pot:{pot_no}", row.entity_key),
        )
        for group_index, entry in enumerate(pot):
            sequence += 1
            group_name = chr(ord("A") + group_index)
            random_value = _hash_rank(record.random_seed, f"pot:{pot_no}", entry.entity_key)
            db.add(DrawPick(session_id=record.id, sequence_no=sequence, entry_id=entry.id, target_group=group_name, target_slot=pot_no, random_value=random_value, created_by=operator, created_at=_now()))
            groups.setdefault(group_name, []).append({"team_id": entry.team_id, "team_name": entry.team_name, "pot_no": pot_no})
    _complete(record, operator, {"groups": groups})


def _draw_pairs(db: Session, record: DrawSession, operator: str, *, seeded: bool) -> None:
    entries = [row for row in _entries(db, record.id) if row.is_active]
    if seeded:
        left = sorted([row for row in entries if row.seed_status == "seeded"], key=lambda row: _hash_rank(record.random_seed, "seeded", row.entity_key))
        right = sorted([row for row in entries if row.seed_status == "unseeded"], key=lambda row: _hash_rank(record.random_seed, "unseeded", row.entity_key))
    else:
        shuffled = sorted(entries, key=lambda row: _hash_rank(record.random_seed, "pair_pool", row.entity_key))
        left, right = shuffled[::2], shuffled[1::2]
    pairs = []
    for index, (first, second) in enumerate(zip(left, right), start=1):
        random_value = _hash_rank(record.random_seed, f"pair:{index}", f"{first.entity_key}:{second.entity_key}")
        db.add(DrawPick(session_id=record.id, sequence_no=index, entry_id=first.id, paired_entry_id=second.id, target_slot=index, random_value=random_value, created_by=operator, created_at=_now()))
        pairs.append({"slot_no": index, "team_a_id": first.team_id, "team_a_name": first.team_name, "team_b_id": second.team_id, "team_b_name": second.team_name})
    _complete(record, operator, {"pairs": pairs})


def _custom_result_entry(entry: DrawPoolEntry, *, slot_no: int | None = None) -> dict[str, Any]:
    return {
        "entity_name": entry.entity_name,
        "entity_type": entry.entity_type,
        "team_name": entry.team_name,
        "team_id": entry.team_id,
        "player_uid": entry.player_uid,
        "slot_no": slot_no,
    }


def _active_picks(db: Session, session_id: int) -> list[DrawPick]:
    return [row for row in _picks(db, session_id) if row.status == "active"]


def _pending_pair(record: DrawSession) -> dict[str, Any] | None:
    pending = _loads(record.result_json).get("pending_pair")
    return pending if isinstance(pending, dict) and pending.get("entry_id") else None


def _result_from_progress(
    db: Session,
    record: DrawSession,
    *,
    pending_pair: dict[str, Any] | None = None,
) -> dict[str, Any]:
    entries = {row.id: row for row in _entries(db, record.id)}
    picks = _active_picks(db, record.id)
    kind = DRAW_TYPES[record.draw_type]["kind"]
    config = _loads(record.config_json)

    if kind == "groups" or (kind == "custom" and str(config.get("mode") or "list") == "groups"):
        group_count = (
            int(DRAW_TYPES[record.draw_type]["group_count"])
            if kind == "groups"
            else int(config.get("group_count") or 2)
        )
        groups: dict[str, list[dict[str, Any]]] = {
            chr(ord("A") + index): [] for index in range(group_count)
        }
        for pick in picks:
            entry = entries.get(pick.entry_id)
            if not entry or not pick.target_group:
                continue
            if kind == "groups":
                item = {"team_id": entry.team_id, "team_name": entry.team_name, "pot_no": pick.target_slot}
            else:
                item = _custom_result_entry(entry, slot_no=pick.target_slot)
            groups.setdefault(str(pick.target_group), []).append(item)
        result: dict[str, Any] = {"groups": groups}
        if kind == "custom":
            result["mode"] = "groups"
        return result

    if kind in {"seeded_pairs", "random_pairs"} or (
        kind == "custom" and str(config.get("mode") or "list") == "pairs"
    ):
        pairs: list[dict[str, Any]] = []
        for pick in picks:
            first = entries.get(pick.entry_id)
            second = entries.get(pick.paired_entry_id) if pick.paired_entry_id else None
            if not first:
                continue
            if kind == "custom":
                pairs.append({
                    "slot_no": pick.target_slot,
                    "entry": _custom_result_entry(first),
                    "paired_entry": _custom_result_entry(second) if second else None,
                })
            else:
                pairs.append({
                    "slot_no": pick.target_slot,
                    "team_a_id": first.team_id,
                    "team_a_name": first.team_name,
                    "team_b_id": second.team_id if second else None,
                    "team_b_name": second.team_name if second else None,
                })
        result = {"pairs": pairs}
        if kind == "custom":
            result["mode"] = "pairs"
        if pending_pair:
            result["pending_pair"] = pending_pair
        return result

    if kind == "custom":
        items = [
            _custom_result_entry(entries[pick.entry_id], slot_no=pick.target_slot)
            for pick in picks
            if pick.entry_id in entries
        ]
        return {
            "mode": "list",
            "list": items,
            "selected_count": len(items),
            "candidate_count": sum(bool(row.is_active) for row in entries.values()),
        }

    if kind == "weighted":
        limit = max(1, min(100, int(config.get("limit") or 15)))
        return {"selected_count": len(picks), "requested_count": limit, "completed_short": False}
    return {}


def _mark_drawing(record: DrawSession, operator: str, result: dict[str, Any]) -> None:
    record.result_json = _dumps(result)
    record.status = "drawing"
    record.updated_by = operator
    record.updated_at = _now()


def _add_pick(
    db: Session,
    record: DrawSession,
    operator: str,
    *,
    sequence_no: int,
    entry: DrawPoolEntry,
    paired_entry: DrawPoolEntry | None = None,
    target_group: str | None = None,
    target_slot: int | None = None,
    namespace: str,
) -> DrawPick:
    click_time = _now()
    pair_key = f":{paired_entry.entity_key}" if paired_entry else ""
    random_value = "click:" + hashlib.sha256(
        f"{record.id}:{sequence_no}:{entry.entity_key}{pair_key}:{click_time.isoformat()}:{secrets.token_hex(8)}".encode("utf-8")
    ).hexdigest()
    pick = DrawPick(
        session_id=record.id,
        sequence_no=sequence_no,
        entry_id=entry.id,
        paired_entry_id=paired_entry.id if paired_entry else None,
        target_group=target_group,
        target_slot=target_slot,
        random_value=random_value,
        created_by=operator,
        created_at=click_time,
    )
    db.add(pick)
    db.flush()
    return pick


def _clicked_entry(allowed: list[DrawPoolEntry], entry_id: int, *, detail: str) -> DrawPoolEntry:
    selected = next((row for row in allowed if int(row.id) == int(entry_id)), None)
    if not selected:
        raise HTTPException(status_code=409, detail=detail)
    return selected


def _used_entry_ids(picks: list[DrawPick], pending: dict[str, Any] | None = None) -> set[int]:
    used = {int(pick.entry_id) for pick in picks}
    used.update(int(pick.paired_entry_id) for pick in picks if pick.paired_entry_id)
    if pending and pending.get("entry_id"):
        used.add(int(pending["entry_id"]))
    return used


def _draw_next_group(db: Session, record: DrawSession, operator: str, entry_id: int) -> tuple[str, bool]:
    definition = DRAW_TYPES[record.draw_type]
    picks = _active_picks(db, record.id)
    total = int(definition["pot_count"]) * int(definition["group_count"])
    if len(picks) >= total:
        raise HTTPException(status_code=409, detail="小组抽签已经完成")
    pot_no = (len(picks) // int(definition["group_count"])) + 1
    group_index = len(picks) % int(definition["group_count"])
    group_name = chr(ord("A") + group_index)
    used = _used_entry_ids(picks)
    pot = [
        row for row in _entries(db, record.id)
        if row.is_active and row.pot_no == pot_no and int(row.id) not in used
    ]
    selected = _clicked_entry(
        pot,
        entry_id,
        detail=f"点击的球队不在当前第{pot_no}档可抽池中，请按画面重新抽取",
    )
    _add_pick(
        db,
        record,
        operator,
        sequence_no=len(picks) + 1,
        entry=selected,
        target_group=group_name,
        target_slot=pot_no,
        namespace=f"pot:{pot_no}",
    )
    result = _result_from_progress(db, record)
    complete = len(_active_picks(db, record.id)) >= total
    (_complete if complete else _mark_drawing)(record, operator, result)
    return f"第{pot_no}档 {group_name}组：{selected.entity_name}", complete


def _draw_next_pair(db: Session, record: DrawSession, operator: str, entry_id: int, *, seeded: bool, custom: bool = False) -> tuple[str, bool]:
    entries = [row for row in _entries(db, record.id) if row.is_active]
    picks = _active_picks(db, record.id)
    pending = _pending_pair(record)
    pair_no = len(picks) + 1

    used = _used_entry_ids(picks, pending)
    available = [row for row in entries if int(row.id) not in used]
    if custom:
        if not available:
            raise HTTPException(status_code=409, detail="配对抽签已经完成")
        selected = _clicked_entry(available, entry_id, detail="点击的候选已不在当前配对池中，请按画面重新抽取")
        side = "first" if not pending else "second"
    elif seeded:
        required_side = "unseeded" if pending else "seeded"
        allowed = [row for row in available if row.seed_status == required_side]
        selected = _clicked_entry(allowed, entry_id, detail=f"当前必须抽取{'非种子队' if pending else '种子队'}，请按画面重新抽取")
        side = "unseeded" if pending else "seeded"
    else:
        if not available:
            raise HTTPException(status_code=409, detail="配对抽签已经完成")
        selected = _clicked_entry(available, entry_id, detail="点击的球队已不在当前配对池中，请按画面重新抽取")
        side = "second" if pending else "first"

    if not pending:
        pending_payload = {
            "entry_id": selected.id,
            "entity_name": selected.entity_name,
            "entity_type": selected.entity_type,
            "team_name": selected.team_name,
            "team_id": selected.team_id,
            "player_uid": selected.player_uid,
            "target_slot": pair_no,
            "side": side,
        }
        remaining = len(entries) - (len(picks) * 2 + 1)
        if custom and remaining == 0:
            _add_pick(
                db,
                record,
                operator,
                sequence_no=pair_no,
                entry=selected,
                target_slot=pair_no,
                namespace=f"custom:pair:{pair_no}:bye",
            )
            result = _result_from_progress(db, record)
            _complete(record, operator, result)
            return f"第{pair_no}组：{selected.entity_name}（轮空）", True
        result = _result_from_progress(db, record, pending_pair=pending_payload)
        _mark_drawing(record, operator, result)
        return f"第{pair_no}组第一方：{selected.entity_name}", False

    first = next((row for row in entries if row.id == int(pending["entry_id"])), None)
    if not first:
        raise HTTPException(status_code=409, detail="待配对候选已失效，请作废抽签后重新开始")
    namespace = f"custom:pair:{pair_no}" if custom else f"pair:{pair_no}"
    _add_pick(
        db,
        record,
        operator,
        sequence_no=pair_no,
        entry=first,
        paired_entry=selected,
        target_slot=pair_no,
        namespace=namespace,
    )
    result = _result_from_progress(db, record)
    complete = len(_active_picks(db, record.id)) * 2 >= len(entries)
    (_complete if complete else _mark_drawing)(record, operator, result)
    return f"第{pair_no}组：{first.entity_name} vs {selected.entity_name}", complete


def _draw_next_custom(db: Session, record: DrawSession, operator: str, entry_id: int) -> tuple[str, bool]:
    config = _loads(record.config_json)
    mode = str(config.get("mode") or "list")
    if mode == "pairs":
        return _draw_next_pair(db, record, operator, entry_id, seeded=False, custom=True)
    entries = [row for row in _entries(db, record.id) if row.is_active]
    picks = _active_picks(db, record.id)
    limit = int(config.get("result_count") or len(entries)) if mode == "list" else len(entries)
    if len(picks) >= limit:
        raise HTTPException(status_code=409, detail="自由抽签已经完成")
    available = [row for row in entries if int(row.id) not in _used_entry_ids(picks)]
    selected = _clicked_entry(available, entry_id, detail="点击的候选已不在当前可抽池中，请按画面重新抽取")
    sequence = len(picks) + 1
    target_group = None
    target_slot = sequence
    namespace = f"custom:list:{sequence}"
    label = f"第{sequence}位：{selected.entity_name}"
    if mode == "groups":
        group_count = int(config.get("group_count") or 2)
        target_group = chr(ord("A") + ((sequence - 1) % group_count))
        target_slot = sum(pick.target_group == target_group for pick in picks) + 1
        namespace = f"custom:group:{target_group}"
        label = f"{target_group}组第{target_slot}位：{selected.entity_name}"
    _add_pick(
        db,
        record,
        operator,
        sequence_no=sequence,
        entry=selected,
        target_group=target_group,
        target_slot=target_slot,
        namespace=namespace,
    )
    result = _result_from_progress(db, record)
    complete = len(_active_picks(db, record.id)) >= limit
    (_complete if complete else _mark_drawing)(record, operator, result)
    return label, complete


def draw_next(db: Session, operator: str, session_id: int, entry_id: int) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.status not in {"locked", "drawing"}:
        raise HTTPException(status_code=409, detail="候选池未锁定或抽签已经完成")
    kind = DRAW_TYPES[record.draw_type]["kind"]
    if kind == "weighted":
        return draw_next_lottery(db, operator, session_id, entry_id)
    if kind == "groups":
        summary, completed = _draw_next_group(db, record, operator, entry_id)
    elif kind == "seeded_pairs":
        summary, completed = _draw_next_pair(db, record, operator, entry_id, seeded=True)
    elif kind == "random_pairs":
        summary, completed = _draw_next_pair(db, record, operator, entry_id, seeded=False)
    elif kind == "custom":
        summary, completed = _draw_next_custom(db, record, operator, entry_id)
    else:
        raise HTTPException(status_code=400, detail="当前抽签类型不支持逐签抽取")
    db.commit()
    db.refresh(record)
    _audit(
        db,
        operator,
        "draw_next",
        f"{'完成抽签' if completed else '确定一签'}：{summary}",
        {"draw_session_id": record.id, "draw_type": record.draw_type, "completed": completed, "clicked_entry_id": entry_id, "selection_method": "click_frame"},
    )
    return _detail(db, record)


def _draw_custom(db: Session, record: DrawSession, operator: str) -> None:
    config = _loads(record.config_json)
    mode = str(config.get("mode") or "list")
    entries = sorted(
        [row for row in _entries(db, record.id) if row.is_active],
        key=lambda row: _hash_rank(record.random_seed, f"custom:{mode}", row.entity_key),
    )
    if mode == "groups":
        group_count = int(config.get("group_count") or 2)
        groups: dict[str, list[dict[str, Any]]] = {chr(ord("A") + index): [] for index in range(group_count)}
        for sequence, entry in enumerate(entries, start=1):
            group_name = chr(ord("A") + ((sequence - 1) % group_count))
            slot_no = len(groups[group_name]) + 1
            random_value = _hash_rank(record.random_seed, f"custom:group:{group_name}", entry.entity_key)
            db.add(DrawPick(session_id=record.id, sequence_no=sequence, entry_id=entry.id, target_group=group_name, target_slot=slot_no, random_value=random_value, created_by=operator, created_at=_now()))
            groups[group_name].append(_custom_result_entry(entry, slot_no=slot_no))
        _complete(record, operator, {"mode": mode, "groups": groups})
        return
    if mode == "pairs":
        pairs: list[dict[str, Any]] = []
        for pair_index, offset in enumerate(range(0, len(entries), 2), start=1):
            first = entries[offset]
            second = entries[offset + 1] if offset + 1 < len(entries) else None
            random_value = _hash_rank(record.random_seed, f"custom:pair:{pair_index}", f"{first.entity_key}:{second.entity_key if second else 'bye'}")
            db.add(DrawPick(session_id=record.id, sequence_no=pair_index, entry_id=first.id, paired_entry_id=second.id if second else None, target_slot=pair_index, random_value=random_value, created_by=operator, created_at=_now()))
            pairs.append({"slot_no": pair_index, "entry": _custom_result_entry(first), "paired_entry": _custom_result_entry(second) if second else None})
        _complete(record, operator, {"mode": mode, "pairs": pairs})
        return
    result_count = int(config.get("result_count") or len(entries))
    selected = entries[:result_count]
    items = []
    for sequence, entry in enumerate(selected, start=1):
        random_value = _hash_rank(record.random_seed, f"custom:list:{sequence}", entry.entity_key)
        db.add(DrawPick(session_id=record.id, sequence_no=sequence, entry_id=entry.id, target_slot=sequence, random_value=random_value, created_by=operator, created_at=_now()))
        items.append(_custom_result_entry(entry, slot_no=sequence))
    _complete(record, operator, {"mode": mode, "list": items, "selected_count": len(items), "candidate_count": len(entries)})


def execute_draw(db: Session, operator: str, session_id: int) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.status not in {"locked", "drawing"}:
        raise HTTPException(status_code=409, detail="候选池未锁定或抽签已经完成")
    raise HTTPException(status_code=409, detail="现场抽签必须逐个点击确定，不能一次性提前生成全部结果")


def _selected_team_ids(db: Session, session_id: int) -> set[int]:
    entry_by_id = {row.id: row for row in _entries(db, session_id)}
    return {
        int(entry_by_id[pick.entry_id].team_id)
        for pick in _picks(db, session_id)
        if pick.status == "active" and entry_by_id.get(pick.entry_id) and entry_by_id[pick.entry_id].team_id
    }


def draw_next_lottery(db: Session, operator: str, session_id: int, entry_id: int) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.draw_type != "lottery":
        raise HTTPException(status_code=400, detail="该接口仅用于乐透抽取")
    if record.status not in {"locked", "drawing"}:
        raise HTTPException(status_code=409, detail="乐透候选池未锁定或已经完成")
    config = _loads(record.config_json)
    limit = max(1, min(100, int(config.get("limit") or 15)))
    active_picks = [row for row in _picks(db, record.id) if row.status == "active"]
    if len(active_picks) >= limit:
        raise HTTPException(status_code=409, detail="已达到本届抽取人数")
    selected_teams = _selected_team_ids(db, record.id)
    available = [row for row in _entries(db, record.id) if row.is_active and row.team_id not in selected_teams]
    if not available:
        _complete(record, operator, {"selected_count": len(active_picks), "requested_count": limit, "completed_short": len(active_picks) < limit})
        db.commit()
        return _detail(db, record)
    sequence = len(_picks(db, record.id)) + 1
    selected = _clicked_entry(available, entry_id, detail="点击的球员已不在当前乐透池中，请按画面重新抽取")
    _add_pick(
        db,
        record,
        operator,
        sequence_no=sequence,
        entry=selected,
        namespace="lottery:click",
    )
    record.status = "drawing"
    record.updated_by = operator
    record.updated_at = _now()
    db.flush()
    active_count = len([row for row in _picks(db, record.id) if row.status == "active"])
    remaining_teams = {row.team_id for row in available if row.team_id != selected.team_id}
    if active_count >= limit or not remaining_teams:
        _complete(record, operator, {"selected_count": active_count, "requested_count": limit, "completed_short": active_count < limit})
    db.commit()
    db.refresh(record)
    _audit(db, operator, "lottery_pick", f"乐透第 {active_count} 位：{selected.entity_name}（{selected.team_name}）", {"draw_session_id": record.id, "player_uid": selected.player_uid, "weight": selected.weight, "clicked_entry_id": entry_id, "selection_method": "click_frame"})
    return _detail(db, record)


def invalidate_lottery_pick(db: Session, operator: str, session_id: int, pick_id: int, reason: str) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.draw_type != "lottery" or record.status not in {"drawing", "completed"}:
        raise HTTPException(status_code=409, detail="当前乐透不能作废单个结果")
    pick = db.query(DrawPick).filter(DrawPick.id == pick_id, DrawPick.session_id == record.id, DrawPick.status == "active").first()
    if not pick:
        raise HTTPException(status_code=404, detail="有效抽取结果不存在")
    clean_reason = str(reason or "").strip()
    if not clean_reason:
        raise HTTPException(status_code=400, detail="请填写作废原因")
    pick.status = "invalidated"
    pick.reason = clean_reason
    pick.invalidated_by = operator
    pick.invalidated_at = _now()
    record.status = "drawing"
    record.completed_at = None
    record.completed_by = None
    record.updated_by = operator
    record.updated_at = _now()
    db.commit()
    _audit(db, operator, "invalidate_pick", f"作废乐透结果：{clean_reason}", {"draw_session_id": record.id, "pick_id": pick.id})
    return _detail(db, record)


def publish_session(db: Session, operator: str, session_id: int) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.status != "completed":
        raise HTTPException(status_code=409, detail="只有已完成抽签可以发布")
    record.status = "published"
    record.published_by = operator
    record.published_at = _now()
    record.updated_by = operator
    record.updated_at = _now()
    db.commit()
    _audit(db, operator, "publish", f"发布抽签结果：{record.name}", {"draw_session_id": record.id})
    return _detail(db, record)


def void_session(db: Session, operator: str, session_id: int, reason: str) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.status == "void":
        return _detail(db, record)
    clean_reason = str(reason or "").strip()
    if not clean_reason:
        raise HTTPException(status_code=400, detail="请填写作废原因")
    record.status = "void"
    record.void_reason = clean_reason
    record.voided_by = operator
    record.voided_at = _now()
    record.updated_by = operator
    record.updated_at = _now()
    db.commit()
    _audit(db, operator, "void", f"作废抽签：{record.name}", {"draw_session_id": record.id, "reason": clean_reason})
    return _detail(db, record)


def build_lottery_pool(db: Session, operator: str, session_id: int, request: LotteryPoolBuildRequest) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.draw_type != "lottery" or record.status != "draft":
        raise HTTPException(status_code=409, detail="只有乐透草稿可以重新筛选候选池")
    team_ids = {int(value) for value in request.team_ids}
    query = db.query(Player).filter(Player.team_id.isnot(None))
    if team_ids:
        query = query.filter(Player.team_id.in_(team_ids))
    excluded = {int(value) for value in request.excluded_player_uids}
    if request.excluded_candidate_list_ids:
        rows = (
            db.query(CandidateListPlayer.uid)
            .join(CandidateList, CandidateList.id == CandidateListPlayer.list_id)
            .filter(
                CandidateList.id.in_(request.excluded_candidate_list_ids),
                CandidateList.status == "published",
                CandidateListPlayer.removed_at.is_(None),
            )
            .all()
        )
        excluded.update(int(row[0]) for row in rows)
    excluded.difference_update(int(value) for value in request.restored_player_uids)
    entries: list[DrawPoolEntryRequest] = []
    for player in query.order_by(Player.team_name, Player.name, Player.uid).all():
        wage = calculate_wage(player.initial_ca or 0, player.ca or 0, player.pa or 0, player.age or 0, player.position or "")
        if float(wage["final_value"]) < float(request.min_final_value) or player.uid in excluded:
            continue
        saves = int(request.self_save_counts.get(str(player.uid), 0) or 0)
        entries.append(DrawPoolEntryRequest(player_uid=player.uid, self_save_count=max(0, saves), metadata={"position": player.position or "", "age": player.age, "ca": player.ca, "pa": player.pa}))
    _replace_entries(db, record, entries)
    record.config_json = _dumps({
        "min_final_value": float(request.min_final_value),
        "team_ids": sorted(team_ids),
        "excluded_candidate_list_ids": sorted(set(request.excluded_candidate_list_ids)),
        "excluded_player_uids": sorted(excluded),
        "restored_player_uids": sorted(set(request.restored_player_uids)),
        "limit": int(request.limit),
    })
    record.updated_by = operator
    record.updated_at = _now()
    db.commit()
    db.refresh(record)
    _audit(db, operator, "build_lottery_pool", f"生成乐透候选池：{len(entries)} 名球员", {"draw_session_id": record.id, "team_count": len({row.team_id for row in _entries(db, record.id) if row.team_id})})
    return _detail(db, record)


def create_lottery_candidate_list(db: Session, operator: str, session_id: int) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.draw_type != "lottery" or record.status not in {"completed", "published"}:
        raise HTTPException(status_code=409, detail="乐透完成后才能生成候选名单草稿")
    if record.candidate_list_id:
        return _detail(db, record)
    entries = {row.id: row for row in _entries(db, record.id)}
    selected = [entries[row.entry_id] for row in _picks(db, record.id) if row.status == "active" and row.entry_id in entries]
    version = resolve_attribute_version(db, None)
    now = _now()
    candidate = CandidateList(
        name=f"{record.name}｜乐透抽取名单",
        description="由抽签模块生成的候选名单草稿；保留抽取顺序，发布前可继续补充或调整。",
        type="lottery",
        status="draft",
        base_data_version=version,
        source_filters_json=_dumps({"draw_session_id": record.id, "pool_hash": record.pool_hash}),
        created_by=operator,
        updated_by=operator,
        created_at=now,
        updated_at=now,
    )
    db.add(candidate)
    db.flush()
    for entry in selected:
        player = db.query(Player).filter(Player.uid == entry.player_uid).first()
        db.add(CandidateListPlayer(
            list_id=candidate.id,
            uid=int(entry.player_uid),
            data_version=version,
            name_snapshot=entry.entity_name,
            club_snapshot="",
            heigo_club_snapshot=entry.team_name or "",
            ca_snapshot=int(player.ca or 0) if player else 0,
            pa_snapshot=int(player.pa or 0) if player else 0,
            added_by=operator,
            added_at=now,
        ))
    record.candidate_list_id = candidate.id
    result = _loads(record.result_json)
    result["candidate_list_id"] = candidate.id
    result["selection_order"] = [entry.player_uid for entry in selected]
    record.result_json = _dumps(result)
    record.updated_by = operator
    record.updated_at = now
    db.commit()
    _audit(db, operator, "candidate_list", f"生成乐透候选名单草稿：{candidate.name}", {"draw_session_id": record.id, "candidate_list_id": candidate.id})
    return _detail(db, record)


def write_to_cup(db: Session, operator: str, session_id: int) -> DrawSessionDetailResponse:
    record = _session_or_404(db, session_id)
    if record.status not in {"completed", "published"} or not record.competition:
        raise HTTPException(status_code=409, detail="杯赛抽签完成后才能写入赛程")
    picks = [row for row in _picks(db, record.id) if row.status == "active"]
    entries = {row.id: row for row in _entries(db, record.id)}
    definition = DRAW_TYPES[record.draw_type]
    if definition["kind"] == "groups":
        existing_matches = db.query(CupMatch).filter(CupMatch.competition == record.competition, CupMatch.stage.like("group_%"), CupMatch.status == "played").count()
        if existing_matches:
            raise HTTPException(status_code=409, detail="杯赛小组已有比分，必须先走杯赛重置流程")
        db.query(CupGroupTeam).filter(CupGroupTeam.competition == record.competition).delete(synchronize_session=False)
        for pick in picks:
            entry = entries[pick.entry_id]
            group_no = ord(str(pick.target_group)) - ord("A") + 1
            db.add(CupGroupTeam(competition=record.competition, group_no=group_no, slot_no=int(pick.target_slot), team_id=entry.team_id, team_name=entry.team_name, created_at=_now(), updated_at=_now()))
        db.commit()
        cup_service.get_group_stage(db, record.competition)
    else:
        stage = "qualifying_round" if record.draw_type == "wumingjian_qualifying" else "round_of_32" if record.draw_type == "wumingjian_r32" else "round_of_16"
        if record.draw_type == "wumingjian_qualifying":
            cup_service._lock_wumingjian_qualification(db)
        played = db.query(CupMatch).filter(CupMatch.competition == record.competition, CupMatch.stage == stage, CupMatch.status == "played").count()
        if played:
            raise HTTPException(status_code=409, detail="该杯赛阶段已有比赛结果，必须先走单独重置流程")
        cup_service.ensure_bracket(db, record.competition)
        for pick in picks:
            first = entries[pick.entry_id]
            second = entries[pick.paired_entry_id] if pick.paired_entry_id else None
            match = db.query(CupMatch).filter(CupMatch.competition == record.competition, CupMatch.stage == stage, CupMatch.slot_no == pick.target_slot).first()
            if not match or not second:
                continue
            match.home_team_id, match.home_team_name = first.team_id, first.team_name
            match.away_team_id, match.away_team_name = second.team_id, second.team_name
            match.home_score = match.away_score = None
            match.winner_team_id = None
            match.winner_team_name = None
            match.status = "scheduled"
            match.notes = "抽签只确定对阵；首回合主场由双方协商。网站录入两回合总比分。"
            match.updated_at = _now()
        db.commit()
    result = _loads(record.result_json)
    result["written_to_cup_at"] = _now().isoformat()
    result["written_to_cup_by"] = operator
    record.result_json = _dumps(result)
    record.updated_by = operator
    record.updated_at = _now()
    db.commit()
    _audit(db, operator, "write_to_cup", f"抽签结果写入{cup_service.CUP_DEFINITIONS[record.competition]['title']}：{record.name}", {"draw_session_id": record.id})
    return _detail(db, record)


def _archive_response(record: SeasonArchive) -> SeasonArchiveDetailResponse:
    return SeasonArchiveDetailResponse(
        id=record.id,
        season_key=record.season_key,
        title=record.title,
        revision_no=record.revision_no,
        parent_archive_id=record.parent_archive_id,
        status=record.status,
        revision_reason=record.revision_reason,
        created_by=record.created_by,
        confirmed_by=record.confirmed_by,
        created_at=record.created_at,
        confirmed_at=record.confirmed_at,
        snapshot=_loads(record.snapshot_json),
        validation=_loads(record.validation_json),
    )


def list_season_archives(db: Session, *, public: bool = False) -> list[SeasonArchiveSummaryResponse]:
    query = db.query(SeasonArchive)
    if public:
        query = query.filter(SeasonArchive.status == "confirmed")
    rows = query.order_by(SeasonArchive.season_key.desc(), SeasonArchive.revision_no.desc()).all()
    return [SeasonArchiveSummaryResponse(**_archive_response(row).model_dump(exclude={"snapshot", "validation"})) for row in rows]


def get_season_archive(db: Session, archive_id: int, *, public: bool = False) -> SeasonArchiveDetailResponse:
    query = db.query(SeasonArchive).filter(SeasonArchive.id == int(archive_id))
    if public:
        query = query.filter(SeasonArchive.status == "confirmed")
    record = query.first()
    if not record:
        raise HTTPException(status_code=404, detail="赛季档案不存在")
    return _archive_response(record)


def _build_archive_snapshot(db: Session, season_key: str) -> tuple[dict[str, Any], dict[str, Any]]:
    standings = match_service.get_standings(db, include_predictions=False)
    standings_rows = [row.model_dump(mode="json") for row in standings.rows]
    player_rankings = player_ranking_service.get_player_rankings(db)
    cup_champions: dict[str, dict[str, Any] | None] = {}
    blockers: list[str] = []
    for competition, definition in cup_service.CUP_DEFINITIONS.items():
        final = db.query(CupMatch).filter(CupMatch.competition == competition, CupMatch.stage == "final", CupMatch.slot_no == 1).first()
        if not final or final.status != "played" or not final.winner_team_id:
            cup_champions[competition] = None
            blockers.append(f"{definition['title']}尚未产生冠军")
        else:
            cup_champions[competition] = {"team_id": final.winner_team_id, "team_name": final.winner_team_name}
    league_completion: dict[str, Any] = {}
    standings_anomalies: list[str] = []
    for level in ("超级", "甲级", "乙级"):
        rows = [row for row in standings.rows if row.level == level]
        expected_played = max((int(row.played) for row in rows), default=0)
        complete = bool(rows) and len({int(row.played) for row in rows}) == 1 and expected_played >= 34
        league_completion[level] = {"complete": complete, "team_count": len(rows), "played_per_team": expected_played}
        if not complete:
            blockers.append(f"{level}联赛尚未全部完成34场")
        expected_ranks = list(range(1, len(rows) + 1))
        actual_ranks = sorted(int(row.rank) for row in rows)
        if actual_ranks != expected_ranks:
            standings_anomalies.append(f"{level}排名序号不连续")
        if any(int(row.played) != int(row.wins) + int(row.draws) + int(row.losses) for row in rows):
            standings_anomalies.append(f"{level}存在场次与胜平负不一致")
    if standings_anomalies:
        blockers.extend(standings_anomalies)
    missing_events = sum(int(row.matches_missing_events) for row in player_rankings.coverage)
    if missing_events:
        blockers.append(f"仍有 {missing_events} 场已赛比赛缺少球员事件")
    teams = db.query(Team).filter(Team.level != "隐藏").order_by(Team.level, Team.name).all()
    snapshot = {
        "season_key": season_key,
        "standings": standings_rows,
        "player_rankings": player_rankings.model_dump(mode="json"),
        "cup_champions": cup_champions,
        "teams": [{"id": row.id, "name": row.name, "manager": row.manager, "level": row.level, "logo_path": row.logo_path} for row in teams],
        "captured_at": _now().isoformat(),
    }
    validation = {"ready": not blockers, "blockers": blockers, "league_completion": league_completion, "standings_anomalies": standings_anomalies, "missing_player_event_matches": missing_events}
    return snapshot, validation


def create_season_archive(db: Session, operator: str, request: SeasonArchiveCreateRequest) -> SeasonArchiveDetailResponse:
    season_key = str(request.season_key or "").strip()
    if not season_key:
        raise HTTPException(status_code=400, detail="赛季标识不能为空")
    parent = None
    if request.revision_of:
        parent = db.query(SeasonArchive).filter(SeasonArchive.id == int(request.revision_of), SeasonArchive.status == "confirmed").first()
        if not parent:
            raise HTTPException(status_code=404, detail="原赛季档案不存在")
        if not str(request.revision_reason or "").strip():
            raise HTTPException(status_code=400, detail="修订版必须填写修订原因")
        season_key = parent.season_key
    latest_revision = max([row[0] for row in db.query(SeasonArchive.revision_no).filter(SeasonArchive.season_key == season_key).all()] or [0])
    snapshot, validation = _build_archive_snapshot(db, season_key)
    if request.confirm and not validation["ready"]:
        raise HTTPException(status_code=409, detail="赛季尚不满足封存条件：" + "；".join(validation["blockers"]))
    record = SeasonArchive(
        season_key=season_key,
        title=str(request.title or "").strip() or f"HEIGO {season_key} 赛季档案",
        revision_no=latest_revision + 1,
        parent_archive_id=parent.id if parent else None,
        status="confirmed" if request.confirm else "draft",
        snapshot_json=_dumps(snapshot),
        validation_json=_dumps(validation),
        revision_reason=str(request.revision_reason or "").strip() or None,
        created_by=operator,
        confirmed_by=operator if request.confirm else None,
        created_at=_now(),
        confirmed_at=_now() if request.confirm else None,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    _audit(db, operator, "archive", f"{'封存' if request.confirm else '生成'}赛季档案：{record.title} v{record.revision_no}", {"season_archive_id": record.id, "ready": validation["ready"]})
    return _archive_response(record)


def confirm_season_archive(db: Session, operator: str, archive_id: int) -> SeasonArchiveDetailResponse:
    record = db.query(SeasonArchive).filter(SeasonArchive.id == int(archive_id)).first()
    if not record:
        raise HTTPException(status_code=404, detail="赛季档案不存在")
    validation = _loads(record.validation_json)
    if not validation.get("ready"):
        raise HTTPException(status_code=409, detail="赛季尚不满足封存条件：" + "；".join(validation.get("blockers") or []))
    record.status = "confirmed"
    record.confirmed_by = operator
    record.confirmed_at = _now()
    db.commit()
    _audit(db, operator, "confirm_archive", f"确认封存赛季档案：{record.title}", {"season_archive_id": record.id})
    return _archive_response(record)


def _latest_archive_snapshot(db: Session) -> dict[str, Any] | None:
    record = db.query(SeasonArchive).filter(SeasonArchive.status == "confirmed").order_by(SeasonArchive.confirmed_at.desc(), SeasonArchive.id.desc()).first()
    return _loads(record.snapshot_json) if record else None


def propose_pool(db: Session, draw_type: str) -> dict[str, Any]:
    if draw_type not in DRAW_TYPES:
        raise HTTPException(status_code=404, detail="抽签类型不存在")
    if draw_type in {"champions_r16", "league_r16"}:
        competition = DRAW_TYPES[draw_type]["competition"]
        stage = cup_service.get_group_stage(db, competition)
        qualifiers = stage.champions_knockout_qualifiers if draw_type == "champions_r16" else stage.league_knockout_qualifiers
        if not stage.qualification_complete:
            raise HTTPException(status_code=409, detail="小组赛尚未全部完成")
        seeded_ids: set[int] = set()
        if draw_type == "champions_r16":
            group_winners = [row for row in qualifiers if row.group_rank == 1]
            others = sorted([row for row in qualifiers if row.group_rank != 1], key=lambda row: (row.group_rank, -row.points, -row.goal_difference, -row.goals_for, -row.wins, row.team_name))
            seeded_ids = {row.team_id for row in group_winners + others[:3]}
        else:
            seeded_ids = {row.team_id for row in qualifiers if row.group_rank == 1 or row.source_competition == "champions_cup"}
        return {"draw_type": draw_type, "entries": [{"team_id": row.team_id, "entity_name": row.team_name, "seed_status": "seeded" if row.team_id in seeded_ids else "unseeded", "metadata": row.model_dump(mode="json")} for row in qualifiers]}
    if draw_type in {"wumingjian_qualifying", "wumingjian_r32"}:
        qualification = cup_service.get_wumingjian_qualification(db)
        teams = qualification.preliminary_eligible_teams if draw_type == "wumingjian_qualifying" else qualification.direct_qualifiers + qualification.preliminary_winners
        return {"draw_type": draw_type, "entries": [{"team_id": row.team_id, "entity_name": row.team_name, "level": row.level, "source_rank": row.source_rank} for row in teams]}
    if draw_type == "lottery":
        return {"draw_type": draw_type, "entries": []}
    if draw_type in CUSTOM_DRAW_TYPES:
        return {"draw_type": draw_type, "entries": []}
    snapshot = _latest_archive_snapshot(db)
    if snapshot:
        archived_standings = snapshot.get("standings") or []
        champions = snapshot.get("cup_champions") or {}
    else:
        archived_standings = [row.model_dump(mode="json") for row in match_service.get_standings(db, include_predictions=False).rows]
        champions = {}
    current_teams = db.query(Team).filter(Team.level != "隐藏").all()
    current_by_id = {int(row.id): row for row in current_teams}
    current_by_name = {str(row.name): row for row in current_teams}
    standings = []
    unmatched = []
    for archived in archived_standings:
        current = current_by_id.get(int(archived.get("team_id") or 0)) or current_by_name.get(str(archived.get("team_name") or ""))
        if not current:
            unmatched.append({"team_id": archived.get("team_id"), "team_name": archived.get("team_name"), "level": archived.get("level"), "rank": archived.get("rank")})
            continue
        standings.append({**archived, "team_id": current.id, "team_name": current.name})
    order = []
    for level in ("超级", "甲级", "乙级"):
        order.extend(sorted([row for row in standings if row.get("level") == level], key=lambda row: int(row.get("rank") or 999)))
    by_id = {int(row["team_id"]): row for row in order if row.get("team_id")}
    by_name = {str(row.get("team_name")): row for row in order}
    first_ids: list[int] = []
    for source in ("league_super", "league_first", "wumingjian_cup", "league_cup", "champions_cup"):
        if source == "league_super":
            row = next((item for item in order if item.get("level") == "超级" and int(item.get("rank") or 0) == 1), None)
        elif source == "league_first":
            row = next((item for item in order if item.get("level") == "甲级" and int(item.get("rank") or 0) == 1), None)
        else:
            champion = champions.get(source) or {}
            row = by_id.get(int(champion.get("team_id") or 0)) or by_name.get(str(champion.get("team_name") or ""))
        if row and int(row["team_id"]) not in first_ids:
            first_ids.append(int(row["team_id"]))
    for row in order:
        if len(first_ids) >= 5:
            break
        if int(row["team_id"]) not in first_ids:
            first_ids.append(int(row["team_id"]))
    remaining = [row for row in order if int(row["team_id"]) not in first_ids]
    second_division_champion = next((row for row in remaining if row.get("level") == "乙级" and int(row.get("rank") or 0) == 1), None)
    ordered_ids = list(first_ids)
    if second_division_champion:
        ordered_ids.append(int(second_division_champion["team_id"]))
    ordered_ids.extend(
        int(row["team_id"])
        for row in remaining
        if not second_division_champion or int(row["team_id"]) != int(second_division_champion["team_id"])
    )
    take = 30 if draw_type == "champions_group" else 54
    selected_ids = ordered_ids[:take]
    if draw_type == "league_group":
        selected_ids = [team_id for team_id in ordered_ids if team_id not in set(ordered_ids[:30])][:24]
    group_count = DRAW_TYPES[draw_type]["group_count"]
    entries = []
    for index, team_id in enumerate(selected_ids):
        row = by_id[team_id]
        entries.append({"team_id": team_id, "entity_name": row.get("team_name"), "level": row.get("level"), "source_rank": row.get("rank"), "pot_no": (index // group_count) + 1})
    return {"draw_type": draw_type, "archive_available": bool(snapshot), "unmatched": unmatched, "resolution": "qualification_progression", "entries": entries}


def export_text(db: Session, session_id: int, *, public: bool = False) -> str:
    detail = get_session(db, session_id, public=public)
    lines = [detail.name, f"类型：{detail.draw_type}", f"随机种子：{detail.random_seed}", f"候选池哈希：{detail.pool_hash or '-'}", ""]
    if detail.result.get("groups"):
        for group, teams in detail.result["groups"].items():
            lines.append(f"{group}组")
            for index, row in enumerate(teams, start=1):
                name = row.get("entity_name") or row.get("team_name") or "-"
                prefix = f"第{row['pot_no']}档" if row.get("pot_no") else f"#{row.get('slot_no') or index}"
                team = row.get("team_name") if row.get("entity_type") == "player" else None
                lines.append(f"  {prefix} {name}{f'｜{team}' if team else ''}")
    else:
        for pick in detail.picks:
            if pick.status != "active":
                continue
            if pick.paired_entry:
                lines.append(f"{pick.target_slot}. {pick.entry.entity_name} vs {pick.paired_entry.entity_name}")
            elif detail.draw_type in CUSTOM_DRAW_TYPES:
                suffix = f"｜{pick.entry.team_name}" if pick.entry.entity_type == "player" and pick.entry.team_name else ""
                bye = "｜轮空" if detail.result.get("mode") == "pairs" else ""
                lines.append(f"{pick.sequence_no}. {pick.entry.entity_name}{suffix}{bye}")
            else:
                lines.append(f"{pick.sequence_no}. {pick.entry.entity_name}｜{pick.entry.team_name or ''}｜{pick.entry.final_value or '-'}M｜权重×{pick.entry.weight:g}")
    return "\n".join(lines)


def export_excel(db: Session, session_id: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    detail = get_session(db, session_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "抽签结果"
    sheet.append([detail.name])
    sheet.append(["抽签类型", detail.draw_type, "随机种子", detail.random_seed])
    sheet.append(["候选池哈希", detail.pool_hash or ""])
    sheet.append([])
    sheet.append(["顺序", "分组/签位", "球队/球员", "所属球队/对手", "档位/种子", "身价", "权重", "状态"])
    for pick in detail.picks:
        sheet.append([
            pick.sequence_no,
            pick.target_group or pick.target_slot or "",
            pick.entry.entity_name,
            pick.paired_entry.entity_name if pick.paired_entry else pick.entry.team_name or "",
            pick.entry.pot_no or pick.entry.seed_status or "",
            pick.entry.final_value if pick.entry.final_value is not None else "",
            pick.entry.weight,
            pick.status,
        ])
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[5]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.column_dimensions["A"].width = 10
    for column in ("B", "C", "D", "E"):
        sheet.column_dimensions[column].width = 24
    for column in ("F", "G", "H"):
        sheet.column_dimensions[column].width = 12
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_png(db: Session, session_id: int) -> bytes:
    from PIL import Image, ImageDraw, ImageFont

    text = export_text(db, session_id)
    lines = text.splitlines()
    width = 1200
    height = max(720, 180 + len(lines) * 42)
    image = Image.new("RGB", (width, height), "#0b1020")
    draw = ImageDraw.Draw(image)
    font_paths = ["/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
    font_path = next((path for path in font_paths if __import__("os").path.exists(path)), font_paths[-1])
    title_font = ImageFont.truetype(font_path, 44)
    body_font = ImageFont.truetype(font_path, 26)
    draw.rounded_rectangle((42, 42, width - 42, height - 42), radius=30, fill="#131a2f", outline="#4f46e5", width=3)
    y = 82
    for index, line in enumerate(lines):
        font = title_font if index == 0 else body_font
        fill = "#f8fafc" if index == 0 else "#cbd5e1"
        draw.text((82, y), line, font=font, fill=fill)
        y += 62 if index == 0 else 42
    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True)
    return output.getvalue()
