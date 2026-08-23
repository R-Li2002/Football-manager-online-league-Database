import json
import io
from datetime import datetime
from typing import Any

from fastapi import HTTPException
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from models import CandidateList, CandidateListPlayer, Player
from repositories.attribute_repository import (
    ATTRIBUTE_RANGE_FIELD_ALLOWLIST,
    POSITION_SCORE_FIELD_ALLOWLIST,
    get_attribute_model_for_versions,
    list_available_attribute_versions,
    resolve_attribute_version,
)
from repositories.player_repository import league_player_membership_filter, map_player_uid_to_team_name
from schemas_read import (
    CandidateListDetailResponse,
    CandidateListMutationResponse,
    CandidateListPlayerPreviewResponse,
    CandidateListPlayerResponse,
    CandidateListPlayersResponse,
    CandidateListPreviewCandidateResponse,
    CandidateListPreviewTokenResponse,
    CandidateListPublishPreviewResponse,
    CandidateListRemovePreviewResponse,
    CandidateListSummaryResponse,
)
from schemas_write import (
    CandidateListBatchRemoveRequest,
    CandidateListPlayerCommitRequest,
    CandidateListPlayerPreviewRequest,
    CandidateListUpsertRequest,
)
from search_normalization import build_search_normalized_keys
from services.operation_audit_service import AUDIT_SOURCE_ADMIN_UI, persist_admin_operation_audit


CANDIDATE_LIST_TYPES = {"transfer", "recommendation", "review", "custom", "lottery"}
CANDIDATE_LIST_STATUSES = {"draft", "published", "archived"}
ACTIVE_PLAYER_LIMIT = 5000


def _now() -> datetime:
    return datetime.now()


def _json_loads(value: str | None) -> dict[str, Any]:
    if not value:
        return {}
    try:
        payload = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def _json_dumps(value: dict[str, Any] | None) -> str:
    return json.dumps(value or {}, ensure_ascii=False, sort_keys=True)


def _active_rows_query(db: Session, list_id: int):
    return (
        db.query(CandidateListPlayer)
        .filter(CandidateListPlayer.list_id == list_id)
        .filter(CandidateListPlayer.removed_at.is_(None))
    )


def _active_rows(db: Session, list_id: int) -> list[CandidateListPlayer]:
    return _active_rows_query(db, list_id).order_by(CandidateListPlayer.added_at.asc(), CandidateListPlayer.id.asc()).all()


def _active_count(db: Session, list_id: int) -> int:
    return _active_rows_query(db, list_id).count()


def _active_exists(db: Session, list_id: int, uid: int, data_version: str | None = None) -> bool:
    return (
        _active_rows_query(db, list_id)
        .filter(CandidateListPlayer.uid == int(uid))
        .first()
        is not None
    )


def _candidate_list_or_404(db: Session, list_id: int, *, public: bool = False) -> CandidateList:
    query = db.query(CandidateList).filter(CandidateList.id == list_id)
    if public:
        query = query.filter(CandidateList.status == "published", CandidateList.archived_at.is_(None))
    record = query.first()
    if not record:
        raise HTTPException(status_code=404, detail="候选名单不存在")
    return record


def _audit(db: Session, *, action: str, operator: str | None, summary: str, details: dict[str, Any] | None = None) -> None:
    bind = db.get_bind()
    if bind is None:
        return
    persist_admin_operation_audit(
        bind,
        category="candidate_list",
        action=action,
        operator=operator,
        status="success",
        summary=summary,
        source=AUDIT_SOURCE_ADMIN_UI,
        operation_label="候选名单",
        extra_details=details or {},
    )


def _serialize_summary(db: Session, record: CandidateList) -> CandidateListSummaryResponse:
    return CandidateListSummaryResponse(
        id=record.id,
        name=record.name,
        description=record.description or "",
        type=record.type or "custom",
        status=record.status or "draft",
        base_data_version=record.base_data_version or "",
        player_count=_active_count(db, record.id),
        published_player_count=int(record.published_player_count or 0),
        created_by=record.created_by,
        updated_by=record.updated_by,
        created_at=record.created_at,
        updated_at=record.updated_at,
        published_at=record.published_at,
        published_by=record.published_by,
        archived_at=record.archived_at,
        locked_at=record.locked_at,
    )


def _serialize_detail(db: Session, record: CandidateList) -> CandidateListDetailResponse:
    summary = _serialize_summary(db, record)
    return CandidateListDetailResponse(
        **summary.model_dump(),
        source_filters=_json_loads(record.source_filters_json),
        last_published_snapshot=_json_loads(record.last_published_snapshot_json),
    )


def list_public_candidate_lists(db: Session) -> list[CandidateListSummaryResponse]:
    records = (
        db.query(CandidateList)
        .filter(CandidateList.status == "published", CandidateList.archived_at.is_(None))
        .order_by(CandidateList.published_at.desc(), CandidateList.id.desc())
        .all()
    )
    return [_serialize_summary(db, record) for record in records]


def list_admin_candidate_lists(db: Session) -> list[CandidateListSummaryResponse]:
    records = db.query(CandidateList).order_by(CandidateList.updated_at.desc(), CandidateList.id.desc()).all()
    return [_serialize_summary(db, record) for record in records]


def get_candidate_list(db: Session, list_id: int, *, public: bool = False) -> CandidateListDetailResponse:
    return _serialize_detail(db, _candidate_list_or_404(db, list_id, public=public))


def _attribute_base_query(db: Session, data_version: str):
    versions = list_available_attribute_versions(db)
    model = get_attribute_model_for_versions(versions)
    query = db.query(model)
    if versions:
        query = query.filter(model.data_version == data_version)
    return model, query


def _query_attributes_by_uids(db: Session, uids: list[int], data_version: str) -> dict[int, Any]:
    if not uids:
        return {}
    model, query = _attribute_base_query(db, data_version)
    rows = query.filter(model.uid.in_(list({int(uid) for uid in uids}))).all()
    return {int(row.uid): row for row in rows}


def _preview_candidate_from_attr(attr: Any, *, data_version: str, heigo_club: str = "") -> CandidateListPreviewCandidateResponse:
    return CandidateListPreviewCandidateResponse(
        uid=int(attr.uid),
        name=str(attr.name or ""),
        data_version=str(getattr(attr, "data_version", data_version) or data_version),
        position=str(attr.position or ""),
        age=attr.age,
        ca=attr.ca,
        pa=attr.pa,
        club=str(attr.club or ""),
        heigo_club=heigo_club,
    )


def _player_response_from_row(
    row: CandidateListPlayer,
    attr: Any | None,
    *,
    view_version: str,
    heigo_club: str,
) -> CandidateListPlayerResponse:
    if attr:
        return CandidateListPlayerResponse(
            uid=int(attr.uid),
            data_version=str(getattr(attr, "data_version", view_version) or view_version),
            name=str(attr.name or row.name_snapshot or ""),
            position=str(attr.position or ""),
            age=attr.age,
            ca=attr.ca,
            pa=attr.pa,
            nationality=str(attr.nationality or ""),
            club=str(attr.club or row.club_snapshot or ""),
            heigo_club=heigo_club or row.heigo_club_snapshot or "",
            missing=False,
            added_at=row.added_at,
        )
    return CandidateListPlayerResponse(
        uid=int(row.uid),
        data_version=view_version,
        name=row.name_snapshot or f"UID {row.uid}",
        position="",
        age=None,
        ca=row.ca_snapshot,
        pa=row.pa_snapshot,
        nationality="",
        club=row.club_snapshot or "",
        heigo_club=row.heigo_club_snapshot or "",
        missing=True,
        added_at=row.added_at,
    )


def get_candidate_list_players(
    db: Session,
    list_id: int,
    *,
    version: str | None = None,
    limit: int = 500,
    offset: int = 0,
    public: bool = False,
) -> CandidateListPlayersResponse:
    record = _candidate_list_or_404(db, list_id, public=public)
    view_version = resolve_attribute_version(db, version or record.base_data_version)
    active_rows = _active_rows(db, list_id)
    total_count = len(active_rows)
    attrs = _query_attributes_by_uids(db, [row.uid for row in active_rows], view_version)
    heigo_players = map_player_uid_to_team_name(db)
    missing_count = len([row for row in active_rows if row.uid not in attrs])
    normalized_limit = max(1, min(1000, int(limit or 500)))
    normalized_offset = max(0, int(offset or 0))
    page_rows = active_rows[normalized_offset : normalized_offset + normalized_limit]
    return CandidateListPlayersResponse(
        list_id=record.id,
        name=record.name,
        data_version=view_version,
        total_count=total_count,
        matched_count=total_count - missing_count,
        missing_count=missing_count,
        limit=normalized_limit,
        offset=normalized_offset,
        items=[
            _player_response_from_row(
                row,
                attrs.get(row.uid),
                view_version=view_version,
                heigo_club=heigo_players.get(row.uid, row.heigo_club_snapshot or ""),
            )
            for row in page_rows
        ],
    )


def build_candidate_list_excel(
    db: Session,
    list_id: int,
    *,
    public: bool = False,
) -> tuple[io.BytesIO, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    record = _candidate_list_or_404(db, list_id, public=public)
    data_version = resolve_attribute_version(db, record.base_data_version)
    active_rows = _active_rows(db, list_id)
    uids = [int(row.uid) for row in active_rows]
    attributes = _query_attributes_by_uids(db, uids, data_version)
    league_players = {
        int(player.uid): player
        for player in (
            db.query(Player)
            .filter(Player.uid.in_(uids), league_player_membership_filter())
            .all()
            if uids
            else []
        )
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候选名单"
    headers = [
        "序号",
        "UID",
        "球员姓名",
        "位置",
        "年龄",
        "国籍",
        "现实俱乐部",
        "HEIGO球队",
        "初始CA",
        "当前CA",
        "当前PA",
        "当前数据来源",
        "数据库版本",
        "加入时间",
    ]
    last_column = get_column_letter(len(headers))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = record.name
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="111827")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = record.description or "HEIGO 候选名单"
    sheet["A2"].font = Font(color="475569")
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = f"数据库版本：{data_version or '-'}｜当前CA、当前PA优先取联赛名单，不在联赛名单时使用球员数据库"
    sheet["A3"].font = Font(size=10, color="64748B")
    sheet["A3"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.append([])
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="25324A")
    header_border = Border(bottom=Side(style="thin", color="5B6B86"))
    for cell in sheet[5]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    for index, candidate_row in enumerate(active_rows, start=1):
        uid = int(candidate_row.uid)
        attr = attributes.get(uid)
        league_player = league_players.get(uid)
        database_ca = getattr(attr, "ca", None) if attr is not None else candidate_row.ca_snapshot
        database_pa = getattr(attr, "pa", None) if attr is not None else candidate_row.pa_snapshot
        league_ca = getattr(league_player, "ca", None) if league_player is not None else None
        league_pa = getattr(league_player, "pa", None) if league_player is not None else None
        current_ca = league_ca if league_ca is not None else database_ca
        current_pa = league_pa if league_pa is not None else database_pa
        if league_player is None:
            current_source = "球员数据库"
        elif league_ca is None or league_pa is None:
            current_source = "联赛名单/数据库回退"
        else:
            current_source = "联赛名单"
        added_at = candidate_row.added_at.strftime("%Y-%m-%d %H:%M:%S") if candidate_row.added_at else ""
        sheet.append([
            index,
            uid,
            str(getattr(attr, "name", None) or getattr(league_player, "name", None) or candidate_row.name_snapshot or f"UID {uid}"),
            str(getattr(attr, "position", None) or getattr(league_player, "position", None) or ""),
            getattr(attr, "age", None) if attr is not None else getattr(league_player, "age", None),
            str(getattr(attr, "nationality", None) or getattr(league_player, "nationality", None) or ""),
            str(getattr(attr, "club", None) or candidate_row.club_snapshot or ""),
            str(getattr(league_player, "team_name", None) or candidate_row.heigo_club_snapshot or "大海"),
            database_ca,
            current_ca,
            current_pa,
            current_source,
            data_version,
            added_at,
        ])

    widths = [8, 12, 24, 16, 8, 14, 24, 22, 11, 11, 11, 22, 16, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center" if cell.column in {1, 2, 5, 9, 10, 11} else "left", vertical="center")
        for column in (9, 10, 11):
            row[column - 1].font = Font(bold=True, color="0F766E" if column == 9 else "1D4ED8")
            row[column - 1].number_format = "0"
        if row[11].value == "联赛名单":
            row[11].fill = PatternFill("solid", fgColor="DCFCE7")
            row[11].font = Font(color="166534", bold=True)
        elif row[11].value == "球员数据库":
            row[11].fill = PatternFill("solid", fgColor="E0E7FF")
            row[11].font = Font(color="3730A3", bold=True)
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A5:{last_column}{max(5, sheet.max_row)}"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"HEIGO_candidate_list_{record.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


def create_candidate_list(
    db: Session,
    admin: str,
    request: CandidateListUpsertRequest,
) -> CandidateListMutationResponse:
    name = str(request.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="候选名单名称不能为空")
    list_type = request.type if request.type in CANDIDATE_LIST_TYPES else "custom"
    data_version = resolve_attribute_version(db, request.base_data_version)
    now = _now()
    record = CandidateList(
        name=name,
        description=str(request.description or "").strip(),
        type=list_type,
        status="draft",
        base_data_version=data_version,
        source_filters_json=_json_dumps(request.source_filters),
        created_by=admin,
        updated_by=admin,
        created_at=now,
        updated_at=now,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    _audit(db, action="create", operator=admin, summary=f"创建候选名单：{name}", details={"candidate_list_id": record.id})
    return CandidateListMutationResponse(success=True, message="候选名单已创建", list=_serialize_detail(db, record))


def update_candidate_list(
    db: Session,
    admin: str,
    list_id: int,
    request: CandidateListUpsertRequest,
) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    if record.locked_at:
        raise HTTPException(status_code=403, detail="候选名单已锁定，不能编辑")
    name = str(request.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="候选名单名称不能为空")
    record.name = name
    record.description = str(request.description or "").strip()
    record.type = request.type if request.type in CANDIDATE_LIST_TYPES else "custom"
    if request.base_data_version:
        record.base_data_version = resolve_attribute_version(db, request.base_data_version)
    record.source_filters_json = _json_dumps(request.source_filters)
    record.updated_by = admin
    record.updated_at = _now()
    db.commit()
    db.refresh(record)
    _audit(db, action="update", operator=admin, summary=f"更新候选名单：{name}", details={"candidate_list_id": record.id})
    return CandidateListMutationResponse(success=True, message="候选名单已更新", list=_serialize_detail(db, record))


def _candidate_rows_for_tokens(db: Session, tokens: list[str], data_version: str) -> list[Any]:
    uid_tokens = [int(token) for token in tokens if token.isdigit()]
    name_tokens = [token for token in tokens if not token.isdigit()]
    model, base_query = _attribute_base_query(db, data_version)
    rows = []
    if uid_tokens:
        rows.extend(base_query.filter(model.uid.in_(uid_tokens)).all())
    name_filters = []
    for token in name_tokens:
        strict_keys, loose_keys = build_search_normalized_keys(token)
        for key in strict_keys:
            name_filters.append(func.heigo_normalize(model.name).contains(key))
        for key in loose_keys:
            name_filters.append(func.heigo_normalize_loose(model.name).contains(key))
        if not strict_keys and not loose_keys:
            name_filters.append(model.name.ilike(f"%{token}%"))
    if name_filters:
        rows.extend(base_query.filter(or_(*name_filters)).limit(ACTIVE_PLAYER_LIMIT).all())
    deduped = {}
    for row in rows:
        deduped[int(row.uid)] = row
    return list(deduped.values())


def _matches_token(token: str, attr: Any) -> tuple[bool, bool]:
    clean = str(token or "").strip()
    if not clean:
        return False, False
    if clean.isdigit():
        return int(attr.uid) == int(clean), True
    strict_keys, loose_keys = build_search_normalized_keys(clean)
    player_strict, player_loose = build_search_normalized_keys(attr.name)
    exact = any(key and key in player_strict for key in strict_keys) or any(key and key in player_loose for key in loose_keys)
    if exact:
        return True, True
    partial = any(key and any(key in player_key for player_key in player_strict) for key in strict_keys)
    partial = partial or any(key and any(key in player_key for player_key in player_loose) for key in loose_keys)
    return partial, False


def _dedupe_candidate_tokens(tokens: list[str] | None = None, uids: list[int] | None = None, *, limit: int = 500) -> list[str]:
    raw_tokens = [str(uid) for uid in uids or []] + [str(token).strip() for token in tokens or []]
    deduped = []
    for token in raw_tokens:
        if token and token not in deduped:
            deduped.append(token)
    return deduped[:limit]


def preview_candidate_list_players(
    db: Session,
    list_id: int,
    request: CandidateListPlayerPreviewRequest,
) -> CandidateListPlayerPreviewResponse:
    record = _candidate_list_or_404(db, list_id)
    data_version = resolve_attribute_version(db, request.version or record.base_data_version)
    tokens = _dedupe_candidate_tokens(request.tokens, request.uids)
    candidate_rows = _candidate_rows_for_tokens(db, tokens, data_version)
    heigo_players = map_player_uid_to_team_name(db)
    matched = []
    ambiguous = []
    unmatched = []
    already_exists = []
    seen_matched_uids = set()
    for token in tokens:
        token_matches = []
        exact_matches = []
        for row in candidate_rows:
            matched_token, exact = _matches_token(token, row)
            if not matched_token:
                continue
            token_matches.append(row)
            if exact:
                exact_matches.append(row)
        auto_matches = exact_matches if exact_matches else []
        if not auto_matches:
            if token_matches:
                ambiguous.append(
                    CandidateListPreviewTokenResponse(
                        token=token,
                        candidates=[
                            _preview_candidate_from_attr(row, data_version=data_version, heigo_club=heigo_players.get(row.uid, ""))
                            for row in token_matches[:20]
                        ],
                    )
                )
            else:
                unmatched.append(token)
            continue
        if len(auto_matches) > 1:
            ambiguous.append(
                CandidateListPreviewTokenResponse(
                    token=token,
                    candidates=[
                        _preview_candidate_from_attr(row, data_version=data_version, heigo_club=heigo_players.get(row.uid, ""))
                        for row in auto_matches[:20]
                    ],
                )
            )
            continue
        row = auto_matches[0]
        if row.uid in seen_matched_uids:
            continue
        seen_matched_uids.add(row.uid)
        candidate = _preview_candidate_from_attr(row, data_version=data_version, heigo_club=heigo_players.get(row.uid, ""))
        if _active_exists(db, list_id, row.uid, data_version):
            already_exists.append(candidate)
        else:
            matched.append(candidate)
    return CandidateListPlayerPreviewResponse(
        matched=matched,
        ambiguous=ambiguous,
        unmatched=unmatched,
        already_exists=already_exists,
        will_add_count=len(matched),
        data_version=data_version,
    )


def _add_attr_to_list(db: Session, record: CandidateList, attr: Any, *, data_version: str, admin: str) -> bool:
    if _active_exists(db, record.id, attr.uid, data_version):
        return False
    heigo_players = map_player_uid_to_team_name(db)
    row = CandidateListPlayer(
        list_id=record.id,
        uid=int(attr.uid),
        data_version=data_version,
        name_snapshot=str(attr.name or ""),
        club_snapshot=str(attr.club or ""),
        heigo_club_snapshot=heigo_players.get(attr.uid, ""),
        ca_snapshot=int(attr.ca or 0),
        pa_snapshot=int(attr.pa or 0),
        added_by=admin,
        added_at=_now(),
    )
    db.add(row)
    return True


def commit_candidate_list_players(
    db: Session,
    admin: str,
    list_id: int,
    request: CandidateListPlayerCommitRequest,
) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    if record.locked_at:
        raise HTTPException(status_code=403, detail="候选名单已锁定，不能编辑")
    data_version = resolve_attribute_version(db, request.version or record.base_data_version)
    preview = preview_candidate_list_players(
        db,
        list_id,
        CandidateListPlayerPreviewRequest(tokens=request.tokens, uids=request.uids, version=data_version),
    )
    uids_to_add = {item.uid for item in preview.matched}
    uids_to_add.update(int(uid) for uid in request.confirmed_uids or [])
    attrs = _query_attributes_by_uids(db, list(uids_to_add), data_version)
    added_count = 0
    for uid in sorted(uids_to_add):
        attr = attrs.get(int(uid))
        if attr and _add_attr_to_list(db, record, attr, data_version=data_version, admin=admin):
            added_count += 1
    record.updated_by = admin
    record.updated_at = _now()
    db.commit()
    db.refresh(record)
    _audit(
        db,
        action="batch_add",
        operator=admin,
        summary=f"候选名单添加 {added_count} 名球员：{record.name}",
        details={
            "candidate_list_id": record.id,
            "candidate_list_name": record.name,
            "added_count": added_count,
            "requested_uids": sorted(uids_to_add),
            "unmatched": preview.unmatched,
            "ambiguous_count": len(preview.ambiguous),
        },
    )
    return CandidateListMutationResponse(
        success=True,
        message=f"已添加 {added_count} 名球员",
        list=_serialize_detail(db, record),
        preview=preview,
        added_count=added_count,
    )


def preview_candidate_list_player_removals(
    db: Session,
    list_id: int,
    request: CandidateListBatchRemoveRequest,
) -> CandidateListRemovePreviewResponse:
    record = _candidate_list_or_404(db, list_id)
    data_version = resolve_attribute_version(db, request.version or record.base_data_version)
    tokens = _dedupe_candidate_tokens(request.tokens, request.uids)
    active_rows = _active_rows(db, list_id)
    active_by_uid = {int(row.uid): row for row in active_rows}
    candidate_rows = _candidate_rows_for_tokens(db, tokens, data_version)
    candidate_by_uid = {int(row.uid): row for row in candidate_rows}
    heigo_players = map_player_uid_to_team_name(db)
    matched = []
    ambiguous = []
    unmatched = []
    not_in_list = []
    seen_remove_uids = set()
    seen_not_in_list_uids = set()

    def append_matched(uid: int) -> None:
        uid = int(uid)
        if uid in seen_remove_uids or uid not in active_by_uid:
            return
        row = active_by_uid[uid]
        matched.append(
            _player_response_from_row(
                row,
                candidate_by_uid.get(uid),
                view_version=data_version,
                heigo_club=heigo_players.get(uid, row.heigo_club_snapshot or ""),
            )
        )
        seen_remove_uids.add(uid)

    def append_not_in_list(attr: Any) -> None:
        uid = int(attr.uid)
        if uid in seen_not_in_list_uids or uid in active_by_uid:
            return
        not_in_list.append(_preview_candidate_from_attr(attr, data_version=data_version, heigo_club=heigo_players.get(uid, "")))
        seen_not_in_list_uids.add(uid)

    for token in tokens:
        if token.isdigit():
            uid = int(token)
            if uid in active_by_uid:
                append_matched(uid)
            elif uid in candidate_by_uid:
                append_not_in_list(candidate_by_uid[uid])
            else:
                unmatched.append(token)
            continue

        token_matches = []
        exact_matches = []
        for row in candidate_rows:
            matched_token, exact = _matches_token(token, row)
            if not matched_token:
                continue
            token_matches.append(row)
            if exact:
                exact_matches.append(row)

        list_exact_matches = [row for row in exact_matches if int(row.uid) in active_by_uid]
        list_matches = [row for row in token_matches if int(row.uid) in active_by_uid]
        if len(list_exact_matches) == 1:
            append_matched(list_exact_matches[0].uid)
            continue
        if len(list_exact_matches) > 1 or list_matches:
            ambiguous.append(
                CandidateListPreviewTokenResponse(
                    token=token,
                    candidates=[
                        _preview_candidate_from_attr(row, data_version=data_version, heigo_club=heigo_players.get(row.uid, ""))
                        for row in (list_exact_matches or list_matches)[:20]
                    ],
                )
            )
            continue
        if token_matches:
            for row in token_matches[:20]:
                append_not_in_list(row)
        else:
            unmatched.append(token)

    return CandidateListRemovePreviewResponse(
        matched=matched,
        ambiguous=ambiguous,
        unmatched=unmatched,
        not_in_list=not_in_list,
        will_remove_count=len(matched),
        data_version=data_version,
    )


def remove_candidate_list_players(
    db: Session,
    admin: str,
    list_id: int,
    request: CandidateListBatchRemoveRequest,
) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    if record.locked_at:
        raise HTTPException(status_code=403, detail="候选名单已锁定，不能编辑")
    uids = sorted({int(uid) for uid in request.uids or []})
    if not uids:
        raise HTTPException(status_code=400, detail="请提供要移除的 UID")
    query = _active_rows_query(db, list_id).filter(CandidateListPlayer.uid.in_(uids))
    rows = query.all()
    now = _now()
    for row in rows:
        row.removed_at = now
        row.removed_by = admin
    record.updated_by = admin
    record.updated_at = now
    db.commit()
    db.refresh(record)
    _audit(
        db,
        action="batch_remove",
        operator=admin,
        summary=f"候选名单移除 {len(rows)} 名球员：{record.name}",
        details={"candidate_list_id": record.id, "candidate_list_name": record.name, "removed_uids": uids},
    )
    return CandidateListMutationResponse(
        success=True,
        message=f"已移除 {len(rows)} 名球员",
        list=_serialize_detail(db, record),
        removed_count=len(rows),
    )


def publish_preview(db: Session, list_id: int) -> CandidateListPublishPreviewResponse:
    record = _candidate_list_or_404(db, list_id)
    active = _active_rows(db, list_id)
    current_uids = sorted({int(row.uid) for row in active})
    snapshot = _json_loads(record.last_published_snapshot_json)
    previous_uids = sorted({int(uid) for uid in snapshot.get("uids", []) if str(uid).isdigit()})
    view_version = resolve_attribute_version(db, record.base_data_version)
    attrs = _query_attributes_by_uids(db, current_uids, view_version)
    current_set = set(current_uids)
    previous_set = set(previous_uids)
    return CandidateListPublishPreviewResponse(
        list_id=record.id,
        name=record.name,
        previous_count=len(previous_set),
        current_count=len(current_set),
        added_uids=sorted(current_set - previous_set),
        removed_uids=sorted(previous_set - current_set),
        kept_count=len(current_set & previous_set),
        missing_count=len([uid for uid in current_uids if uid not in attrs]),
    )


def publish_candidate_list(db: Session, admin: str, list_id: int) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    if record.locked_at:
        raise HTTPException(status_code=403, detail="候选名单已锁定，不能发布")
    preview = publish_preview(db, list_id)
    now = _now()
    record.status = "published"
    record.published_at = now
    record.published_by = admin
    record.archived_at = None
    record.archived_by = None
    record.published_player_count = preview.current_count
    record.last_published_snapshot_json = _json_dumps(
        {
            "uids": sorted({row.uid for row in _active_rows(db, list_id)}),
            "published_at": now.isoformat(),
            "published_by": admin,
            "base_data_version": record.base_data_version,
        }
    )
    record.updated_by = admin
    record.updated_at = now
    db.commit()
    db.refresh(record)
    _audit(db, action="publish", operator=admin, summary=f"发布候选名单：{record.name}", details=preview.model_dump(mode="json"))
    return CandidateListMutationResponse(success=True, message="候选名单已发布", list=_serialize_detail(db, record))


def unpublish_candidate_list(db: Session, admin: str, list_id: int) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    if record.locked_at:
        raise HTTPException(status_code=403, detail="候选名单已锁定，不能取消发布")
    record.status = "draft"
    record.updated_by = admin
    record.updated_at = _now()
    db.commit()
    db.refresh(record)
    _audit(db, action="unpublish", operator=admin, summary=f"取消发布候选名单：{record.name}", details={"candidate_list_id": record.id})
    return CandidateListMutationResponse(success=True, message="候选名单已取消发布", list=_serialize_detail(db, record))


def archive_candidate_list(db: Session, admin: str, list_id: int) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    record.status = "archived"
    record.archived_at = _now()
    record.archived_by = admin
    record.updated_by = admin
    record.updated_at = record.archived_at
    db.commit()
    db.refresh(record)
    _audit(db, action="archive", operator=admin, summary=f"归档候选名单：{record.name}", details={"candidate_list_id": record.id})
    return CandidateListMutationResponse(success=True, message="候选名单已归档", list=_serialize_detail(db, record))


def duplicate_candidate_list(db: Session, admin: str, list_id: int) -> CandidateListMutationResponse:
    source = _candidate_list_or_404(db, list_id)
    now = _now()
    clone = CandidateList(
        name=f"{source.name} - 副本",
        description=source.description,
        type=source.type,
        status="draft",
        base_data_version=source.base_data_version,
        source_filters_json=source.source_filters_json,
        created_by=admin,
        updated_by=admin,
        created_at=now,
        updated_at=now,
    )
    db.add(clone)
    db.flush()
    for row in _active_rows(db, source.id):
        db.add(
            CandidateListPlayer(
                list_id=clone.id,
                uid=row.uid,
                data_version=row.data_version,
                name_snapshot=row.name_snapshot,
                club_snapshot=row.club_snapshot,
                heigo_club_snapshot=row.heigo_club_snapshot,
                ca_snapshot=row.ca_snapshot,
                pa_snapshot=row.pa_snapshot,
                added_by=admin,
                added_at=now,
            )
        )
    db.commit()
    db.refresh(clone)
    _audit(
        db,
        action="duplicate",
        operator=admin,
        summary=f"复制候选名单：{source.name}",
        details={"source_candidate_list_id": source.id, "candidate_list_id": clone.id},
    )
    return CandidateListMutationResponse(success=True, message="候选名单副本已创建", list=_serialize_detail(db, clone))


def lock_candidate_list(db: Session, admin: str, list_id: int) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    record.locked_at = _now()
    record.locked_by = admin
    record.updated_by = admin
    record.updated_at = record.locked_at
    db.commit()
    db.refresh(record)
    _audit(db, action="lock", operator=admin, summary=f"锁定候选名单：{record.name}", details={"candidate_list_id": record.id})
    return CandidateListMutationResponse(success=True, message="候选名单已锁定", list=_serialize_detail(db, record))


def unlock_candidate_list(db: Session, admin: str, list_id: int) -> CandidateListMutationResponse:
    record = _candidate_list_or_404(db, list_id)
    record.locked_at = None
    record.locked_by = None
    record.updated_by = admin
    record.updated_at = _now()
    db.commit()
    db.refresh(record)
    _audit(db, action="unlock", operator=admin, summary=f"解锁候选名单：{record.name}", details={"candidate_list_id": record.id})
    return CandidateListMutationResponse(success=True, message="候选名单已解锁", list=_serialize_detail(db, record))
