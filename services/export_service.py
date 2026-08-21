import io
import re
from copy import copy
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from imports_runtime.constants import SUPPORTED_INFO_KEY_ALIASES, WORKBOOK_SHEET_LEAGUE_PLAYERS, WORKBOOK_SHEET_OVERVIEW
from imports_runtime.validators import normalize_header
from models import LeagueInfo, Team
from repositories.match_repository import list_matches
from repositories.player_repository import list_league_players
from repositories.team_repository import list_visible_teams
from schemas_read import PlayerExportRow, TeamExportRow
from services.import_service import resolve_import_root
from league_settings import get_league_wage_caps
from services.league_service import REALTIME_TEAM_STAT_SCOPES, collect_team_stat_overlays, get_team_effective_wage_cap
from services import site_note_service, suspension_service

LEVEL_ORDER = {"超级": 1, "甲级": 2, "乙级": 3}
VISIBLE_LEVEL = "隐藏"

SCHEDULE_STATUS_LABELS = {
    "scheduled": "未赛",
    "played": "已赛",
    "postponed": "延期",
    "cancelled": "取消",
    "home_forfeit": "主队判负",
    "away_forfeit": "客队判负",
    "double_forfeit": "双方判负",
}

OVERVIEW_TEAM_FORMULA_LABELS = {
    "球队人数",
    "门将人数",
    "工资",
    "额外工资",
    "税后",
    "最终工资",
    "8M",
    "7M",
    "伪名",
    "总身价",
    "平均身价",
    "平均CA",
    "平均PA",
    "成长总计",
}

OVERVIEW_SUMMARY_FORMULA_KEYS = {
    "总工资",
    "总平均工资",
    "总身价",
    "总平均身价",
    "身价极差",
    "总平均CA",
    "CA极差",
    "总平均PA",
    "PA极差",
    "总平均成长",
}

OVERVIEW_TEAM_FORMULA_HEADERS = {normalize_header(label) for label in OVERVIEW_TEAM_FORMULA_LABELS}


def _overview_range_arguments(column: str, used_rows_by_level: dict[str, list[int]]) -> str:
    ranges = []
    for level in LEVEL_ORDER:
        rows = used_rows_by_level.get(level, [])
        if rows:
            ranges.append(f"{column}{rows[0]}:{column}{rows[-1]}")
    return ",".join(ranges)


def _set_overview_summary_formulas(sheet, used_rows_by_level: dict[str, list[int]]) -> None:
    team_name_ranges = _overview_range_arguments("F", used_rows_by_level)
    wage_ranges = _overview_range_arguments("M", used_rows_by_level)
    value_ranges = _overview_range_arguments("Q", used_rows_by_level)
    avg_value_ranges = _overview_range_arguments("R", used_rows_by_level)
    avg_ca_ranges = _overview_range_arguments("S", used_rows_by_level)
    avg_pa_ranges = _overview_range_arguments("T", used_rows_by_level)
    growth_ranges = _overview_range_arguments("U", used_rows_by_level)
    if not team_name_ranges:
        return
    formulas = {
        "总工资": f"=SUM({wage_ranges})",
        "总平均工资": f"=IFERROR(SUM({wage_ranges})/COUNTA({team_name_ranges}),0)",
        "总身价": f"=SUM({value_ranges})",
        "总平均身价": f"=SUM({avg_value_ranges})",
        "身价极差": f"=MAX({value_ranges})-MIN({value_ranges})",
        "总平均CA": f"=SUM({avg_ca_ranges})",
        "CA极差": f"=MAX({avg_ca_ranges})-MIN({avg_ca_ranges})",
        "总平均PA": f"=SUM({avg_pa_ranges})",
        "PA极差": f"=MAX({avg_pa_ranges})-MIN({avg_pa_ranges})",
        "总平均成长": f"=IFERROR(SUM({growth_ranges})/COUNTA({team_name_ranges}),0)",
    }
    for row_no in range(1, sheet.max_row + 1):
        key = str(sheet.cell(row_no, 1).value or "").strip()
        if key in formulas:
            sheet.cell(row_no, 2).value = formulas[key]


def _get_export_teams(db: Session) -> list[Team]:
    teams = list_visible_teams(db, VISIBLE_LEVEL)
    return sorted(teams, key=lambda team: (LEVEL_ORDER.get(team.level, 99), team.name))


def _get_export_players(db: Session):
    return list_league_players(db)


def _find_latest_roster_template() -> Path | None:
    root = resolve_import_root()
    candidates = [
        path
        for pattern in ("*.xlsx", "*.xlsm")
        for path in root.glob(pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    candidates.sort(key=lambda path: (path.stat().st_mtime, path.stat().st_size, path.name), reverse=True)
    for path in candidates:
        try:
            with pd.ExcelFile(path) as workbook:
                sheet_names = {normalize_header(name) for name in workbook.sheet_names}
            if normalize_header(WORKBOOK_SHEET_OVERVIEW) in sheet_names and normalize_header(WORKBOOK_SHEET_LEAGUE_PLAYERS) in sheet_names:
                return path
        except Exception:
            continue
    return None


def _header_columns(sheet, row_no: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column_no in range(1, sheet.max_column + 1):
        key = normalize_header(sheet.cell(row_no, column_no).value)
        if key and key not in columns:
            columns[key] = column_no
    return columns


def _copy_template_row(sheet, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column_no in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column_no)
        target = sheet.cell(target_row, column_no)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)
        if source.border:
            target.border = copy(source.border)
        if isinstance(source.value, str) and source.value.startswith("="):
            try:
                target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
            except Exception:
                target.value = source.value
        else:
            target.value = source.value


def _mark_workbook_for_full_recalculation(workbook) -> None:
    """Make Excel/WPS discard template caches and rebuild every formula chain."""
    try:
        calculation = workbook.calculation
        calculation.calcMode = "auto"
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcOnSave = True
        calculation.calcCompleted = False
        calculation.fullPrecision = True
        # A copied template can carry a recent calculation-engine id. Some WPS
        # versions then trust stale cached results even when formulas changed.
        calculation.calcId = 0
    except AttributeError:
        pass


def _find_header_column(sheet, row_no: int, labels: tuple[str, ...], *, prefer_last: bool = False) -> int | None:
    normalized_labels = {normalize_header(label) for label in labels}
    matches = [
        column_no
        for column_no in range(1, sheet.max_column + 1)
        if normalize_header(sheet.cell(row_no, column_no).value) in normalized_labels
    ]
    if not matches:
        return None
    return matches[-1] if prefer_last else matches[0]


def _set_player_financial_formulas(sheet, row_no: int, headers: dict[str, int]) -> None:
    required = {
        "age": _find_header_column(sheet, 1, ("年龄",)),
        "initial_ca": _find_header_column(sheet, 1, ("初始CA",)),
        "current_ca": _find_header_column(sheet, 1, ("当前CA",)),
        "pa": _find_header_column(sheet, 1, ("PA",)),
        "position": _find_header_column(sheet, 1, ("位置",)),
        "initial_value": _find_header_column(sheet, 1, ("初始身价",)),
        "current_value": _find_header_column(sheet, 1, ("当前身价",)),
        "potential_value": _find_header_column(sheet, 1, ("潜力身价",)),
        "final_value": _find_header_column(sheet, 1, ("身价",)),
        "coefficient": _find_header_column(sheet, 1, ("系数",)),
        "wage": _find_header_column(sheet, 1, ("工资",)),
        "initial_field": _find_header_column(sheet, 1, ("初始",)),
        "slot": _find_header_column(sheet, 1, ("名额",), prefer_last=True),
    }
    if any(column_no is None for column_no in required.values()):
        return

    refs = {key: f"{get_column_letter(column_no)}{row_no}" for key, column_no in required.items()}
    growth_limit = f"'{WORKBOOK_SHEET_OVERVIEW}'!$B$4"
    initial_value_formula = f'=IF({refs["initial_ca"]}<115,1,INT(({refs["initial_ca"]}-95)/10))'
    current_value_formula = f'=IF({refs["current_ca"]}<115,1,INT(({refs["current_ca"]}-95)/10))'
    potential_value_formula = f'=IF({refs["pa"]}<115,1,INT(({refs["pa"]}-95)/10))'
    final_value_formula = (
        f'=IF(OR({refs["age"]}>{growth_limit},{refs["pa"]}<140),'
        f'{refs["current_value"]},({refs["current_value"]}+{refs["potential_value"]})/2)'
    )
    initial_field_formula = (
        f'=IF({refs["age"]}>25,{refs["initial_value"]},'
        f'({refs["initial_value"]}+{refs["potential_value"]})/2)'
    )
    slot_formula = (
        f'=IF({refs["initial_field"]}>=8,"8M",IF({refs["initial_field"]}>=7,"7M",'
        f'IF(OR(AND({refs["age"]}<={growth_limit},{refs["pa"]}>164),'
        f'AND({refs["age"]}>{growth_limit},{refs["age"]}<26,{refs["current_ca"]}>164)),"伪名","")))'
    )
    coefficient_formula = (
        f'=IF(AND(UPPER(TRIM({refs["position"]}))="GK",{refs["slot"]}<>""),0.1,'
        f'IF({refs["final_value"]}=1,0.1,IF({refs["initial_field"]}=1,0.1,'
        f'IF(({refs["initial_field"]}+{refs["current_value"]})/2=1,0.1,'
        f'IF({refs["slot"]}="8M",0.15,IF({refs["slot"]}="7M",0.13,'
        f'IF({refs["slot"]}="伪名",0.11,IF({refs["age"]}>{growth_limit},'
        f'IF(INT(({refs["current_ca"]}-95)/10)>5,0.09,0.07),'
        f'IF(INT(({refs["pa"]}-95)/10)>5,0.09,0.07)))))))))'
    )

    sheet.cell(row_no, required["initial_value"]).value = initial_value_formula
    sheet.cell(row_no, required["current_value"]).value = current_value_formula
    sheet.cell(row_no, required["potential_value"]).value = potential_value_formula
    sheet.cell(row_no, required["final_value"]).value = final_value_formula
    sheet.cell(row_no, required["initial_field"]).value = initial_field_formula
    sheet.cell(row_no, required["slot"]).value = slot_formula
    sheet.cell(row_no, required["coefficient"]).value = coefficient_formula
    sheet.cell(row_no, required["wage"]).value = f'=ROUND({refs["final_value"]}*{refs["coefficient"]},3)'

    first_slot_column = _find_header_column(sheet, 1, ("名额",))
    if first_slot_column and first_slot_column != required["slot"]:
        sheet.cell(row_no, first_slot_column).value = f'={refs["slot"]}'


def _set_overview_team_formulas(workbook, sheet, row_no: int, level: str, headers: dict[str, int], extra_wage: float) -> None:
    player_sheet = workbook[WORKBOOK_SHEET_LEAGUE_PLAYERS]
    player_team_column = _find_header_column(player_sheet, 1, ("联赛球队", "俱乐部", "更新俱乐部"))
    player_position_column = _find_header_column(player_sheet, 1, ("位置",))
    player_wage_column = _find_header_column(player_sheet, 1, ("工资",))
    player_slot_column = _find_header_column(player_sheet, 1, ("名额",), prefer_last=True)
    player_value_column = _find_header_column(player_sheet, 1, ("身价",))
    player_ca_column = _find_header_column(player_sheet, 1, ("当前CA",))
    player_pa_column = _find_header_column(player_sheet, 1, ("PA",))
    player_initial_ca_column = _find_header_column(player_sheet, 1, ("初始CA",))
    required_player_columns = (
        player_team_column,
        player_position_column,
        player_wage_column,
        player_slot_column,
        player_value_column,
        player_ca_column,
        player_pa_column,
        player_initial_ca_column,
    )
    if any(column_no is None for column_no in required_player_columns):
        return

    def overview_ref(label: str) -> str | None:
        column_no = headers.get(normalize_header(label))
        return f"{get_column_letter(column_no)}{row_no}" if column_no else None

    refs = {label: overview_ref(label) for label in OVERVIEW_TEAM_FORMULA_LABELS | {"球队名", "工资帽"}}
    if any(refs.get(label) is None for label in OVERVIEW_TEAM_FORMULA_LABELS | {"球队名", "工资帽"}):
        return

    team_column = get_column_letter(player_team_column)
    position_column = get_column_letter(player_position_column)
    wage_column = get_column_letter(player_wage_column)
    slot_column = get_column_letter(player_slot_column)
    value_column = get_column_letter(player_value_column)
    ca_column = get_column_letter(player_ca_column)
    pa_column = get_column_letter(player_pa_column)
    initial_ca_column = get_column_letter(player_initial_ca_column)
    player_sheet_ref = f"'{WORKBOOK_SHEET_LEAGUE_PLAYERS}'"
    team_ref = refs["球队名"]
    min_wage = {"超级": 8.0, "甲级": 7.5, "乙级": 6.5}.get(level, 8.0)

    formulas = {
        "球队人数": f'=COUNTIF({player_sheet_ref}!{team_column}:{team_column},{team_ref})',
        "门将人数": f'=COUNTIFS({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{position_column}:{position_column},"*GK*")',
        "工资": f'=ROUND(SUMIF({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{wage_column}:{wage_column}),3)',
        "额外工资": f'=ROUND({float(extra_wage or 0):g},3)',
        "8M": f'=COUNTIFS({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{slot_column}:{slot_column},"8M")',
        "7M": f'=COUNTIFS({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{slot_column}:{slot_column},"7M")',
        "伪名": f'=COUNTIFS({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{slot_column}:{slot_column},"伪名")',
        "总身价": f'=SUMIF({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{value_column}:{value_column})',
        "平均身价": f'=IFERROR({refs["总身价"]}/{refs["球队人数"]},0)',
        "平均CA": f'=IFERROR(SUMIF({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{ca_column}:{ca_column})/{refs["球队人数"]},0)',
        "平均PA": f'=IFERROR(SUMIF({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{pa_column}:{pa_column})/{refs["球队人数"]},0)',
        "成长总计": (
            f'=SUMIF({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{ca_column}:{ca_column})-'
            f'SUMIF({player_sheet_ref}!{team_column}:{team_column},{team_ref},{player_sheet_ref}!{initial_ca_column}:{initial_ca_column})'
        ),
    }
    total_wage = f'ROUND({refs["工资"]}+{refs["额外工资"]},3)'
    overflow = f'ROUND({total_wage}-{refs["工资帽"]},3)'
    formulas["税后"] = (
        f'=IF({overflow}>0.3,"拍卖",IF({total_wage}>{refs["工资帽"]},'
        f'MAX({total_wage},ROUND({overflow}*10+{refs["工资"]},3)),{total_wage}))'
    )
    formulas["最终工资"] = f'=IF({refs["税后"]}="拍卖","拍卖",MAX({min_wage:g},{refs["税后"]}))'

    for label, formula in formulas.items():
        sheet[refs[label]].value = formula


def _update_template_overview(workbook, db: Session, export_teams: list[Team], realtime_overlays: dict) -> None:
    sheet = workbook[WORKBOOK_SHEET_OVERVIEW]
    headers = _header_columns(sheet, 2)
    wage_cap_header = normalize_header("工资帽")
    if wage_cap_header not in headers:
        wage_cap_column = sheet.max_column + 1
        source_header_column = headers.get(normalize_header("备注"), sheet.max_column)
        source_header = sheet.cell(2, source_header_column)
        target_header = sheet.cell(2, wage_cap_column, "工资帽")
        if source_header.has_style:
            target_header._style = copy(source_header._style)
        target_header.font = copy(source_header.font)
        target_header.fill = copy(source_header.fill)
        target_header.border = copy(source_header.border)
        target_header.alignment = copy(source_header.alignment)
        target_header.number_format = source_header.number_format
        sheet.column_dimensions[get_column_letter(wage_cap_column)].width = 12
        headers = _header_columns(sheet, 2)
    team_name_column = headers.get(normalize_header("球队名"))
    level_column = headers.get(normalize_header("级别"))
    if not team_name_column or not level_column:
        raise ValueError("导入名单模板的信息总览缺少球队名列")

    wage_caps = get_league_wage_caps(db)
    team_rows = []
    for index, team in enumerate(export_teams, start=1):
        overlay = realtime_overlays.get(team.id, {})
        team_rows.append({
            "序号": index,
            "级别": team.level,
            "球队名": team.name,
            "主教": team.manager or "",
            "球队人数": team.team_size,
            "门将人数": team.gk_count,
            "工资": team.wage,
            "额外工资": team.extra_wage or 0,
            "工资帽": get_team_effective_wage_cap(team, wage_caps),
            "税后": team.after_tax or 0,
            "最终工资": team.final_wage,
            "8M": team.count_8m,
            "7M": team.count_7m,
            "伪名": team.count_fake,
            "总身价": overlay.get("total_value", team.total_value),
            "平均身价": overlay.get("avg_value", team.avg_value),
            "平均CA": overlay.get("avg_ca", team.avg_ca),
            "平均PA": overlay.get("avg_pa", team.avg_pa),
            "成长总计": overlay.get("total_growth", team.total_growth),
            "备注": team.notes or "",
        })

    slot_rows_by_level = {
        level: [
            row_no
            for row_no in range(3, sheet.max_row + 1)
            if str(sheet.cell(row_no, level_column).value or "").strip() == level
        ]
        for level in LEVEL_ORDER
    }
    teams_by_level = {
        level: [values for values in team_rows if values["级别"] == level]
        for level in LEVEL_ORDER
    }
    for level, values in teams_by_level.items():
        if len(values) > len(slot_rows_by_level[level]):
            raise ValueError(f"导入名单模板的{level}球队槽位不足：需要 {len(values)}，模板只有 {len(slot_rows_by_level[level])}")

    linked_team_name_column = None
    for level_rows in slot_rows_by_level.values():
        for row_no in level_rows:
            formula = sheet.cell(row_no, team_name_column).value
            match = re.fullmatch(r"=([A-Z]+)(\d+)", str(formula or "").strip(), flags=re.IGNORECASE)
            if match and int(match.group(2)) == row_no:
                linked_team_name_column = column_index_from_string(match.group(1).upper())
                break
        if linked_team_name_column:
            break

    for level, slot_rows in slot_rows_by_level.items():
        for row_no in slot_rows:
            if linked_team_name_column:
                sheet.cell(row_no, linked_team_name_column).value = None
            for normalized_label, column_no in headers.items():
                if column_no >= 4:
                    cell = sheet.cell(row_no, column_no)
                    if (
                        normalized_label in OVERVIEW_TEAM_FORMULA_HEADERS
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        continue
                    cell.value = None

    for level in LEVEL_ORDER:
        for level_index, (row_offset, values) in enumerate(zip(slot_rows_by_level[level], teams_by_level[level]), start=1):
            if linked_team_name_column:
                sheet.cell(row_offset, linked_team_name_column).value = values["球队名"]
            for label, value in values.items():
                column_no = headers.get(normalize_header(label))
                if column_no:
                    cell = sheet.cell(row_offset, column_no)
                    if label in OVERVIEW_TEAM_FORMULA_LABELS:
                        if label == "额外工资":
                            cell.value = f"={float(value or 0):g}"
                        elif isinstance(cell.value, str) and cell.value.startswith("="):
                            continue
                        else:
                            cell.value = value
                    else:
                        cell.value = level_index if label == "序号" else value

            _set_overview_team_formulas(
                workbook,
                sheet,
                row_offset,
                values["级别"],
                headers,
                values["额外工资"],
            )

    used_rows_by_level = {
        level: slot_rows_by_level[level][:len(teams_by_level[level])]
        for level in LEVEL_ORDER
    }
    _set_overview_summary_formulas(sheet, used_rows_by_level)

    league_info = {record.key: record.value for record in db.query(LeagueInfo).all()}
    for row_no in range(1, sheet.max_row + 1):
        raw_key = sheet.cell(row_no, 1).value
        normalized_key = SUPPORTED_INFO_KEY_ALIASES.get(str(raw_key or "").strip(), str(raw_key or "").strip())
        if normalized_key in league_info:
            value_cell = sheet.cell(row_no, 2)
            if normalized_key in OVERVIEW_SUMMARY_FORMULA_KEYS and isinstance(value_cell.value, str) and value_cell.value.startswith("="):
                continue
            value_cell.value = league_info[normalized_key]


def _parse_uid(value) -> int | None:
    try:
        if value in (None, ""):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _update_template_players(workbook, players) -> None:
    sheet = workbook[WORKBOOK_SHEET_LEAGUE_PLAYERS]
    headers = _header_columns(sheet, 1)
    uid_column = headers.get(normalize_header("编号"))
    if not uid_column:
        raise ValueError("导入名单模板的联赛名单缺少编号列")

    existing_rows: dict[int, int] = {}
    for row_no in range(2, sheet.max_row + 1):
        uid = _parse_uid(sheet.cell(row_no, uid_column).value)
        if uid is not None:
            existing_rows[uid] = row_no

    current_uids = {int(player.uid) for player in players}
    for uid, row_no in existing_rows.items():
        if uid in current_uids:
            continue
        for column_no in range(1, sheet.max_column + 1):
            sheet.cell(row_no, column_no).value = None

    template_row = max(existing_rows.values(), default=2)
    next_row = max(sheet.max_row, template_row) + 1
    field_columns = {
        "编号": "uid",
        "姓名": "name",
        "年龄": "age",
        "初始CA": "initial_ca",
        "当前CA": "ca",
        "PA": "pa",
        "位置": "position",
        "国籍": "nationality",
    }

    for player in players:
        row_no = existing_rows.get(int(player.uid))
        if row_no is None:
            row_no = next_row
            next_row += 1
            _copy_template_row(sheet, template_row, row_no)
        for label, attribute in field_columns.items():
            column_no = headers.get(normalize_header(label))
            if column_no:
                sheet.cell(row_no, column_no).value = getattr(player, attribute)
        for label in ("联赛球队", "俱乐部", "更新俱乐部"):
            column_no = headers.get(normalize_header(label))
            if column_no:
                sheet.cell(row_no, column_no).value = player.team_name
        _set_player_financial_formulas(sheet, row_no, headers)


def _build_template_export(db: Session, template_path: Path, export_teams: list[Team], players, realtime_overlays: dict):
    keep_vba = template_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(template_path, data_only=False, keep_vba=keep_vba, keep_links=True)
    _update_template_overview(workbook, db, export_teams, realtime_overlays)
    _update_template_players(workbook, players)
    _mark_workbook_for_full_recalculation(workbook)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    suffix = ".xlsm" if keep_vba else ".xlsx"
    filename = f"heigo_roster_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}{suffix}"
    return output, filename


def _build_standard_export(db: Session, export_teams: list[Team], players, realtime_overlays: dict):
    output = io.BytesIO()
    wage_caps = get_league_wage_caps(db)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        team_rows = [
            TeamExportRow(
                level=team.level,
                team_name=team.name,
                manager=team.manager or "",
                team_size=team.team_size,
                gk_count=team.gk_count,
                extra_wage=team.extra_wage or 0,
                wage_cap=get_team_effective_wage_cap(team, wage_caps),
                after_tax=team.after_tax or 0,
                final_wage=team.final_wage,
                count_8m=team.count_8m,
                count_7m=team.count_7m,
                count_fake=team.count_fake,
                total_value=realtime_overlays.get(team.id, {}).get("total_value", team.total_value),
                avg_value=realtime_overlays.get(team.id, {}).get("avg_value", team.avg_value),
                avg_ca=realtime_overlays.get(team.id, {}).get("avg_ca", team.avg_ca),
                avg_pa=realtime_overlays.get(team.id, {}).get("avg_pa", team.avg_pa),
                total_growth=realtime_overlays.get(team.id, {}).get("total_growth", team.total_growth),
                notes=team.notes or "",
            ).model_dump(by_alias=True)
            for team in export_teams
        ]
        pd.DataFrame(team_rows).to_excel(writer, sheet_name="信息总览", index=False, startrow=1)

        player_rows = [
            PlayerExportRow(
                uid=player.uid,
                name=player.name,
                age=player.age,
                initial_ca=player.initial_ca,
                ca=player.ca,
                pa=player.pa,
                position=player.position,
                nationality=player.nationality,
                team_name=player.team_name,
                wage=player.wage,
                slot_type=player.slot_type or "",
            ).model_dump(by_alias=True)
            for player in players
        ]
        pd.DataFrame(player_rows).to_excel(writer, sheet_name="联赛名单", index=False)

    output.seek(0)
    filename = f"heigo_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


def build_export_excel(db: Session):
    export_teams = _get_export_teams(db)
    players = _get_export_players(db)
    realtime_overlays = collect_team_stat_overlays(db, export_teams, stat_scopes=REALTIME_TEAM_STAT_SCOPES)
    template_path = _find_latest_roster_template()
    if template_path is not None:
        return _build_template_export(db, template_path, export_teams, players, realtime_overlays)
    return _build_standard_export(db, export_teams, players, realtime_overlays)


def build_suspensions_excel(db: Session, level: str):
    clean_level = str(level or "").strip()
    if clean_level not in LEVEL_ORDER:
        raise ValueError("伤停导出仅支持超级、甲级、乙级")

    suspension_payload = suspension_service.get_suspensions(db)
    notes_by_key = {item.key: item.text for item in site_note_service.list_site_notes(db)}
    level_note = notes_by_key.get(site_note_service.build_suspension_note_key(clean_level), "")
    teams = [team for team in suspension_payload.teams if team.level == clean_level]
    summary_rows = []
    detail_rows = []

    for team in teams:
        team_note = "" if team.is_orphaned else notes_by_key.get(
            site_note_service.build_suspension_team_note_key(team.team_id),
            "",
        )
        summary_rows.append(
            {
                "级别": team.level,
                "球队": team.team_name,
                "主教练": team.manager or "",
                "1张黄牌人数": len(team.one_yellow),
                "2张黄牌人数": len(team.two_yellows),
                "停赛人数": len(team.suspended),
                "球队更新备注": team_note,
                "级别更新说明": level_note,
            }
        )
        groups = (
            ("1张黄牌", team.one_yellow),
            ("2张黄牌", team.two_yellows),
            ("停赛", team.suspended),
        )
        for status_label, records in groups:
            for record in records:
                detail_rows.append(
                    {
                        "级别": record.level,
                        "球队": team.team_name,
                        "主教练": team.manager or "",
                        "状态分类": status_label,
                        "球员UID": record.player_uid,
                        "球员": record.player_name,
                        "额外黄牌数": record.yellow_cards,
                        "3黄停赛": "是" if record.yellow_card_suspended else "否",
                        "红牌停赛": "是" if record.red_card_suspended else "否",
                        "红伤停赛": "是" if record.red_injury_suspended else "否",
                        "球员备注": record.notes or "",
                        "球队更新备注": team_note,
                        "更新时间": record.updated_at.strftime("%Y-%m-%d %H:%M") if record.updated_at else "",
                    }
                )

    summary_columns = ["级别", "球队", "主教练", "1张黄牌人数", "2张黄牌人数", "停赛人数", "球队更新备注", "级别更新说明"]
    detail_columns = ["级别", "球队", "主教练", "状态分类", "球员UID", "球员", "额外黄牌数", "3黄停赛", "红牌停赛", "红伤停赛", "球员备注", "球队更新备注", "更新时间"]
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows, columns=summary_columns).to_excel(writer, sheet_name="球队汇总", index=False)
        pd.DataFrame(detail_rows, columns=detail_columns).to_excel(writer, sheet_name="伤停明细", index=False)
        for worksheet in writer.sheets.values():
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)

    output.seek(0)
    filename = f"HEIGO_suspensions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename


def _standings_sumproduct(team_cell: str, expression: str, last_schedule_row: int) -> str:
    home = f"'赛程'!$B$4:$B${last_schedule_row}"
    away = f"'赛程'!$C$4:$C${last_schedule_row}"
    home_score = f"'赛程'!$D$4:$D${last_schedule_row}"
    away_score = f"'赛程'!$E$4:$E${last_schedule_row}"
    status = f"'赛程'!$F$4:$F${last_schedule_row}"
    active = (
        f"--ISNUMBER({home_score}),--ISNUMBER({away_score}),"
        f"--({status}<>\"延期\"),--({status}<>\"取消\")"
    )
    return expression.format(
        team=team_cell,
        home=home,
        away=away,
        hs=home_score,
        aws=away_score,
        status=status,
        active=active,
    )


def build_standings_excel(db: Session, level: str):
    clean_level = str(level or "").strip()
    if clean_level not in LEVEL_ORDER:
        raise ValueError("积分榜导出仅支持超级、甲级、乙级")

    teams = sorted(
        (team for team in _get_export_teams(db) if team.level == clean_level),
        key=lambda team: team.name,
    )
    matches = list_matches(db, level=clean_level)
    workbook = Workbook()
    standings_sheet = workbook.active
    standings_sheet.title = "主积分榜"
    schedule_sheet = workbook.create_sheet("赛程")

    surface = "FFFFFF"
    surface_muted = "F8F9FA"
    surface_soft = "F4F6FA"
    text_primary = "2C3E50"
    accent_green = "27AE60"
    accent_orange = "E67E22"
    accent = {"超级": "9C82FF", "甲级": "55AEFF", "乙级": "45CF9A"}[clean_level]
    accent_deep = {"超级": "6B4FE0", "甲级": "2779D7", "乙级": "18875F"}[clean_level]
    pale_accent = {"超级": "F2EEFF", "甲级": "EEF7FF", "乙级": "EDFAF5"}[clean_level]
    editable_fill = "EAF7EF"
    pale_blue = "EEF7FF"
    pale_red = "FDEEEE"
    promotion_fill = "EAF7EF"
    relegation_fill = "FDEEEE"
    white = "FFFFFF"
    muted = "5A6C7D"
    thin_side = Side(style="thin", color="DDE3EA")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    accent_bottom_border = Border(bottom=Side(style="medium", color=accent_deep))

    standings_sheet.sheet_view.showGridLines = False
    schedule_sheet.sheet_view.showGridLines = False
    standings_sheet.sheet_properties.tabColor = accent
    schedule_sheet.sheet_properties.tabColor = accent_green

    # Reserve enough formula range for later score entry without adding a third helper sheet.
    last_schedule_row = max(1003, len(matches) + 103)
    standings_sheet.merge_cells("A1:K1")
    standings_sheet["A1"] = f"HEIGO {clean_level}联赛主积分榜"
    standings_sheet["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color=accent_deep)
    standings_sheet["A1"].fill = PatternFill("solid", fgColor=pale_accent)
    standings_sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    standings_sheet["A1"].border = accent_bottom_border
    standings_sheet.row_dimensions[1].height = 34
    standings_sheet.merge_cells("A2:K2")
    standings_sheet["A2"] = "修改“赛程”中的比分和状态后，本表会自动计算；如需查看最新顺序，请按“排名”列升序排序。"
    standings_sheet["A2"].font = Font(size=10, color=muted)
    standings_sheet["A2"].fill = PatternFill("solid", fgColor=surface)
    standings_sheet["A2"].alignment = Alignment(vertical="center")
    standings_sheet["A2"].border = Border(bottom=thin_side)

    standing_headers = ["排名", "球队", "主教练", "场次", "胜", "平", "负", "进球", "失球", "净胜球", "积分"]
    for column, header in enumerate(standing_headers, start=1):
        cell = standings_sheet.cell(row=3, column=column, value=header)
        cell.font = Font(name="Microsoft YaHei", bold=True, color=text_primary)
        cell.fill = PatternFill("solid", fgColor=surface_muted)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(style="medium", color=accent_deep))

    for index, team in enumerate(teams, start=4):
        standings_sheet.cell(index, 2, team.name)
        standings_sheet.cell(index, 3, team.manager or "")
        team_cell = f"$B{index}"
        standings_sheet.cell(index, 4, _standings_sumproduct(
            team_cell,
            "=SUMPRODUCT(--((({home}={team})+({away}={team}))>0),{active})",
            last_schedule_row,
        ))
        normal_status = (
            "({status}<>\"主队判负\")*({status}<>\"客队判负\")*"
            "({status}<>\"双方判负\")"
        )
        standings_sheet.cell(index, 5, _standings_sumproduct(
            team_cell,
            "=SUMPRODUCT(--({home}={team}),{active},--((({status}=\"客队判负\")+("
            + normal_status + "*({hs}>{aws})))>0))+SUMPRODUCT(--({away}={team}),{active},"
            "--((" + normal_status + "*({aws}>{hs}))>0))",
            last_schedule_row,
        ))
        standings_sheet.cell(index, 6, _standings_sumproduct(
            team_cell,
            "=SUMPRODUCT(--((({home}={team})+({away}={team}))>0),{active},"
            "--((({status}=\"主队判负\")+(" + normal_status + "*({hs}={aws})))>0))",
            last_schedule_row,
        ))
        standings_sheet.cell(index, 7, f"=D{index}-E{index}-F{index}")
        standings_sheet.cell(index, 8, _standings_sumproduct(
            team_cell,
            "=SUMPRODUCT(--({home}={team}),{active},{hs})+SUMPRODUCT(--({away}={team}),{active},{aws})",
            last_schedule_row,
        ))
        standings_sheet.cell(index, 9, _standings_sumproduct(
            team_cell,
            "=SUMPRODUCT(--({home}={team}),{active},{aws})+SUMPRODUCT(--({away}={team}),{active},{hs})",
            last_schedule_row,
        ))
        standings_sheet.cell(index, 10, f"=H{index}-I{index}")
        standings_sheet.cell(index, 11, f"=E{index}*3+F{index}")

    first_team_row = 4
    last_team_row = max(first_team_row, len(teams) + 3)
    for row in range(first_team_row, len(teams) + 4):
        standings_sheet.cell(row, 1, (
            f'=1+COUNTIF($K${first_team_row}:$K${last_team_row},">"&K{row})'
            f'+COUNTIFS($K${first_team_row}:$K${last_team_row},K{row},$J${first_team_row}:$J${last_team_row},">"&J{row})'
            f'+COUNTIFS($K${first_team_row}:$K${last_team_row},K{row},$J${first_team_row}:$J${last_team_row},J{row},$H${first_team_row}:$H${last_team_row},">"&H{row})'
            f'+COUNTIFS($K${first_team_row}:$K${last_team_row},K{row},$J${first_team_row}:$J${last_team_row},J{row},$H${first_team_row}:$H${last_team_row},H{row},$E${first_team_row}:$E${last_team_row},">"&E{row})'
            f'+COUNTIFS($K${first_team_row}:$K${last_team_row},K{row},$J${first_team_row}:$J${last_team_row},J{row},$H${first_team_row}:$H${last_team_row},H{row},$E${first_team_row}:$E${last_team_row},E{row},$B${first_team_row}:$B${last_team_row},"<"&B{row})'
        ))
        for column in range(1, 12):
            cell = standings_sheet.cell(row, column)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if column in (2, 3) else "center", vertical="center")
            cell.font = Font(name="Microsoft YaHei", size=10, color=text_primary)
            cell.fill = PatternFill("solid", fgColor=surface_muted if row % 2 == 0 else surface)
        standings_sheet.cell(row, 1).font = Font(name="Microsoft YaHei", bold=True, color=accent_deep)
        standings_sheet.cell(row, 1).fill = PatternFill("solid", fgColor=pale_accent)
        standings_sheet.cell(row, 2).font = Font(name="Microsoft YaHei", bold=True, color=accent_green)
        standings_sheet.cell(row, 11).font = Font(name="Microsoft YaHei", bold=True, color=accent_orange)

    standings_sheet.freeze_panes = "D4"
    if teams:
        standings_sheet.auto_filter.ref = f"A3:K{last_team_row}"
        standings_sheet.conditional_formatting.add(
            f"A4:K{last_team_row}",
            FormulaRule(formula=["$A4<=5"], fill=PatternFill("solid", fgColor=promotion_fill), stopIfTrue=True),
        )
        standings_sheet.conditional_formatting.add(
            f"A4:K{last_team_row}",
            FormulaRule(
                formula=[f"$A4>COUNTA($B$4:$B${last_team_row})-5"],
                fill=PatternFill("solid", fgColor=relegation_fill),
                stopIfTrue=True,
            ),
        )
    standings_widths = [9, 24, 18, 9, 8, 8, 8, 9, 9, 10, 10]
    for column, width in enumerate(standings_widths, start=1):
        standings_sheet.column_dimensions[get_column_letter(column)].width = width

    schedule_sheet.merge_cells("A1:H1")
    schedule_sheet["A1"] = f"HEIGO {clean_level}联赛赛程与赛果"
    schedule_sheet["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color=accent_deep)
    schedule_sheet["A1"].fill = PatternFill("solid", fgColor=pale_accent)
    schedule_sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    schedule_sheet["A1"].border = accent_bottom_border
    schedule_sheet.row_dimensions[1].height = 34
    schedule_sheet.merge_cells("A2:H2")
    schedule_sheet["A2"] = "黄色区域可编辑：填写双方比分即可计入积分榜；延期、取消或判负时，请同时选择对应状态。"
    schedule_sheet["A2"].font = Font(size=10, color=muted)
    schedule_sheet["A2"].fill = PatternFill("solid", fgColor=editable_fill)
    schedule_sheet["A2"].alignment = Alignment(vertical="center")
    schedule_sheet["A2"].border = Border(bottom=thin_side)
    schedule_headers = ["轮次", "主队", "客队", "主队比分", "客队比分", "状态", "比赛日期", "备注"]
    for column, header in enumerate(schedule_headers, start=1):
        cell = schedule_sheet.cell(row=3, column=column, value=header)
        cell.font = Font(name="Microsoft YaHei", bold=True, color=text_primary)
        cell.fill = PatternFill("solid", fgColor=surface_muted)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(style="medium", color=accent_deep))

    for row, match in enumerate(matches, start=4):
        values = [
            match.round_no,
            match.home_team_name,
            match.away_team_name,
            match.home_score,
            match.away_score,
            SCHEDULE_STATUS_LABELS.get(match.status, match.status or "未赛"),
            match.match_date,
            match.notes or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = schedule_sheet.cell(row=row, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if column in (2, 3, 8) else "center", vertical="center")
            cell.font = Font(name="Microsoft YaHei", size=10, color=text_primary)
            if column in (4, 5, 6):
                cell.fill = PatternFill("solid", fgColor=editable_fill)
            elif row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=surface_muted)
            else:
                cell.fill = PatternFill("solid", fgColor=surface)
        if match.match_date:
            schedule_sheet.cell(row, 7).number_format = "yyyy-mm-dd hh:mm"

    validation_last_row = max(len(matches) + 3, 203)
    status_validation = DataValidation(
        type="list",
        formula1='"未赛,已赛,延期,取消,主队判负,客队判负,双方判负"',
        allow_blank=True,
    )
    status_validation.promptTitle = "选择比赛状态"
    status_validation.prompt = "正常比赛填写比分即可；特殊情况请选择延期、取消或判负状态。"
    status_validation.errorTitle = "状态不受支持"
    status_validation.error = "请从下拉列表选择比赛状态。"
    status_validation.errorStyle = "stop"
    status_validation.showErrorMessage = True
    status_validation.showInputMessage = True
    schedule_sheet.add_data_validation(status_validation)
    status_validation.add(f"F4:F{validation_last_row}")
    for row in range(4, validation_last_row + 1):
        for column in (4, 5, 6):
            schedule_sheet.cell(row, column).fill = PatternFill("solid", fgColor=editable_fill)
            schedule_sheet.cell(row, column).protection = Protection(locked=False)

    if matches:
        schedule_sheet.auto_filter.ref = f"A3:H{len(matches) + 3}"
    schedule_sheet.freeze_panes = "D4"
    schedule_widths = [9, 25, 25, 12, 12, 15, 20, 30]
    for column, width in enumerate(schedule_widths, start=1):
        schedule_sheet.column_dimensions[get_column_letter(column)].width = width
    schedule_sheet.conditional_formatting.add(
        f"F4:F{validation_last_row}",
        CellIsRule(operator="equal", formula=['"延期"'], fill=PatternFill("solid", fgColor=pale_blue)),
    )
    schedule_sheet.conditional_formatting.add(
        f"F4:F{validation_last_row}",
        CellIsRule(operator="equal", formula=['"取消"'], fill=PatternFill("solid", fgColor=pale_red)),
    )

    _mark_workbook_for_full_recalculation(workbook)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    level_code = {"超级": "S", "甲级": "A", "乙级": "B"}[clean_level]
    filename = f"HEIGO_{level_code}_standings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output, filename
